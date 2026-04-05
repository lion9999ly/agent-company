"""
@description: 飞书长连接客户端 - 使用官方SDK，非研发任务走GPT-5.4直连
@dependencies: lark-oapi, requests, PIL, python-dotenv
@last_modified: 2026-03-20

使用方法:
    python scripts/feishu_sdk_client.py
"""

# 加载 .env 环境变量（必须在其他 import 之前）
from dotenv import load_dotenv
load_dotenv()

import os
import sys

# 强制刷新日志（解决日志不可见问题）
try:
    sys.stdout.reconfigure(line_buffering=True)
except Exception:
    import functools
    print = functools.partial(print, flush=True)

import json
import uuid
import time
import re
import requests
import tempfile
import threading
from pathlib import Path
from io import BytesIO
from datetime import datetime
from PIL import Image

sys.path.insert(0, str(Path(__file__).parent.parent))

# 文件日志函数（解决 cmd.exe 不显示 print 的问题）
_LOG_FILE = Path(__file__).parent.parent / ".ai-state" / "feishu_debug.log"

def log(msg: str):
    """写入文件日志"""
    try:
        with open(_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
    except Exception:
        pass
    print(msg, flush=True)  # 同时尝试 print

from feishu_bridge.ocr_middleware import process_image_to_text
from src.graph.router import app as langgraph_app
from src.utils.model_gateway import get_model_gateway
from src.tools.fix_executor import get_pending_proposals, approve_and_execute, reject_proposal, format_proposal_for_feishu

try:
    import lark_oapi as lark
except ImportError:
    print("请安装: pip install lark-oapi")
    sys.exit(1)


# 配置 - 从环境变量读取，无默认值
APP_ID = os.getenv("FEISHU_APP_ID", "")
APP_SECRET = os.getenv("FEISHU_APP_SECRET", "")
PROCESSED_IDS_FILE = Path(__file__).parent.parent / ".ai-state" / "processed_message_ids.json"

# 全局客户端
client = None

# 消息去重集合
processed_ids = set()

# 记录最近一次研发任务的 ID 和经验卡片路径，用于关联用户评价
_last_task_memory = {"task_id": None, "memory_dir": None}

# 研发任务并发锁，同一时间只允许一个研发任务执行

# 回复上下文：群聊时用 chat_id，私聊时用 open_id
_reply_context = {"target": None, "type": "open_id"}
_rd_task_running = False

# 长任务标志（JDM学习、深度研究等）
_long_task_running = False

# 待执行命令缓存（白天拒绝的重任务，等待夜间或强制执行）
_pending_commands = {}


def load_processed_ids() -> set:
    """从磁盘加载已处理的消息ID"""
    if PROCESSED_IDS_FILE.exists():
        try:
            data = json.loads(PROCESSED_IDS_FILE.read_text(encoding="utf-8"))
            return set(data[-500:])  # 只保留最近500条
        except Exception:
            pass
    return set()


def save_processed_id(message_id: str) -> None:
    """保存消息ID到磁盘"""
    PROCESSED_IDS_FILE.parent.mkdir(parents=True, exist_ok=True)
    ids = list(processed_ids)[-500:]  # 只保留最近500条
    PROCESSED_IDS_FILE.write_text(json.dumps(ids, ensure_ascii=False), encoding="utf-8")


processed_ids = load_processed_ids()


def get_tenant_access_token() -> str:
    """获取 tenant_access_token"""
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    resp = requests.post(url, json={"app_id": APP_ID, "app_secret": APP_SECRET}, timeout=10)
    return resp.json().get("tenant_access_token", "")


def _send_file_to_feishu(target_id: str, file_path, id_type: str = "open_id", file_type: str = "stream"):
    """上传文件到飞书并发送到对话"""
    from pathlib import Path
    import json as _json

    file_path = Path(file_path)
    if not file_path.exists():
        print(f"[Feishu] File not found: {file_path}")
        return False

    # 根据文件扩展名确定 MIME 类型
    suffix = file_path.suffix.lower()
    mime_types = {
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.csv': 'text/csv',
        '.pdf': 'application/pdf',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.json': 'application/json',
        '.txt': 'text/plain',
        '.md': 'text/markdown',
    }
    mime_type = mime_types.get(suffix, 'application/octet-stream')

    try:
        # Step 1: 上传文件到飞书
        upload_url = "https://open.feishu.cn/open-apis/im/v1/files"
        token = get_tenant_access_token()

        if not token:
            print("[Feishu] Failed to get token")
            return False

        headers = {"Authorization": f"Bearer {token}"}

        with open(file_path, 'rb') as f:
            files = {
                'file': (file_path.name, f, mime_type),
                'file_type': (None, file_type),
                'file_name': (None, file_path.name),
            }
            resp = requests.post(upload_url, headers=headers, files=files, timeout=60)

        if resp.status_code != 200:
            print(f"[Feishu] Upload failed: {resp.status_code}")
            return False

        result = resp.json()
        if result.get("code") != 0:
            print(f"[Feishu] Upload failed: {result.get('msg')}")
            return False

        file_key = result["data"]["file_key"]

        # Step 2: 发送文件消息
        send_url = f"https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type={id_type}"

        payload = {
            "receive_id": target_id,
            "msg_type": "file",
            "content": _json.dumps({"file_key": file_key})
        }

        resp2 = requests.post(send_url, headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }, json=payload, timeout=30)

        if resp2.status_code == 200 and resp2.json().get("code") == 0:
            print(f"[Feishu] File sent: {file_path.name}")
            return True
        else:
            print(f"[Feishu] Send failed: {resp2.json()}")
            return False
    except Exception as e:
        print(f"[Feishu] File send error: {e}")
        return False


def _clean_synthesis_output(result_text: str) -> str:
    """清理整合输出：提取 CPO 整合部分，兜底发完整内容

    原则：宁可发完整内容（含 Agent 原始标记），绝不能发空内容。
    """
    import re

    if not result_text or len(result_text.strip()) < 50:
        return result_text

    # 如果输出不包含 Agent 标记，直接返回
    if "=== CTO" not in result_text and "===CMO" not in result_text and "=== CDO" not in result_text:
        return result_text

    # 尝试找 CPO 整合部分（通常在各 Agent 方案之后）
    # 方案1：找 "---" 分隔后的内容
    parts = re.split(r'\n---+\n', result_text)
    if len(parts) > 1:
        last_part = parts[-1].strip()
        if len(last_part) > 200:
            return last_part

    # 方案2：找最后一个 "=== " 之后的内容
    last_section_match = re.search(r'===\s*[^=]+\s*===\s*([\s\S]+)$', result_text)
    if last_section_match:
        last_content = last_section_match.group(1).strip()
        if len(last_content) > 200:
            return last_content

    # 方案3：移除各 Agent 原始段落标记，保留内容
    # 只移除标记行，不移除内容
    cleaned = re.sub(r'={2,}\s*(CTO|CMO|CDO)\s*(技术方案|市场策略|设计方案|方案)?\s*={2,}', '', result_text)
    cleaned = re.sub(r'(CTO|CMO|CDO)(认为|建议|方案|指出|提出)[：:]\s*', '', cleaned)
    cleaned = cleaned.strip()

    # 兜底：清理后如果太短，返回原始内容
    if len(cleaned) < 200:
        return result_text

    return cleaned


def _send_rd_result(synthesis_output: str, task_id: str, task_goal: str,
                     reply_target: str, reply_type: str):
    """发送研发任务结果：优先 Excel，降级树形文本，绝不发原始 JSON

    Args:
        synthesis_output: CPO 整合输出
        task_id: 任务 ID
        task_goal: 任务目标
        reply_target: 回复目标 ID
        reply_type: 回复类型
    """
    import re as _re
    import json as _json

    # Step 1: 尝试检测 JSON 并导出 Excel
    json_match = _re.search(r'\[[\s\S]*\]', synthesis_output)

    if json_match:
        try:
            items = _json.loads(json_match.group())
            if isinstance(items, list) and len(items) > 0:
                print(f"[RD Result] 检测到 JSON: {len(items)} 条")

                # 导出 Excel
                xlsx_path = _export_to_excel(items, task_id, task_goal)
                print(f"[RD Result] Excel 生成: {xlsx_path}")

                # 发送 Excel 文件到飞书
                file_sent = _send_file_to_feishu(reply_target, xlsx_path, reply_type)

                if file_sent:
                    # 发送树形摘要作为预览
                    summary = _format_items_as_tree(items)
                    send_reply(reply_target,
                        f"📊 功能PRD清单已生成（{len(items)} 条功能），Excel 文件已发送。\n\n"
                        f"预览：\n{summary[:3000]}",
                        reply_type)
                else:
                    # 文件发送失败，发树形文本
                    summary = _format_items_as_tree(items)
                    send_reply(reply_target,
                        f"📋 功能PRD清单（{len(items)} 条功能）：\n\n{summary[:4000]}\n\n"
                        f"⚠️ Excel 文件发送失败，已保存在服务器。",
                        reply_type)

                return
        except Exception as e:
            print(f"[RD Result] JSON/Excel 处理失败: {e}")
            import traceback
            traceback.print_exc()

    # Step 2: 非 JSON 内容，过滤掉 Agent 原始标记后发送
    clean = synthesis_output

    # 去掉 "=== CTO/CMO/CDO xxx ===" 标记
    clean = _re.sub(r'===\s*(CTO|CMO|CDO)\s*[^=]*===', '', clean)
    # 去掉 "=== 汇聚状态 ==="
    clean = _re.sub(r'===\s*汇聚[^=]*===', '', clean)
    clean = clean.strip()

    # 兜底：永远不发空内容
    if len(clean) < 100:
        clean = synthesis_output  # 不清理了，发原始内容

    send_reply(reply_target, clean[:4000], reply_type)


def _format_items_as_tree(items: list, max_items: int = 50) -> str:
    """将功能清单格式化为树形文本摘要

    绝对不发原始 JSON，只发格式化的可读文本。
    """
    if not items or not isinstance(items, list):
        return "（无有效数据）"

    lines = [f"📋 功能PRD清单摘要（共 {len(items)} 条）\n"]
    current_l1 = ""

    for item in items[:max_items]:
        level = item.get("level", "")
        name = item.get("name", "")
        priority = item.get("priority", "")

        if level == "L1":
            current_l1 = name
            lines.append(f"\n▎{name} [{priority}]")
        elif level == "L2":
            lines.append(f"  ├ {name} [{priority}]")
        elif level == "L3":
            lines.append(f"  │  └ {name} [{priority}]")
        else:
            # 兜底：直接显示
            lines.append(f"  • {name} [{priority}]")

    if len(items) > max_items:
        lines.append(f"\n... 还有 {len(items) - max_items} 条，完整清单见 Excel 文件")

    return "\n".join(lines)


def _export_to_excel(items: list, task_id: str, task_goal: str = "") -> str:
    """将结构化功能清单导出为格式化的 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "功能PRD清单"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 优先级颜色
    priority_fills = {
        "P0": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "P1": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        "P2": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        "P3": PatternFill(start_color="808080", end_color="808080", fill_type="solid"),
    }
    priority_fonts = {
        "P0": Font(bold=True, color="FFFFFF"),
        "P1": Font(bold=True, color="FFFFFF"),
        "P2": Font(color="FFFFFF"),
        "P3": Font(color="FFFFFF"),
    }

    # L1 行样式（加粗背景）
    l1_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    l1_font = Font(bold=True, size=11)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 写表头
    headers = ["模块", "层级", "父功能", "功能名称", "优先级", "交互方式", "功能描述", "验收标准", "依赖功能", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, item in enumerate(items, 2):
        values = [
            item.get("module", ""),
            item.get("level", ""),
            item.get("parent", ""),
            item.get("name", ""),
            item.get("priority", ""),
            item.get("interaction", ""),
            item.get("description", ""),
            item.get("acceptance", ""),
            item.get("dependencies", ""),
            item.get("note", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(val))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # L1 行样式
        level = item.get("level", "")
        if level == "L1":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = l1_fill
                ws.cell(row=row_idx, column=col).font = l1_font

        # 优先级颜色
        priority = item.get("priority", "").upper()
        if priority in priority_fills:
            p_cell = ws.cell(row=row_idx, column=5)
            p_cell.fill = priority_fills[priority]
            p_cell.font = priority_fonts[priority]
            p_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    ws.column_dimensions['A'].width = 16  # 模块
    ws.column_dimensions['B'].width = 8   # 层级
    ws.column_dimensions['C'].width = 16  # 父功能
    ws.column_dimensions['D'].width = 24  # 功能名称
    ws.column_dimensions['E'].width = 10  # 优先级
    ws.column_dimensions['F'].width = 14  # 交互方式
    ws.column_dimensions['G'].width = 40  # 功能描述
    ws.column_dimensions['H'].width = 35  # 验收标准
    ws.column_dimensions['I'].width = 16  # 依赖功能
    ws.column_dimensions['J'].width = 20  # 备注

    # 冻结表头
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:J{len(items) + 1}"

    # 保存
    export_dir = Path(__file__).parent.parent / ".ai-state" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = export_dir / f"prd_{task_id}.xlsx"
    wb.save(str(xlsx_path))

    print(f"[Export] Excel: {len(items)} items -> {xlsx_path.name}")
    return str(xlsx_path)


def download_feishu_image(image_key: str, message_id: str) -> bytes:
    token = get_tenant_access_token()
    # 飞书消息图片需要用 message_id + file_key 下载
    url = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/resources/{image_key}?type=image"
    headers = {"Authorization": f"Bearer {token}"}
    print(f"  下载URL: {url}")
    try:
        response = requests.get(url, headers=headers, timeout=30)
        print(f"  HTTP状态: {response.status_code}")
        if response.status_code == 200:
            return response.content
        else:
            print(f"  下载响应: {response.text[:200]}")
    except Exception as e:
        print(f"  下载异常: {e}")
    return None


def call_llm_chat(text: str) -> str:
    """非研发任务：带知识库上下文的智能对话"""
    from src.tools.knowledge_base import search_knowledge, format_knowledge_for_prompt, KB_ROOT

    gateway = get_model_gateway()

    # 搜索知识库
    kb_entries = search_knowledge(text, limit=5)
    kb_context = format_knowledge_for_prompt(kb_entries) if kb_entries else ""

    # 读取产品锚点
    product_anchor = ""
    try:
        for f in KB_ROOT.rglob("*.json"):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                tags = data.get("tags", [])
                if "internal" in tags and ("prd" in tags or "product_definition" in tags):
                    product_anchor = data.get("content", "")[:1000]
                    break
            except:
                continue
    except:
        pass

    # 构建有上下文的 prompt
    context_prompt = (
        f"你是智能摩托车全盔项目的 AI 合伙人。\n"
        f"基于项目知识库和产品定义回答用户的问题。\n"
        f"如果问题和项目相关，引用知识库中的具体数据回答。\n"
        f"如果问题和项目无关，也可以正常回答，但保持专业。\n"
        f"回答要简洁有力，像合伙人之间对话，不要像客服。\n\n"
    )

    if product_anchor:
        context_prompt += f"## 产品定义\n{product_anchor[:500]}\n\n"
    if kb_context:
        context_prompt += f"{kb_context[:3000]}\n\n"

    context_prompt += f"## 用户问题\n{text}"

    result = gateway.call_azure_openai("cpo", context_prompt,
        "你是项目合伙人Leo's Agent。回复简洁有力，300-500字，像面对面说话。", "smart_chat")
    if result.get("success"):
        return result["response"]
    return f"[回复失败: {result.get('error', '未知错误')}]"


def handle_fix_command(text: str, open_id: str, reply_target: str = None) -> bool:
    """处理修复审批指令，返回 True 表示已处理"""
    if reply_target is None:
        reply_target = open_id
    text_lower = text.strip()
    if text_lower.startswith("批准 fix_") or text_lower.startswith("approve fix_"):
        fix_id = text_lower.split()[-1]
        result = approve_and_execute(fix_id)
        if result["success"]:
            send_reply(reply_target, f"✅ 修复 {fix_id} 已执行\n文件: {result['file']}")
        else:
            send_reply(reply_target, f"❌ 执行失败: {result['error']}")
        return True
    elif text_lower.startswith("驳回 fix_") or text_lower.startswith("reject fix_"):
        parts = text_lower.split(maxsplit=2)
        fix_id = parts[1] if len(parts) > 1 else ""
        reason = parts[2] if len(parts) > 2 else ""
        result = reject_proposal(fix_id, reason)
        if result["success"]:
            send_reply(reply_target, f"🚫 修复 {fix_id} 已驳回")
        else:
            send_reply(reply_target, f"❌ 驳回失败: {result['error']}")
        return True
    elif text_lower in ("待审批", "pending", "审批列表"):
        proposals = get_pending_proposals()
        if not proposals:
            send_reply(reply_target, "📋 当前无待审批的修复提案")
        else:
            for p in proposals:
                send_reply(reply_target, format_proposal_for_feishu(p))
        return True
    return False


def handle_rating(text: str, open_id: str, reply_target: str = None) -> bool:
    """处理用户评价，回填到最近的经验卡片"""
    if reply_target is None:
        reply_target = open_id
    # 调试：记录原始输入
    print(f"  [handle_rating] text={repr(text)}, len={len(text)}")
    text = text.strip()
    rating_map = {"a": "A", "b": "B", "c": "C", "d": "D",
                  "A": "A", "B": "B", "C": "C", "D": "D"}
    first_char = text[0] if text else ""
    print(f"  [handle_rating] first_char={repr(first_char)}")
    if first_char not in rating_map:
        return False
    rating = rating_map[first_char]
    feedback = text[1:].strip().lstrip(".").lstrip("。").lstrip("、").lstrip(",").strip() if len(text) > 1 else ""
    global _last_task_memory
    if not _last_task_memory.get("memory_dir"):
        send_reply(reply_target, f"📝 收到评价 {rating}，但没有找到最近的任务记录")
        return True
    memory_dir = Path(_last_task_memory["memory_dir"])
    if not memory_dir.exists():
        send_reply(reply_target, f"📝 收到评价 {rating}，记忆目录不存在")
        return True
    # 找到最近的经验卡片
    files = sorted(memory_dir.glob("*.json"), reverse=True)
    if not files:
        send_reply(reply_target, f"📝 收到评价 {rating}，但没有经验卡片")
        return True
    latest = files[0]
    try:
        data = json.loads(latest.read_text(encoding="utf-8"))
        data["user_rating"] = rating
        if feedback:
            data["user_feedback"] = feedback
        latest.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        rating_labels = {"A": "可直接使用", "B": "需要小改", "C": "方向对但不够深", "D": "方向有问题"}
        msg = f"📝 评价已记录：{rating} ({rating_labels.get(rating, '')})"
        if feedback:
            msg += f"\n反馈：{feedback}"
        send_reply(reply_target, msg)

        # === 评价驱动进化 ===
        if rating in ("C", "D"):
            # 差评 → 自动分析失败原因，写入经验教训
            def _analyze_failure():
                print("[Evolution] _analyze_failure 线程已启动")
                try:
                    from src.utils.model_gateway import get_model_gateway
                    from src.tools.knowledge_base import add_knowledge
                    from src.utils.critic_rules import add_critic_rule

                    gw = get_model_gateway()
                    task_goal = data.get("task_goal", "")
                    synthesis = data.get("synthesis_output", "")
                    user_feedback = feedback if feedback else f"用户评价{rating}"

                    analysis_prompt = (
                        f"一个研发任务收到了差评（{rating}）。请分析失败原因并提取教训和检查规则。\n\n"
                        f"## 任务目标\n{task_goal}\n\n"
                        f"## Agent 输出（摘要）\n{synthesis[:2000]}\n\n"
                        f"## 用户反馈\n{user_feedback}\n\n"
                        f"请输出两部分：\n\n"
                        f"## PART 1: 失败分析（自然语言）\n"
                        f"1. 失败根因（一句话）\n"
                        f"2. Agent 哪个环节出了问题（CPO规划/CTO技术/CMO市场/Critic评审/知识库不足）\n"
                        f"3. 下次遇到类似任务应该怎么做\n"
                        f"4. 需要补充什么知识\n"
                        f"控制在 300 字以内。\n\n"
                        f"## PART 2: 检查规则（JSON 格式，单独一行）\n"
                        f"基于这次失败，生成一条 Critic 评审必须检查的规则。\n"
                        f"格式：CHECK_RULE_JSON:{{\"check_description\": \"具体描述下次 Critic 应该检查什么\", \"trigger_context\": \"什么类型的任务应该触发这条规则\"}}\n"
                        f"规则要具体可执行，不要泛泛的'要更仔细'。\n"
                        f"示例：CHECK_RULE_JSON:{{\"check_description\": \"mesh 相关任务必须先分析 Cardo DMC 方案可行性，不能直接跳到放弃 Mesh\", \"trigger_context\": \"mesh 对讲 intercom 通讯\"}}"
                    )

                    print("[Evolution] 开始调用 LLM 分析...")
                    result = gw.call_azure_openai("cpo", analysis_prompt, "你是质量分析师。输出失败分析和一条检查规则。", "failure_analysis")
                    print(f"[Evolution] LLM 返回: success={result.get('success')}")

                    if not result.get("success"):
                        print(f"[Evolution] LLM 调用失败: {result.get('error', '')[:200]}")
                        # Fallback: 直接用用户反馈生成规则
                        rule_path = add_critic_rule(
                            check_description=f"用户反馈: {user_feedback[:100]}",
                            trigger_context=task_goal[:100],
                            severity="must_check",
                            source=f"user_rating_{rating}",
                            source_task_id=_last_task_memory.get("task_id", "")
                        )
                        if rule_path:
                            print(f"[Evolution] Fallback 规则已生成")
                            send_reply(reply_target, "📋 已从评价中提取检查规则（简化版）。")
                        else:
                            send_reply(reply_target, "📋 规则已存在或达上限。")
                        return

                    if result.get("success"):
                        response_text = result["response"]

                        # 写入教训（原有逻辑）
                        add_knowledge(
                            title=f"[教训] {task_goal[:40]}（评价{rating}）",
                            domain="lessons",
                            content=response_text,
                            tags=["evolution", "failure", f"rating_{rating.lower()}"],
                            source="user_feedback_analysis",
                            confidence="high",
                            caller="user_feedback_analysis"
                        )
                        print(f"[Evolution] 差评分析完成，已写入知识库")

                        # === 新增：提取并写入检查规则 ===
                        import re
                        rule_match = re.search(r'CHECK_RULE_JSON:\s*(\{.*?\})', response_text, re.DOTALL)
                        if rule_match:
                            try:
                                import json as _json
                                rule_data = _json.loads(rule_match.group(1))
                                rule_path = add_critic_rule(
                                    check_description=rule_data.get("check_description", ""),
                                    trigger_context=rule_data.get("trigger_context", ""),
                                    severity="must_check",
                                    source="user_rating_D" if rating == "D" else "user_rating_C",
                                    source_task_id=_last_task_memory.get("task_id", "")
                                )
                                if rule_path:
                                    print(f"[Evolution] 检查规则已生成: {rule_data.get('check_description', '')[:60]}")
                                    send_reply(reply_target, f"📋 已从评价中提取检查规则，Critic 下次会自动检查。")
                                else:
                                    print(f"[Evolution] 规则已存在或达上限，跳过")
                            except Exception as rule_err:
                                print(f"[Evolution] 规则解析失败: {rule_err}")
                        else:
                            print(f"[Evolution] LLM 未输出 CHECK_RULE_JSON，跳过规则生成")

                        send_reply(reply_target, f"已分析任务失败原因并记录为经验教训，下次类似任务会注意。")

                        # === Enhanced: Auto attribution analysis for D ratings ===
                        if rating == "D":
                            try:
                                attribution_prompt = (
                                    f"A R&D task received poor rating(D). Analyze root cause and categorize.\n\n"
                                    f"User requirement: {task_goal[:500]}\n\n"
                                    f"System output (summary): {synthesis[:1000]}\n\n"
                                    f"Categorize from these dimensions (multi-select):\n"
                                    f"1. KNOWLEDGE_GAP: Knowledge base lacks relevant data\n"
                                    f"2. FORMAT_WRONG: Output format doesn't match user requirement\n"
                                    f"3. GOAL_MISALIGN: Didn't align with user goal\n"
                                    f"4. SPECULATIVE: Output too much speculative content\n"
                                    f"5. INCOMPLETE: Missing explicitly requested content\n"
                                    f"6. TOO_VERBOSE: Output too verbose, not concise\n"
                                    f"7. WRONG_ROUTE: Used wrong processing path (e.g., should use fast track but used multi-agent)\n\n"
                                    f"Output JSON: {{'causes':['cause1','cause2'],'fix_actions':['fix1','fix2'],'knowledge_gaps':['gap1']}}\n"
                                    f"Only output JSON."
                                )

                                attribution = gw.call_azure_openai("cpo", attribution_prompt,
                                    "Analyze failure causes. Only output JSON.", "failure_attribution")

                                if attribution.get("success"):
                                    import re as _re
                                    try:
                                        attr_result = _json.loads(_re.search(r'\{[\s\S]*\}', attribution["response"]).group())

                                        causes = attr_result.get("causes", [])
                                        knowledge_gaps = attr_result.get("knowledge_gaps", [])

                                        print(f"[Evolution] D rating attribution: {causes}")

                                        # If knowledge gap, auto trigger deep dive
                                        if "KNOWLEDGE_GAP" in causes and knowledge_gaps:
                                            try:
                                                from scripts.self_test import auto_deep_dive_weak_areas
                                                weak = [{"question": g, "domain": "components",
                                                         "suggested_searches": [g]} for g in knowledge_gaps[:3]]
                                                auto_deep_dive_weak_areas(weak)
                                                print(f"[Evolution] Auto deep dive triggered for knowledge gaps: {knowledge_gaps}")
                                            except ImportError:
                                                print("[Evolution] self_test module not available for auto deep dive")

                                        # Save attribution result for trend analysis
                                        evolution_dir = Path(__file__).parent.parent / ".ai-state" / "evolution"
                                        evolution_dir.mkdir(parents=True, exist_ok=True)
                                        (evolution_dir / f"d_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json").write_text(
                                            _json.dumps({"task": task_goal[:200], "causes": causes,
                                                       "fixes": attr_result.get("fix_actions", []),
                                                       "gaps": knowledge_gaps}, ensure_ascii=False, indent=2),
                                            encoding="utf-8"
                                        )
                                    except Exception as parse_err:
                                        print(f"[Evolution] Attribution parse failed: {parse_err}")
                            except Exception as attr_err:
                                print(f"[Evolution] Attribution analysis failed: {attr_err}")

                except Exception as e:
                    print(f"[Evolution] 差评分析失败: {e}")
                    import traceback
                    print(traceback.format_exc())

            threading.Thread(target=_analyze_failure, daemon=True).start()

        elif rating == "A":
            # 好评 → 提取成功模式
            def _extract_success():
                try:
                    from src.utils.model_gateway import get_model_gateway
                    from src.tools.knowledge_base import add_knowledge

                    gw = get_model_gateway()
                    task_goal = data.get("task_goal", "")
                    synthesis = data.get("synthesis_output", "")

                    success_prompt = (
                        f"一个研发任务收到了满分评价（A）。请提取成功模式。\n\n"
                        f"## 任务目标\n{task_goal}\n\n"
                        f"## Agent 输出（摘要）\n{synthesis[:2000]}\n\n"
                        f"请输出：\n"
                        f"1. 这个任务为什么做得好（一句话）\n"
                        f"2. 哪些做法值得复制到其他任务\n"
                        f"3. 成功的关键因素是什么\n"
                        f"控制在 200 字以内。"
                    )

                    result = gw.call_azure_openai("cpo", success_prompt, "你是质量分析师。", "success_analysis")

                    if result.get("success"):
                        add_knowledge(
                            title=f"[成功模式] {task_goal[:40]}",
                            domain="lessons",
                            content=result["response"],
                            tags=["evolution", "success", "rating_a"],
                            source="user_feedback_analysis",
                            confidence="high",
                            caller="user_feedback_analysis"
                        )
                        print(f"[Evolution] 成功模式提取完成")

                        # === 新增：从成功模式提取推荐做法规则 ===
                        try:
                            from src.utils.critic_rules import add_critic_rule
                            # 用 LLM 从成功模式中提取规则
                            rule_prompt = (
                                f"以下是一个成功任务的分析。请提取一条 Critic 评审的推荐做法规则。\n\n"
                                f"## 任务目标\n{task_goal}\n\n"
                                f"## 成功分析\n{result['response'][:1000]}\n\n"
                                f"输出一行 JSON：CHECK_RULE_JSON:{{\"check_description\": \"推荐做法\", \"trigger_context\": \"触发场景\"}}"
                            )
                            rule_result = gw.call_azure_openai("cpo", rule_prompt, "只输出一行 CHECK_RULE_JSON。", "success_rule")
                            if rule_result.get("success"):
                                import re
                                rule_match = re.search(r'CHECK_RULE_JSON:\s*(\{.*?\})', rule_result["response"], re.DOTALL)
                                if rule_match:
                                    import json as _json
                                    rule_data = _json.loads(rule_match.group(1))
                                    add_critic_rule(
                                        check_description=rule_data.get("check_description", ""),
                                        trigger_context=rule_data.get("trigger_context", ""),
                                        severity="should_check",
                                        source="user_rating_A",
                                        source_task_id=_last_task_memory.get("task_id", "")
                                    )
                        except Exception as rule_err:
                            print(f"[Evolution] A评价规则提取失败: {rule_err}")
                except Exception as e:
                    print(f"[Evolution] 成功分析失败: {e}")

            threading.Thread(target=_extract_success, daemon=True).start()

    except Exception as e:
        send_reply(reply_target, f"📝 评价记录失败：{e}")
    return True


def is_rd_task(text: str) -> bool:
    """判断是否为需要多Agent协作的研发任务"""
    keywords = [
        "设计方案", "技术方案", "架构", "需求分析",
        "市场策略", "GTM", "竞品分析", "产品规划",
        "研发", "开发方案", "实现方案", "技术选型",
        "设计", "模块", "方案", "选型", "评估",
        "原型", "验证", "测试方案", "硬件", "固件",
        "通信", "传感器", "算法", "HUD", "蓝牙",
        "导航", "供应链", "认证", "量产"
    ]
    return any(kw in text for kw in keywords)


def _is_likely_article(text: str, has_url: bool = False) -> bool:
    """判断是否为应该导入知识库的长文章（而非研发任务）

    规则：
    1. 纯文本 > 800 字，且无明确指令词 → 文章
    2. 包含 URL 且正文 > 500 字 → 文章
    3. 短消息（< 500 字）或有指令词 → 不是文章
    """
    text_len = len(text.strip())

    # 指令词开头
    command_prefixes = ["研究", "分析一下", "帮我", "请", "设计", "方案", "@dev"]
    has_command = any(text.strip().startswith(p) for p in command_prefixes)

    # 规则 3：短消息或有指令词 → 不是文章
    if text_len < 500 or has_command:
        return False

    # 规则 2：有 URL 且正文 > 500 字 → 文章
    if has_url and text_len > 500:
        return True

    # 规则 1：纯文本 > 800 字 → 文章
    if text_len > 800:
        return True

    return False


def _ocr_image_fallback(image_data: bytes) -> str:
    """OCR 降级方案：使用 pytesseract 提取图片中的文字"""
    try:
        import pytesseract
        from PIL import Image
        from io import BytesIO

        img = Image.open(BytesIO(image_data))
        # 使用中文+英文识别
        text = pytesseract.image_to_string(img, lang='chi_sim+eng')
        if text.strip():
            return text.strip()[:200]  # 限制长度
    except ImportError:
        print("  [OCR] pytesseract 未安装，跳过 OCR")
    except Exception as e:
        print(f"  [OCR] 识别失败: {e}")
    return None


def _save_pending_image(image_data: bytes, image_key: str, message_id: str) -> str:
    """保存待处理的图片到 pending_images 目录"""
    pending_dir = Path(__file__).parent.parent / ".ai-state" / "pending_images"
    pending_dir.mkdir(parents=True, exist_ok=True)

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{image_key[:20]}.jpg"
    filepath = pending_dir / filename

    try:
        filepath.write_bytes(image_data)
        return str(filepath)
    except Exception as e:
        print(f"  [Pending] 保存失败: {e}")
        return None


def _extract_images_from_post(post_content: list, message_id: str, open_id: str) -> list:
    """从 post 消息中提取并理解图片内容（带完整降级策略）

    降级链：Gemini Vision → 重试 → Flash Vision → OCR → 保存待处理
    返回：图片理解结果列表（每张图片都有结果）
    """
    image_keys = []
    for block in post_content:
        for element in block:
            if element.get("tag") == "img":
                img_key = element.get("image_key", "")
                if img_key:
                    image_keys.append(img_key)

    if not image_keys:
        return []

    # 不限制图片数量，全部处理
    print(f"  [图片处理] 检测到 {len(image_keys)} 张图片")
    image_descriptions = []
    gateway = get_model_gateway()
    vision_prompt = "请用中文简洁描述这张图片的内容，重点关注：1. 如果是产品图，描述产品形态、关键特征；2. 如果是图表，描述数据含义；3. 如果是截图，描述关键信息。回复控制在100字以内。"

    for i, img_key in enumerate(image_keys):
        result_desc = None
        image_data = None

        try:
            # Step 1: 下载图片
            image_data = download_feishu_image(img_key, message_id)
            if not image_data:
                result_desc = f"[图片{i+1}: 下载失败]"
                image_descriptions.append(result_desc)
                continue

            # Step 2: 压缩图片（加速传输）
            try:
                image_data = compress_image(image_data, max_size=1024)
            except Exception as comp_err:
                print(f"  图片 {i+1} 压缩失败: {comp_err}，使用原图")

            # Step 3: 尝试 Gemini Vision (gemini_2_5_flash)
            result = gateway.call_gemini_vision("gemini_2_5_flash", image_data, vision_prompt)

            if result.get("success"):
                desc = result.get("response", "")[:150]
                result_desc = f"[图片{i+1}: {desc}]"
                print(f"  图片 {i+1} Flash Vision 成功")
            else:
                print(f"  图片 {i+1} Flash Vision 失败: {result.get('error', '')[:50]}，尝试重试...")

                # Step 4: 重试一次 Flash Vision
                import time
                time.sleep(1)
                result = gateway.call_gemini_vision("gemini_2_5_flash", image_data, vision_prompt)

                if result.get("success"):
                    desc = result.get("response", "")[:150]
                    result_desc = f"[图片{i+1}: {desc}]"
                    print(f"  图片 {i+1} Flash Vision 重试成功")
                else:
                    print(f"  图片 {i+1} Flash Vision 重试仍失败，尝试 OCR...")

                    # Step 5: 降级到 OCR
                    ocr_text = _ocr_image_fallback(image_data)
                    if ocr_text:
                        result_desc = f"[图片{i+1} OCR文字]: {ocr_text}"
                        print(f"  图片 {i+1} OCR 成功")
                    else:
                        # Step 6: 保存待处理图片
                        saved_path = _save_pending_image(image_data, img_key, message_id)
                        if saved_path:
                            result_desc = f"[图片{i+1}: 暂时无法识别，已保存待后续处理]"
                            print(f"  图片 {i+1} 已保存到 {saved_path}")
                        else:
                            result_desc = f"[图片{i+1}: 处理失败]"

        except Exception as e:
            print(f"  图片 {i+1} 处理异常: {e}")
            # 尝试保存已下载的图片
            if image_data:
                saved_path = _save_pending_image(image_data, img_key, message_id)
                if saved_path:
                    result_desc = f"[图片{i+1}: 处理异常，已保存待处理]"
                else:
                    result_desc = f"[图片{i+1}: 处理失败]"
            else:
                result_desc = f"[图片{i+1}: 处理失败]"

        # 确保每张图片都有结果
        if result_desc:
            image_descriptions.append(result_desc)
        else:
            image_descriptions.append(f"[图片{i+1}: 未知错误]")

    print(f"  [图片处理] 完成，{len(image_descriptions)} 张图片已处理")
    return image_descriptions


def _import_article_to_kb(text: str, image_descriptions: list, open_id: str, title: str = None) -> str:
    """将文章导入知识库

    返回：导入结果消息
    """
    from src.tools.knowledge_base import add_knowledge

    # 合并文本和图片理解
    full_content = text
    if image_descriptions:
        full_content += "\n\n--- 图片内容 ---\n" + "\n".join(image_descriptions)

    # 根据内容判断 domain
    domain = "lessons"
    text_lower = text.lower()
    if any(kw in text_lower for kw in ["竞品", "对手", "竞争", "市场", "行业"]):
        domain = "competitors"
    elif any(kw in text_lower for kw in ["芯片", "传感器", "模组", "bom", "供应商", "成本"]):
        domain = "components"
    elif any(kw in text_lower for kw in ["认证", "标准", "法规", "合规"]):
        domain = "standards"

    # 自动生成标题（取前 30 字）
    if not title:
        title = text.strip()[:50].replace("\n", " ")
        if len(text) > 50:
            title += "..."

    # 存入知识库
    try:
        path = add_knowledge(
            title=title,
            domain=domain,
            content=full_content,
            tags=["feishu_import", "article"],
            source="飞书分享",
            confidence="medium",
            caller="user_share"
        )
        return f"已学习：{title[:40]}\n存入 [{domain}] 共 {len(full_content)} 字"
    except Exception as e:
        return f"导入失败: {e}"


def _try_generate_design_image(cdo_output: dict, reply_target: str, reply_type: str = None) -> None:
    """从 CDO 输出中提取 AI_IMAGE_PROMPT 并生成图片"""
    import base64 as b64
    design = cdo_output.get("execution", {}).get("cdo_output", {}).get("design_proposal", "")
    if "[AI_IMAGE_PROMPT]" not in design:
        return
    prompt_section = design.split("[AI_IMAGE_PROMPT]")[-1].strip()
    lines = [l.strip() for l in prompt_section.split("\n") if l.strip() and not l.strip().startswith("[")]
    if not lines:
        return
    # 取第一个 prompt 生图
    first_prompt = lines[0].lstrip("1. ").lstrip("- ")
    if len(first_prompt) < 10:
        return
    send_reply(reply_target, "🖼️ 正在生成设计概念图...", reply_type)
    from src.tools.tool_registry import get_tool_registry
    result = get_tool_registry().call("image_generation", first_prompt)
    if result.get("success") and result.get("image_base64"):
        image_bytes = b64.b64decode(result["image_base64"])
        send_image_reply(reply_target, image_bytes, reply_type or "open_id")
    else:
        send_reply(reply_target, f"[概念图生成失败: {result.get('error', '未知')}]", reply_type)


def _stream_langgraph(app, initial_state: dict, reply_target: str = None, reply_type: str = None, config: dict = None) -> dict:
    """LangGraph stream 模式执行，精简进度消息"""
    # 确保 config 包含 thread_id（checkpoint 必须）
    if config is None:
        config = {}
    if "configurable" not in config:
        config["configurable"] = {}
    if "thread_id" not in config["configurable"]:
        config["configurable"]["thread_id"] = f"stream_{uuid.uuid4().hex[:8]}"

    # 只保留关键阶段的进度消息
    progress_map = {
        "cpo_plan": None,  # 不发送
        "cto_coder": "📖 研究中: 技术方案",
        "cmo_strategist": "📖 研究中: 市场策略",
        "cdo_designer": "📖 研究中: 设计方案",
        "state_merge": None,  # 不发送
        "cpo_synthesis": "📖 研究中: 方案整合",
        "cpo_critic": None,  # 不发送
        "memory_writer": None,  # 不发送
    }
    # 记录已发送的阶段，避免重复
    sent_stages = set()
    final_state = None
    for event in app.stream(initial_state, stream_mode="updates", config=config):
        for node_name, node_output in event.items():
            final_state = node_output if isinstance(node_output, dict) else final_state
            msg = progress_map.get(node_name)
            if reply_target and msg and msg not in sent_stages:
                send_reply(reply_target, msg, reply_type)
                sent_stages.add(msg)
    return final_state


def call_langgraph(text: str, task_role: str = "cto", reply_target: str = None, reply_type: str = None) -> str:
    """调用 LangGraph 多Agent工作流，支持流式进度反馈"""
    task_id = f"task_{uuid.uuid4().hex[:8]}"

    # 判断需要 CTO、CMO 还是双Agent
    if any(kw in text for kw in ["市场", "GTM", "竞品", "定价", "用户"]):
        roles = ["cmo"]
    elif any(kw in text for kw in ["技术", "架构", "设计", "开发", "实现"]):
        roles = ["cto"]
    else:
        roles = ["cto", "cmo"]  # 默认双Agent

    sub_tasks = {}
    for role in roles:
        sub_id = f"{role}_{task_id}"
        sub_tasks[sub_id] = {
            "subtask_id": sub_id,
            "target_role": role,
            "task_description": text,
            "depends_on": [],
            "is_core_dependency": False,
            "dependency_timeout_sec": 120,
            "output_schema": {},
            "acceptance_criteria": {},
            "tool_white_list": []
        }

    initial_state = {
        "metadata": {
            "task_id": task_id,
            "global_status": "pending",
            "max_retry_threshold": 3
        },
        "sub_tasks": sub_tasks,
        "task_contract": {
            "task_goal": text
        }
    }

    # Checkpoint config：用 task_id 作为 thread_id，支持断点续传
    checkpoint_config = {"configurable": {"thread_id": task_id}}

    try:
        print(f"[LangGraph] 启动任务 {task_id}，角色：{roles}")
        # 优先使用 stream 模式，失败时 fallback 到 invoke
        try:
            result = _stream_langgraph(langgraph_app, initial_state, reply_target, reply_type,
                                        config=checkpoint_config)
        except Exception as stream_err:
            print(f"[LangGraph] stream 模式失败，fallback invoke: {stream_err}")
            result = langgraph_app.invoke(initial_state, config=checkpoint_config)

        if result is None:
            return "[多Agent任务未产生输出]"

        # 提取输出
        execution = result.get("execution", {})
        synthesis_output = execution.get("synthesis_output", "")
        merge_summary = execution.get("merge_summary", "")
        critic_feedback = execution.get("critic_feedback", "")
        critic_decision = execution.get("critic_decision", "")

        # 优先展示整合结论，没有则 fallback 到原始汇聚
        final_output = synthesis_output if synthesis_output else merge_summary

        # Critic PASS 后，从整合后的方案中提取 AI_IMAGE_PROMPT 生成概念图
        if reply_target and "[AI_IMAGE_PROMPT]" in str(final_output):
            _try_generate_design_image({"execution": {"cdo_output": {"design_proposal": final_output}}}, reply_target, reply_type)

        # 记录最近任务信息，用于后续评价关联
        global _last_task_memory
        _last_task_memory = {
            "task_id": task_id,
            "memory_dir": str(Path(__file__).parent.parent / ".ai-state" / "memory")
        }

        if final_output:
            char_count = len(final_output)
            header = f"[R&D] Done: {text[:40]}{'...' if len(text) > 40 else ''}"
            return f"{header}\n\n{final_output}\n\nReport saved ({char_count} chars). Send '日报' to check knowledge base."

        # 没有 merge_summary 时手动拼装
        parts = []
        cto_out = execution.get("cto_output", {})
        cmo_out = execution.get("cmo_output", {})
        cdo_out = execution.get("cdo_output", {})

        # 尝试多种字段名
        cto_text = cto_out.get("protocol_code") or cto_out.get("output") or cto_out.get("result") or ""
        cmo_text = cmo_out.get("market_strategy") or cmo_out.get("output") or cmo_out.get("result") or ""
        cdo_text = cdo_out.get("design_proposal") or cdo_out.get("output") or cdo_out.get("result") or ""

        if cto_text:
            parts.append(f"**CTO 技术方案**\n{str(cto_text)[:4000]}")
        if cmo_text:
            parts.append(f"**CMO 市场策略**\n{str(cmo_text)[:4000]}")
        if cdo_text:
            parts.append(f"**CDO 设计方案**\n{str(cdo_text)[:4000]}")

        # 如果有 critic_feedback 但没有其他输出，也展示
        if not parts and critic_feedback:
            parts.append(f"**Critic 反馈**\n{critic_feedback[:2000]}")

        if parts:
            result_text = "\n\n---\n\n".join(parts)
            # 格式化输出
            char_count = len(result_text)
            header = f"[R&D] Done: {text[:40]}{'...' if len(text) > 40 else ''}"
            return f"{header}\n\n{result_text}\n\nReport saved ({char_count} chars). Send '日报' to check knowledge base."

        # 最终保底
        return "[R&D] Task completed but no output generated. Check task description or logs."

    except Exception as e:
        return f"[LangGraph调用失败: {e}]"


def _run_rd_task_background(text: str, open_id: str, reply_target: str = None, reply_type: str = None) -> None:
    """后台线程执行研发任务，不阻塞主线程"""
    global _rd_task_running
    # 如果没传 reply_target，降级到 open_id（向后兼容）
    if reply_target is None:
        reply_target = open_id
    try:
        _rd_task_running = True
        reply = call_langgraph(text, reply_target=reply_target, reply_type=reply_type)
        if reply:
            # === 先检测 JSON 并导出 Excel，再决定发什么 ===
            import re as _re
            import json as _json
            import uuid as _uuid

            # 使用贪婪匹配获取完整 JSON 数组（从第一个 [ 到最后一个 ]）
            json_match = _re.search(r'\[[\s\S]*\]', reply)
            xlsx_path = None
            items = None

            if json_match:
                try:
                    items = _json.loads(json_match.group())
                    if isinstance(items, list) and len(items) > 0:
                        task_id = f"{_uuid.uuid4().hex[:8]}"
                        xlsx_path = _export_to_excel(items, task_id, text)
                        print(f"[Export] Excel 生成: {xlsx_path}, {len(items)} 条功能")
                except Exception as e:
                    print(f"[Export] JSON 解析失败: {e}")
                    items = None
                    xlsx_path = None

            # 根据 Excel 生成结果决定发送内容
            if xlsx_path:
                # Excel 生成成功，发送文件
                file_sent = _send_file_to_feishu(reply_target, xlsx_path, reply_type)
                print(f"[Export] 飞书文件发送: {'成功' if file_sent else '失败'}")

                if file_sent:
                    send_reply(reply_target, f"📊 功能PRD清单已导出为 Excel（{len(items)} 条功能），文件已发送。", reply_type)
                else:
                    # 文件发送失败，发送格式化文本摘要
                    print("[Export] 文件发送失败，降级到文本摘要")
                    summary = _format_items_as_tree(items)
                    send_reply(reply_target, summary[:4000], reply_type)
                    send_reply(reply_target, f"⚠️ 飞书文件发送失败，Excel 已保存在服务器: {xlsx_path}", reply_type)
            else:
                # 没有 JSON 或 Excel 生成失败，发送格式化文本（不发原始 JSON）
                clean_reply = _clean_synthesis_output(reply)
                # 如果清理后还是原始 JSON 格式，转为树形摘要
                if clean_reply.startswith('[') and '{' in clean_reply:
                    # 尝试解析并格式化
                    try:
                        items_retry = _json.loads(clean_reply)
                        if isinstance(items_retry, list):
                            clean_reply = _format_items_as_tree(items_retry)
                    except:
                        pass
                send_reply(reply_target, clean_reply[:4000], reply_type)

        send_reply(reply_target, "[Rating] Please rate this solution:\nA. Ready to use\nB. Needs minor changes\nC. Right direction but not deep enough\nD. Wrong direction\nReply with letter, C/D please include reason.", reply_type)
    except Exception as e:
        send_reply(reply_target, f"[Error] R&D task failed: {str(e)[:300]}", reply_type)
        print(f"[RD_TASK] Background execution error: {e}")
    finally:
        _rd_task_running = False


def compress_image(image_bytes: bytes, max_size: int = 1024) -> bytes:
    """压缩图片：限制最大边为 max_size 像素，JPEG 质量 85"""
    img = Image.open(BytesIO(image_bytes))
    if max(img.size) > max_size:
        ratio = max_size / max(img.size)
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size, Image.LANCZOS)
    output = BytesIO()
    img = img.convert("RGB")
    img.save(output, format="JPEG", quality=85)
    return output.getvalue()


def _has_shareable_url(text: str) -> bool:
    """检测消息是否包含可分享的 URL（不是提问）"""
    import re
    urls = re.findall(r'https?://[^\s<>"{}|\\^`\[\]]+', text)
    if not urls:
        return False
    # 去掉 URL 后剩余文字
    remaining = text
    for url in urls:
        remaining = remaining.replace(url, "").strip()
    # 清理小红书/微信的固定后缀文字
    noise_phrases = ["复制后打开", "查看笔记", "点击链接", "打开看看", "分享给你"]
    for phrase in noise_phrases:
        remaining = remaining.replace(phrase, "").strip()
    # 问号在最后 10 个字符内 = 可能在提问；问号在标题中间 = 不是提问
    if remaining.rstrip().endswith("?") or remaining.rstrip().endswith("？"):
        return False
    # 如果剩余文字超过 300 字且没有 URL 关键特征，可能是普通长文
    if len(remaining) > 300:
        return False
    return True


def handle_share_content(open_id: str, text: str = "", url: str = "", reply_target: str = None) -> None:
    """处理用户分享的 URL 或内容，自动入库"""
    if reply_target is None:
        reply_target = open_id
    # 统一导入所有模块（避免变量作用域问题）
    import re
    import json as _json
    import requests as _req
    import concurrent.futures
    from pathlib import Path
    from urllib.parse import urlparse
    from src.utils.model_gateway import get_model_gateway
    from src.tools.knowledge_base import add_knowledge
    from src.tools.tool_registry import get_tool_registry

    # 初始化全局对象
    registry = get_tool_registry()
    gateway = get_model_gateway()

    log(f"handle_share_content 被调用, text={text[:80] if text else url[:80]}")

    # 分享去重：同一 URL 5 分钟内不重复处理
    import hashlib
    import time as _time
    url_match_check = re.search(r'https?://[^\s<>"{}|\\^`\[\]]+', text or "")
    if url_match_check:
        url_hash = hashlib.md5(url_match_check.group(0).encode()).hexdigest()[:8]
        dedup_file = Path(__file__).parent.parent / ".ai-state" / f"share_dedup_{url_hash}"
        if dedup_file.exists():
            age = _time.time() - dedup_file.stat().st_mtime
            if age < 300:
                log(f"分享去重: 5分钟内已处理")
                return
        dedup_file.write_text(str(_time.time()))

    content_to_process = text or url
    if not content_to_process:
        send_reply(reply_target, "[Share] No content to process")
        return

    # 情况1：包含 URL，三步流程
    url_match = re.search(r'https?://[^\s<>"{}|\\^`\[\]]+', text or url or content_to_process)
    if url_match:
        url = url_match.group(0)
        send_reply(reply_target, "[Share] 正在读取和学习...")
        source_type = "url"

        try:
            domain = urlparse(url).netloc

            # 从消息文字提取标题线索
            text_without_url = re.sub(r'https?://[^\s]+', '', text).strip()
            for noise in ["复制后打开", "查看笔记", "点击链接", "打开看看", "分享给你",
                           "复制后打开【小红书】查看笔记", "【小红书】", "【微信】", "【知乎】"]:
                text_without_url = text_without_url.replace(noise, "").strip()
            text_without_url = text_without_url.strip("！!。.…，,\n\r")

            title_clue = text_without_url if len(text_without_url) > 3 else ""
            log(f"标题线索: '{title_clue}', 域名: {domain}")

            # === STEP 1：平台分流搜索（核心）===
            if title_clue and len(title_clue) > 10:
                # 带域名前缀让 platform_search 识别平台，但搜索时只用标题
                search_input = f"[{domain}] {title_clue}"
            else:
                # 纯裸链接，用 URL 搜
                search_input = url
            log(f"搜索输入: '{search_input[:60]}'")
            platform_result = registry.call("platform_search", search_input)

            platform_content = ""
            if platform_result.get("success"):
                platform_content = platform_result["data"]
                log(f"平台搜索成功: {len(platform_content)} 字, 平台={platform_result.get('platform','')}")

            # === STEP 2：如果有标题线索，用关键词再深搜一轮 ===
            extra_content = ""
            all_text = f"{title_clue}\n{platform_content[:500]}" if platform_content else title_clue

            if all_text and len(all_text) > 10:
                kw_result = gateway.call_azure_openai(
                    "cpo",
                    f"从以下内容提取核心主题的 3-5 个关键词（品牌/产品/技术/人名）。\n"
                    f"只输出关键词，空格分隔。\n"
                    f"禁止输出平台名（微信/小红书等）。\n\n{all_text[:1500]}",
                    "只输出关键词。",
                    "share_extract_keywords"
                )
                if kw_result.get("success"):
                    keywords = kw_result["response"].strip()
                    log(f"关键词: {keywords}")
                    if len(keywords) > 3:
                        deep = registry.call("deep_research", keywords)
                        if deep.get("success"):
                            extra_content = deep["data"][:2000]
                            log(f"深度搜索补充: {len(extra_content)} 字")

            # === STEP 3：合并入库 ===
            parts = []
            if platform_content:
                parts.append(platform_content[:3000])
            if extra_content:
                parts.append(extra_content)

            if not parts:
                send_reply(reply_target, "[Share] 搜索未获得有效内容，请截图发给我")
                return

            content = "\n\n".join(parts)
            log(f"合并内容: {len(content)} 字")

        except Exception as e:
            error_msg = f"[Share] 处理异常: {type(e).__name__}: {e}"
            log(error_msg)
            error_log = Path(__file__).parent.parent / ".ai-state" / "share_error.log"
            with open(error_log, "a", encoding="utf-8") as f:
                import traceback as _tb
                import time as _time
                f.write(f"\n[{_time.strftime('%Y-%m-%d %H:%M:%S')}] {error_msg}\n")
                _tb.print_exc(file=f)
            send_reply(reply_target, f"[Share] 处理出错: {str(e)[:200]}")
            return
    else:
        # 情况2：纯文本
        content = content_to_process[:3000]

    # LLM 提炼
    refine_result = gateway.call_azure_openai(
        "cpo",
        f"以下是用户分享的内容摘要：\n{content[:3000]}\n\n"
        f"重要约束：\n"
        f"- title 必须准确反映内容的实际主题，包含核心实体名（品牌/产品/技术名）\n"
        f"- 绝对禁止编造不存在的概念作为标题\n"
        f"- 如果信息量太少无法确定主题，title 写'待补充:' + 你能确定的部分\n"
        f"- 相关性判断要宽泛：AI、智能出行、穿戴设备、工业设计、商业模式都算 medium 以上\n\n"
        f"相关性判断原则：用户主动分享的内容默认有价值。\n"
        f"- high: 直接相关（头盔、骑行、灯光、传感器、安全装备）\n"
        f"- medium: 间接相关（AI技术、智能出行、汽车智能化、穿戴设备、工业设计、开源工具、商业模式、用户体验、新材料）\n"
        f"- 只有纯美食、纯娱乐、纯体育赛事才判 relevant=false\n"
        f"- 用户主动分享 = 用户认为有价值，尽量判 relevant=true\n\n"
        f"如果 relevant=true，按 JSON 回复：\n"
        f'{{"title": "包含具体名称的标题", "domain": "competitors或components或standards或lessons", '
        f'"tags": ["标签"], "summary": "200字摘要", "relevant": true, "confidence": "high或medium"}}\n'
        f"如果无关，回复：{{\"relevant\": false}}",
        "只输出 JSON。",
        "share_refine"
    )

    if not refine_result.get("success"):
        send_reply(reply_target, f"[Share] Content fetched but refine failed\n{content[:300]}")
        return

    try:
        resp = refine_result["response"].strip()
        resp = re.sub(r'^```json\s*', '', resp)
        resp = re.sub(r'\s*```$', '', resp)
        data = _json.loads(resp)

        if data.get("relevant", False):
            add_knowledge(
                title=data.get("title", "User Share"),
                domain=data.get("domain", "lessons"),
                content=data.get("summary", content[:500]),
                tags=data.get("tags", []) + ["user_share"],
                source=f"user_share:url",
                confidence="high",
                caller="user_share"
            )
            send_reply(reply_target, f"[OK] Saved: {data.get('title', '')}")
            send_reply(reply_target, "If you have questions about this, just ask me")
        else:
            # 用户主动分享的内容，即使 LLM 判定无关，也存入知识库备查
            add_knowledge(
                title=data.get("title", "用户分享") if isinstance(data, dict) else "用户分享",
                domain="lessons",
                content=content[:500],
                tags=["user_share", "low_relevance"],
                source="user_share:url",
                confidence="low",
                caller="user_share"
            )
            send_reply(reply_target, f"[OK] Saved (low relevance): {data.get('title', '') if isinstance(data, dict) else 'user share'}")
            send_reply(reply_target, "If you have questions about this, just ask me")
    except Exception as e:
        send_reply(reply_target, f"[Share] Parse error: {e}\n{content[:200]}")


def handle_image_message(open_id: str, image_key: str, message_id: str, reply_target: str = None) -> None:
    """处理图片消息：优先Gemini Vision，降级OCR，智能路由"""
    if reply_target is None:
        reply_target = open_id
    import traceback
    try:
        send_reply(reply_target, "🖼️ 正在识别图片...")
        image_data = download_feishu_image(image_key, message_id)
        if not image_data:
            send_reply(reply_target, "❌ 图片下载失败，请重试")
            return

        # 压缩后调用 Gemini Vision
        compressed = compress_image(image_data)
        result = get_model_gateway().call_gemini_vision("gemini_3_pro", compressed,
            "请详细描述这张图片的内容。如果图片包含文字，也请提取出来。")
        if result.get("success"):
            vision_text = result["response"]
            print(f"  [Gemini Vision] 成功: {vision_text[:100]}...")
        else:
            # 降级 OCR
            print(f"  [Gemini Vision] 失败，降级OCR")
            with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as f:
                f.write(image_data)
                tmp_path = f.name
            vision_text = f"[OCR识别结果]\n{process_image_to_text(tmp_path)}"

        # 图片只做描述和入库，不触发研发任务
        # 自动入库逻辑
        gateway = get_model_gateway()
        refine_result = gateway.call_azure_openai(
            "cpo",
            f"以下是一张用户分享图片的 AI 描述：\n{vision_text[:3000]}\n\n"
            f"判断这是否与智能骑行头盔相关。如果相关，按 JSON 回复：\n"
            f'{{"title": "20字标题", "domain": "competitors或components或standards或lessons", "tags": ["标签"], "summary": "200字摘要", "relevant": true}}\n'
            f"如果完全无关，回复：{{\"relevant\": false}}",
            "只输出 JSON。",
            "image_share_refine"
        )

        saved_title = ""
        if refine_result.get("success"):
            try:
                import re
                resp = refine_result["response"].strip()
                resp = re.sub(r'^```json\s*', '', resp)
                resp = re.sub(r'\s*```$', '', resp)
                import json as _json
                data = _json.loads(resp)
                if data.get("relevant", False):
                    from src.tools.knowledge_base import add_knowledge
                    add_knowledge(
                        title=data.get("title", "User Image"),
                        domain=data.get("domain", "lessons"),
                        content=data.get("summary", vision_text[:500]),
                        tags=data.get("tags", []) + ["user_share"],
                        source="user_share:image",
                        confidence="high",
                        caller="user_share"
                    )
                    saved_title = data.get("title", "")
                    send_reply(reply_target, f"[Image] {vision_text[:300]}\n\n[OK] Saved: {saved_title}")
                else:
                    send_reply(reply_target, f"[Image] {vision_text[:500]}")
            except Exception as e2:
                print(f"[Image] JSON解析失败: {e2}")
                send_reply(reply_target, f"[Image] {vision_text[:500]}")
        else:
            send_reply(reply_target, f"[Image] {vision_text[:500]}")

        send_reply(reply_target, "If you have questions, just ask me")
    except Exception as e:
        print(f"[Image] 处理失败: {e}")
        print(f"[Image] 详细: {traceback.format_exc()}")
        send_reply(reply_target, f"图片处理失败: {str(e)[:200]}")


def download_audio(message_id: str, file_key: str = "") -> bytes:
    """从飞书下载语音消息的音频文件"""
    token = get_tenant_access_token()
    if not token:
        print("  [语音下载失败: 无 token]")
        return b""
    headers = {"Authorization": f"Bearer {token}"}

    # 方式1：通过 file_key 下载（推荐）
    if file_key:
        url = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/resources/{file_key}"
        params = {"type": "file"}
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            print(f"  [语音下载 方式1: status={resp.status_code}, size={len(resp.content)}]")
            if resp.status_code == 200 and len(resp.content) > 100:
                return resp.content
        except Exception as e:
            print(f"  [语音下载 方式1 失败: {e}]")

    # 方式2：不带 file_key，尝试直接下载
    url2 = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/resources/file"
    params2 = {"type": "file"}
    try:
        resp2 = requests.get(url2, headers=headers, params=params2, timeout=30)
        print(f"  [语音下载 方式2: status={resp2.status_code}, size={len(resp2.content)}]")
        if resp2.status_code == 200 and len(resp2.content) > 100:
            return resp2.content
        else:
            print(f"  [语音下载 方式2 响应: {resp2.text[:300]}]")
    except Exception as e:
        print(f"  [语音下载 方式2 失败: {e}]")

    return b""


def handle_audio_message(open_id: str, message_id: str, content: str = "", reply_target: str = None) -> None:
    """处理语音消息：下载→Gemini音频理解→智能路由"""
    if reply_target is None:
        reply_target = open_id
    send_reply(reply_target, "🎤 收到语音，正在识别...")

    # 提取 file_key
    file_key = ""
    try:
        content_data = json.loads(content) if isinstance(content, str) else content
        file_key = content_data.get("file_key", "")
        print(f"  [语音 file_key: {file_key}]")
    except Exception:
        print(f"  [语音 content 解析失败: {content[:200]}]")

    audio_bytes = download_audio(message_id, file_key)
    if not audio_bytes:
        send_reply(reply_target, "❌ 语音下载失败，请重试")
        return

    print(f"  [语音下载成功: {len(audio_bytes)} bytes]")

    result = get_model_gateway().call_gemini_audio(
        "gemini_3_pro",
        audio_bytes,
        "请准确转写这段语音的内容，只输出语音中的文字内容，不要添加任何解释。如果是中文语音就输出中文。",
        "",
        "audio_transcribe"
    )

    if not result.get("success"):
        send_reply(reply_target, f"❌ 语音识别失败: {result.get('error', '未知')[:200]}")
        return

    transcribed_text = result["response"].strip()
    print(f"  [语音转写: {transcribed_text[:100]}]")
    send_reply(reply_target, f"🎤 语音识别结果：{transcribed_text}")

    if not transcribed_text:
        return

    if handle_fix_command(transcribed_text, open_id, reply_target):
        pass
    # 评价匹配：必须是单字母 或 字母+空格+理由，且必须有待评价任务
    elif (lambda _s=transcribed_text.strip().upper(): len(_s) >= 1 and _s[0] in ("A", "B", "C", "D") and (len(_s) == 1 or (len(_s) > 1 and _s[1] == " ")))() and _last_task_memory.get("memory_dir") and handle_rating(transcribed_text, open_id, reply_target):
        pass
    elif is_rd_task(transcribed_text):
        if _rd_task_running:
            send_reply(reply_target, "⏳ 上一个研发任务还在执行中，请稍后再试")
        else:
            send_reply(reply_target, "🚀 检测到研发任务，启动多Agent工作流...")
            threading.Thread(
                target=_run_rd_task_background,
                args=(transcribed_text, open_id, reply_target),
                daemon=True
            ).start()
    else:
        send_reply(reply_target, "⏳ 处理中...")
        reply = call_llm_chat(transcribed_text)
        send_reply(reply_target, reply)


def _smart_route_and_reply(text: str, open_id: str, chat_id: str, chat_type: str,
                            reply_target: str, reply_type: str):
    """非精确指令的智能路由（text 和 post 消息共用）"""
    from src.utils.conversation_memory import get_conversation_memory
    from src.utils.capability_registry import get_capabilities_for_intent, get_capabilities_summary, CAPABILITIES
    from src.utils.intent_router import classify_intent
    from src.tools.knowledge_base import search_knowledge, format_knowledge_for_prompt, KB_ROOT

    gateway = get_model_gateway()
    mem = get_conversation_memory()
    session_id = chat_id if chat_type == "group" else open_id

    mem.add_user_message(session_id, text)
    history = mem.get_history_for_prompt(session_id)

    pending_tool = mem.get_context(session_id, "pending_tool")

    if pending_tool:
        intent_result = {
            "intent": "tool_call", "tool": pending_tool,
            "needs_more_input": False, "reasoning": "用户提供了上轮等待的输入"
        }
        mem.clear_context(session_id, "pending_tool")
    else:
        caps_desc = get_capabilities_for_intent()
        intent_result = classify_intent(text, history, caps_desc, gateway)

    intent = intent_result.get("intent", "chat")
    tool = intent_result.get("tool", "none")
    needs_more = intent_result.get("needs_more_input", False)

    log(f"[Intent] {intent}, tool={tool}, needs_more={needs_more}")

    kb_entries = search_knowledge(text, limit=8)
    kb_context = format_knowledge_for_prompt(kb_entries) if kb_entries else ""

    # 知识库强匹配 override
    if kb_entries and intent != "tool_call":
        has_strong_match = any(
            "技术档案" in e.get("title", "") or
            "decision_tree" in str(e.get("tags", [])) or
            "anchor" in str(e.get("tags", [])) or
            "芯片" in e.get("title", "") or
            "AR1" in str(e.get("content", "")) or
            "AR2" in str(e.get("content", ""))
            for e in kb_entries
        )
        if has_strong_match:
            intent = "knowledge_search"
            log(f"[Intent Override] 知识库强匹配，切换到 knowledge_search")

    # === 路由执行 ===
    if intent == "tool_call" and tool == "image_generation":
        if needs_more:
            mem.set_context(session_id, "pending_tool", "image_generation")
            reply_text = intent_result.get("what_to_ask", "请发送图片描述（prompt）。")
            send_reply(reply_target, reply_text, reply_type)
            mem.add_bot_message(session_id, reply_text, "ask_input")
        else:
            send_reply(reply_target, "正在生成图片...", reply_type)
            from src.tools.tool_registry import get_tool_registry
            import base64
            img_result = get_tool_registry().call("image_generation", text)
            if img_result.get("success") and img_result.get("image_base64"):
                img_bytes = base64.b64decode(img_result["image_base64"])
                send_image_reply(reply_target, img_bytes, reply_type)
                reply_text = "图片已生成。要调整什么地方吗？"
            else:
                reply_text = f"图片生成失败: {img_result.get('error', '未知')[:200]}"
            send_reply(reply_target, reply_text, reply_type)
            mem.add_bot_message(session_id, reply_text, "image_generation")

    elif intent in ("tool_call", "knowledge_search") and (tool == "knowledge_search" or intent == "knowledge_search"):
        send_reply(reply_target, "正在查阅知识库...", reply_type)

        def _kb_bg():
            try:
                if kb_entries:
                    answer_prompt = (
                        f"基于以下知识库内容回答用户问题。引用具体数据。\n"
                        f"speculative 条目引用时标注'这是推测'。\n\n"
                        f"{kb_context[:4000]}\n\n用户问题：{text}\n\n"
                        f"回复 300-500 字，口语化，像微信聊天。"
                    )
                    answer_result = gateway.call_azure_openai("cpo", answer_prompt,
                        "你是项目合伙人。简洁回答，引用数据。", "kb_answer")
                    reply_text = answer_result.get("response", "") if answer_result.get("success") else "查询失败"
                    if not reply_text:
                        reply_text = f"知识库中暂无「{text[:20]}」相关信息。要深入研究吗？"
                else:
                    reply_text = f"知识库中暂无「{text[:20]}」相关信息。要深入研究吗？"
                    mem.set_context(session_id, "pending_research_topic", text)
                send_reply(reply_target, reply_text, reply_type)
                mem.add_bot_message(session_id, reply_text, "knowledge_search")
            except Exception as e:
                send_reply(reply_target, f"查询出错: {str(e)[:100]}", reply_type)

        threading.Thread(target=_kb_bg, daemon=True).start()

    elif intent == "research":
        if _rd_task_running:
            send_reply(reply_target, "上一个研发任务还在执行中", reply_type)
        else:
            send_reply(reply_target, "检测到研发任务，启动多 Agent 工作流...", reply_type)
            mem.add_bot_message(session_id, "启动研发任务", "research")
            threading.Thread(
                target=_run_rd_task_background,
                args=(text, open_id, reply_target, reply_type),
                daemon=True
            ).start()

    else:
        # chat 兜底
        send_reply(reply_target, "思考中...", reply_type)

        def _chat_bg():
            try:
                caps_summary = get_capabilities_summary()
                chat_prompt = (
                    f"你是智能摩托车全盔项目的 AI 合伙人。\n\n"
                    f"## 对话历史\n{history}\n\n"
                    f"## 我的能力\n{caps_summary}\n\n"
                )
                if kb_context:
                    chat_prompt += f"## 相关知识\n{kb_context[:3000]}\n\n"
                chat_prompt += f"## 用户消息\n{text}\n\n简洁专业地回答，300-500字。"

                result = gateway.call_azure_openai("cpo", chat_prompt,
                    "你是项目合伙人Leo's Agent。回复简洁有力，像面对面说话。", "smart_chat")
                reply_text = result.get("response", "抱歉，我没理解。") if result.get("success") else "服务暂时不可用"
                send_reply(reply_target, reply_text, reply_type)
                mem.add_bot_message(session_id, reply_text, "chat")

                # === Chat 洞察自动存档 ===
                # 如果用户消息包含行业动态信号词，且机器人回复包含对项目的启发，自动存入知识库
                try:
                    signal_words = ["发布", "上线", "推出", "宣布", "融资", "合作",
                                    "MCP", "skill", "plugin", "开源", "GitHub",
                                    "对我们", "启示", "启发", "借鉴", "参考"]
                    has_signal = any(w in text for w in signal_words)
                    has_insight = any(w in reply_text for w in ["对我们", "启示", "启发", "建议", "借鉴", "意味着"])

                    if has_signal and has_insight and len(reply_text) > 200:
                        from src.tools.knowledge_base import add_knowledge
                        archive_content = f"## 用户分享\n{text[:500]}\n\n## AI 分析\n{reply_text[:2000]}"
                        add_knowledge(
                            title=f"[洞察] {text[:40]}",
                            domain="lessons",
                            content=archive_content,
                            tags=["chat_insight", "auto_archive", "industry_signal"],
                            source="chat_auto_archive",
                            confidence="medium",
                            caller="user_share",
                        )
                        print(f"[Chat] 洞察已自动存档: {text[:40]}")
                except Exception as archive_err:
                    print(f"[Chat] 自动存档失败: {archive_err}")
            except Exception as e:
                send_reply(reply_target, f"回复出错: {str(e)[:100]}", reply_type)

        threading.Thread(target=_chat_bg, daemon=True).start()


def handle_message(event):
    """处理接收到的消息"""
    global client, processed_ids

    # Path 已在文件顶部 import，直接使用全局变量

    # 静默模式：飞书后台任务不播放提示音
    from src.utils.notifier import set_silent
    set_silent(True)

    # P0 优先级标记：用户实时任务
    from src.utils.task_priority import get_priority_manager
    pm = get_priority_manager()
    pm.p0_start()

    try:
        # 写入心跳
        try:
            _heartbeat_path = Path(__file__).parent.parent / ".ai-state" / "heartbeat.txt"
            _heartbeat_path.write_text(datetime.now().isoformat(), encoding="utf-8")
        except Exception:
            pass

        print(f"\n{'='*50}")
        print(f"收到消息!")

        # 正确访问事件数据
        message = event.event.message
        sender = event.event.sender

        # 过滤机器人自己发的消息
        sender_type = getattr(sender, 'sender_type', '')
        if sender_type == 'app':
            print("[Skip] 跳过机器人自己的消息")
            return

        # 消息去重
        message_id = message.message_id
        if message_id in processed_ids:
            return
        processed_ids.add(message_id)
        save_processed_id(message_id)

        log(f"收到消息! msg_id={message_id}")

        # 消息类型
        msg_type = message.message_type
        raw_content = message.content
        log(f"msg_type={msg_type}, content_len={len(raw_content) if raw_content else 0}")

        # 诊断：记录所有消息原始内容到文件
        try:
            _debug_log = Path(__file__).parent.parent / ".ai-state" / "message_debug.log"
            with open(_debug_log, "a", encoding="utf-8") as _df:
                _df.write(f"\n[{datetime.now()}] msg_type={msg_type}\n")
                _df.write(f"  raw_content={raw_content}\n")
        except Exception:
            pass

        if isinstance(raw_content, str):
            content = json.loads(raw_content) if raw_content else {}
        else:
            content = {}

        # 发送者信息
        sender_id = sender.sender_id
        open_id = sender_id.open_id
        chat_id = message.chat_id

        # === 群聊/私聊区分 ===
        chat_type = getattr(message, 'chat_type', 'p2p')  # p2p 或 group
        print(f"  Chat Type: {chat_type}")

        # 群聊中检查是否 @了机器人
        if chat_type == "group":
            mentions = getattr(message, 'mentions', [])
            if not mentions:
                # 没有任何 @，忽略群消息
                print("  [群聊] 没有 @，忽略")
                return
            print(f"  [群聊] 检测到 @，处理消息")

        # 设置回复目标
        if chat_type == "group":
            reply_target = chat_id
            reply_type = "chat_id"
        else:
            reply_target = open_id
            reply_type = "open_id"

        # 设置全局回复上下文（让其他函数的 send_reply 调用自动使用正确的目标）
        _reply_context["target"] = reply_target
        _reply_context["type"] = reply_type

        # DEBUG: 消息类型和内容预览
        print(f"  [DEBUG] msg_type={msg_type}, content_preview={str(message.content)[:200]}")

        print(f"  消息类型: {msg_type}")
        print(f"  Open ID: {open_id}")
        print(f"  Chat ID: {chat_id}")

        if msg_type == "text":
            text = content.get("text", "")

            # === 清理群聊 @mention ===
            if chat_type == "group":
                # 方式1：去掉飞书格式的 @_user_xxx 标记
                text = re.sub(r'@_user_\d+\s*', '', text).strip()
                # 方式2：去掉普通 @xxx 标记
                text = re.sub(r'@[^\s]+\s*', '', text).strip()
                print(f"  [群聊] 清理 @mention 后: {text[:50]}")

            log(f"text 路由, 长度={len(text)}, _has_url={_has_shareable_url(text)}")
            print(f"  [DEBUG] text 路由开始, 长度={len(text)}, 前50字: {text[:50]}")
            print(f"  内容: {text}")

            # 诊断：记录 text 内容
            try:
                with open(_debug_log, "a", encoding="utf-8") as _df:
                    _df.write(f"  text={text[:500] if text else 'None'}\n")
            except Exception:
                pass

            # 意图路由：审批指令优先，评价其次，学习触发，研发任务走 LangGraph，其余走 LLM
            if handle_fix_command(text, open_id, reply_target):
                pass
            elif text.strip() in ("帮助", "help", "?", "？", "能力", "你能做什么"):
                from src.utils.capability_registry import get_capabilities_summary
                send_reply(reply_target, get_capabilities_summary(), reply_type)
            # 评价匹配：必须是单字母 或 字母+空格+理由，且必须有待评价任务
            elif (lambda _s=text.strip().upper(): len(_s) >= 1 and _s[0] in ("A", "B", "C", "D") and (len(_s) == 1 or (len(_s) > 1 and _s[1] == " ")))() and _last_task_memory.get("memory_dir") and handle_rating(text, open_id, reply_target):
                pass
            elif text.strip() in ("学习", "每日学习", "daily learning"):
                # 动态获取主题数量
                _topics_path = Path(__file__).parent.parent / ".ai-state" / "knowledge" / "learning_topics.json"
                _topic_count = 10  # 默认值
                if _topics_path.exists():
                    try:
                        _topics_data = json.loads(_topics_path.read_text(encoding="utf-8"))
                        _topic_count = len(_topics_data.get("topics", []))
                    except Exception:
                        pass
                send_reply(reply_target, f"📚 正在执行每日学习（{_topic_count}个主题，预计3-5分钟）...")
                from scripts.daily_learning import run_daily_learning
                report = run_daily_learning(
                    progress_callback=lambda msg: send_reply(reply_target, msg)
                )
                send_reply(reply_target, report)
            elif text.strip() in ("重置学习", "reset learning", "重学"):
                from scripts.daily_learning import COVERED_TOPICS_FILE
                if COVERED_TOPICS_FILE.exists():
                    COVERED_TOPICS_FILE.unlink()
                send_reply(reply_target, "已重置学习覆盖记录，下次学习将重新搜索所有固定主题")
            elif text.strip() in ("深度学习", "夜间学习", "night learning"):
                send_reply(reply_target, "[NightLearn] 启动深度学习（三阶段：深化+拓展+跨界）...")
                from scripts.daily_learning import run_night_deep_learning
                def _run_night():
                    report = run_night_deep_learning(
                        progress_callback=lambda msg: send_reply(reply_target, msg)
                    )
                    send_reply(reply_target, report)
                threading.Thread(target=_run_night, daemon=True).start()
            elif text.strip() in ("对齐", "对齐报告", "alignment"):
                def _run_align():
                    try:
                        from scripts.daily_learning import generate_alignment_report
                        report = generate_alignment_report()
                        send_reply(reply_target, report)
                    except Exception as e:
                        send_reply(reply_target, f"对齐报告生成失败: {e}")
                threading.Thread(target=_run_align, daemon=True).start()
                send_reply(reply_target, "📊 正在生成对齐报告...")
            elif text.strip().startswith("关注 ") or text.strip().startswith("关注："):
                topic_text = text.strip().replace("关注 ", "").replace("关注：", "").strip()
                if topic_text:
                    _topics_path = Path(__file__).parent.parent / ".ai-state" / "knowledge" / "learning_topics.json"
                    if _topics_path.exists():
                        _data = json.loads(_topics_path.read_text(encoding="utf-8"))
                    else:
                        _data = {"version": "1.0", "topics": []}
                    _data["topics"].append({"query": topic_text, "domain": "lessons", "tags": ["用户关注"]})
                    _data["updated_at"] = datetime.now().strftime("%Y-%m-%d")
                    _topics_path.write_text(json.dumps(_data, ensure_ascii=False, indent=2), encoding="utf-8")
                    send_reply(reply_target, f"✅ 已添加学习关注：{topic_text}\n下次学习时会搜索此主题")
                else:
                    send_reply(reply_target, "请输入关注的主题，如：关注 骑行头盔 AR导航")
            elif text.strip() in ("导入文档", "导入", "import"):
                send_reply(reply_target, "📂 正在扫描文档收件箱...")
                from scripts.doc_importer import scan_and_import
                report = scan_and_import(
                    progress_callback=lambda msg: send_reply(reply_target, msg)
                )
                if report:
                    send_reply(reply_target, report)
                else:
                    send_reply(reply_target, "[Info] 收件箱为空，无文件需要处理。\n请把文件放入 .ai-state/inbox/ 目录")
            elif text.strip() in ("复盘", "团队复盘", "review"):
                send_reply(reply_target, "[Review] 启动团队复盘分析...")
                from scripts.self_evolution import run_review
                report = run_review(
                    progress_callback=lambda msg: send_reply(reply_target, msg)
                )
                send_reply(reply_target, report)
            elif text.strip() in ("长任务", "任务状态", "task status"):
                hb_file = Path(__file__).parent.parent / ".ai-state" / "long_task_heartbeat.json"
                if hb_file.exists():
                    data = json.loads(hb_file.read_text(encoding="utf-8"))
                    status = data.get("status", "unknown")
                    name = data.get("task_name", "?")
                    current = data.get("current", 0)
                    total = data.get("total", 0)
                    elapsed = data.get("elapsed_sec", 0)
                    pct = f" ({current*100//total}%)" if total > 0 else ""
                    send_reply(reply_target,
                        f"📊 长任务状态: {name}\n"
                        f"状态: {status} | 进度: {current}/{total}{pct}\n"
                        f"耗时: {elapsed//60}分钟 | 错误: {data.get('errors', 0)}", reply_type)
                else:
                    send_reply(reply_target, "当前没有运行中的长任务", reply_type)
            elif text.strip() in ("技术雷达", "平台监控", "tech radar", "platform monitor"):
                send_reply(reply_target, "📡 启动技术雷达扫描（GitHub + ClawHub），预计 5-10 分钟...", reply_type)
                def _radar():
                    try:
                        from scripts.platform_monitor import run_platform_monitor
                        report = run_platform_monitor(
                            since_days=7,
                            progress_callback=lambda msg: print(f"  {msg}"),
                            feishu_notify=lambda msg: send_reply(reply_target, msg, reply_type)
                        )
                        send_reply(reply_target, report[:3500], reply_type)
                    except Exception as e:
                        send_reply(reply_target, f"❌ 技术雷达失败: {e}", reply_type)
                threading.Thread(target=_radar, daemon=True).start()
            elif text.strip() in ("知识库", "知识库状态", "kb", "knowledge"):
                from src.tools.knowledge_base import get_knowledge_stats, KB_ROOT
                import json as _json
                stats = get_knowledge_stats()
                total = sum(stats.values())

                lines = [f"[KB] 知识库状态 ({total} 条)"]
                for domain in sorted(stats.keys()):
                    lines.append(f"  {domain}: {stats[domain]}")

                # 最近 10 条新增（带编号）
                lines.append(f"\n--- 最近新增 (发送 '删除 KB 编号' 可删除) ---")
                all_files = sorted(KB_ROOT.rglob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
                for i, f in enumerate(all_files[:10], 1):
                    try:
                        data = _json.loads(f.read_text(encoding="utf-8"))
                        title = data.get("title", f.stem)[:45]
                        domain = f.parent.name
                        source = data.get("source", "?")[:15]
                        # 用完整文件名作为短 ID（确保唯一）
                        short_id = f.stem
                        lines.append(f"  {i}. [{domain}] {title}")
                        lines.append(f"     💡 删除: 删除 KB {short_id}")
                    except Exception:
                        continue

                lines.append(f"\n发送「知识库 详细」查看覆盖率")
                send_reply(reply_target, "\n".join(lines))
            elif text.strip() in ("规则库", "检查规则", "critic rules", "rules"):
                from src.utils.critic_rules import get_rules_summary
                send_reply(reply_target, get_rules_summary(), reply_type)
            elif text.strip() in ("恢复任务", "resume", "恢复"):
                try:
                    from src.graph.router import HAS_CHECKPOINT, _checkpoint_db
                    if not HAS_CHECKPOINT:
                        send_reply(reply_target, "断点续传未启用（缺少 langgraph-checkpoint-sqlite）", reply_type)
                    else:
                        import sqlite3
                        conn = sqlite3.connect(str(_checkpoint_db))
                        try:
                            cursor = conn.execute(
                                "SELECT DISTINCT thread_id FROM checkpoints ORDER BY created_at DESC LIMIT 5"
                            )
                            threads = [row[0] for row in cursor.fetchall()]
                        except sqlite3.OperationalError:
                            threads = []
                        conn.close()

                        if not threads:
                            send_reply(reply_target, "没有可恢复的任务", reply_type)
                        else:
                            lines = ["📋 最近的任务 checkpoint：\n"]
                            for i, tid in enumerate(threads, 1):
                                lines.append(f"  {i}. {tid}")
                            lines.append(f"\n发送「恢复 task_xxx」恢复指定任务")
                            send_reply(reply_target, "\n".join(lines), reply_type)
                except Exception as e:
                    send_reply(reply_target, f"检查 checkpoint 失败: {e}", reply_type)

            elif text.strip().startswith("恢复 task_") or text.strip().startswith("resume task_"):
                task_id = text.strip().split()[-1]
                try:
                    from src.graph.router import app as langgraph_app, HAS_CHECKPOINT
                    if not HAS_CHECKPOINT:
                        send_reply(reply_target, "断点续传未启用", reply_type)
                    else:
                        send_reply(reply_target, f"🔄 正在恢复任务 {task_id}...", reply_type)

                        def _resume_bg():
                            global _rd_task_running
                            _rd_task_running = True
                            try:
                                config = {"configurable": {"thread_id": task_id}}
                                result = langgraph_app.invoke(None, config=config)

                                if result:
                                    execution = result.get("execution", {})
                                    synthesis = execution.get("synthesis_output", "")
                                    if synthesis:
                                        send_reply(reply_target, f"✅ 任务 {task_id} 恢复完成\n\n{synthesis[:3000]}", reply_type)
                                    else:
                                        send_reply(reply_target, f"✅ 任务 {task_id} 恢复完成，但无输出", reply_type)
                                else:
                                    send_reply(reply_target, f"任务 {task_id} 恢复失败：无结果", reply_type)
                            except Exception as e:
                                send_reply(reply_target, f"❌ 恢复失败: {str(e)[:300]}", reply_type)
                            finally:
                                _rd_task_running = False

                        threading.Thread(target=_resume_bg, daemon=True).start()
                except Exception as e:
                    send_reply(reply_target, f"恢复失败: {e}", reply_type)
            elif text.strip() in ("知识图谱", "kg expand", "深挖芯片", "深挖", "自主深挖"):
                send_reply(reply_target, "🔬 启动自主深挖：分析知识库薄弱方向，自动生成深挖计划...")
                def _deep():
                    try:
                        from scripts.knowledge_graph_expander import run_autonomous_deep_dive
                        report = run_autonomous_deep_dive(progress_callback=None)
                        send_reply(reply_target, f"✅ 自主深挖完成\n{report[:3500]}")
                    except Exception as e:
                        import traceback
                        send_reply(reply_target, f"❌ 自主深挖失败: {e}\n{traceback.format_exc()[:500]}")
                threading.Thread(target=_deep, daemon=True).start()
            elif text.strip() == "强制深挖":
                # 保留强制深挖，但改为调用自主深挖
                send_reply(reply_target, "⚠️ 强制执行自主深挖...")
                def _kg_force():
                    try:
                        from scripts.knowledge_graph_expander import run_autonomous_deep_dive
                        report = run_autonomous_deep_dive(progress_callback=None)
                        send_reply(reply_target, f"✅ 自主深挖完成\n{report[:3500]}")
                    except Exception as e:
                        import traceback
                        send_reply(reply_target, f"❌ 自主深挖失败: {e}\n{traceback.format_exc()[:500]}")
                threading.Thread(target=_kg_force, daemon=True).start()
            elif text.strip() in ("完整性检测", "completeness", "缺口检测"):
                send_reply(reply_target, "开始知识自完整性检测，寻找家族性缺口...")
                def _check():
                    try:
                        from scripts.knowledge_completeness_checker import run_completeness_check
                        report = run_completeness_check(progress_callback=lambda msg: send_reply(reply_target, msg))
                        send_reply(reply_target, report[:4000])
                    except Exception as e:
                        import traceback
                        send_reply(reply_target, f"完整性检测失败: {e}\n{traceback.format_exc()[:500]}")
                threading.Thread(target=_check, daemon=True).start()
            elif text.strip() in ("进化记录", "evolution", "进化"):
                from src.tools.knowledge_base import search_knowledge
                entries = search_knowledge("evolution 教训 成功模式", limit=10)
                evo_entries = [e for e in entries if "evolution" in e.get("tags", [])]

                if not evo_entries:
                    send_reply(reply_target, "📊 暂无进化记录。给研发任务打 A 或 D 评价后，系统会自动分析并记录。")
                else:
                    lines = ["📊 Agent 进化记录\n"]
                    for e in evo_entries[:8]:
                        tags = e.get("tags", [])
                        icon = "✅" if "success" in tags else "❌" if "failure" in tags else "📝"
                        lines.append(f"{icon} {e['title']}")
                        lines.append(f"   {e['content'][:150]}...\n")
                    send_reply(reply_target, "\n".join(lines))
            elif text.strip().startswith("删除 KB ") or text.strip().startswith("删除KB"):
                from src.tools.knowledge_base import KB_ROOT
                import json as _json
                import re as _re
                target_id = text.strip().replace("删除 KB ", "").replace("删除KB", "").strip()

                # 去掉可能的方括号、编号前缀等
                target_id = target_id.strip("[]()（） .")
                # 如果用户复制了整行，尝试提取 ID 部分
                if "删除 KB" in target_id or "删除KB" in target_id:
                    target_id = target_id.replace("删除 KB", "").replace("删除KB", "").strip()
                if "ID:" in target_id:
                    target_id = target_id.split("ID:")[1].split()[0].strip()
                # 提取日期格式的 ID（如 20260322_112）
                _match = _re.search(r'(\d{8}_\d{3,})', target_id)
                if _match:
                    target_id = _match.group(1)

                if not target_id:
                    send_reply(reply_target, "请提供知识条目 ID，如：删除 KB 20260322_0941")
                else:
                    # 搜索匹配的文件
                    deleted = []
                    for f in KB_ROOT.rglob("*.json"):
                        if target_id in f.stem:
                            try:
                                data = _json.loads(f.read_text(encoding="utf-8"))
                                title = data.get("title", "")[:50]
                                f.unlink()
                                deleted.append(f"[{f.parent.name}] {title}")
                            except Exception:
                                continue

                    if deleted:
                        send_reply(reply_target, f"[OK] 已删除 {len(deleted)} 条:\n" + "\n".join(deleted))
                    else:
                        send_reply(reply_target, f"[FAIL] 未找到 ID 包含 '{target_id}' 的条目")
            elif text.strip() in ("审计", "audit", "知识库审计"):
                from scripts.daily_learning import audit_knowledge_base
                audit = audit_knowledge_base()
                reply = (
                    f"[Audit] 知识库审计\n"
                    f"总量: {audit['total']} 条\n"
                    f"深度: 浅{audit['shallow']}({audit['shallow_pct']}%) | 中{audit['medium']} | 深{audit['deep']}\n"
                    f"质量: 无数据{audit['no_data']} | 重复{audit['duplicates']}\n"
                    f"待深化: {len(audit.get('shallow_entries', []))} 条"
                )
                send_reply(reply_target, reply)
            elif text.strip() in ("自动研究", "auto research"):
                from scripts.daily_learning import generate_alignment_report, auto_schedule_research
                send_reply(reply_target, "[AutoResearch] 正在基于最新对齐报告自动规划研究任务...")
                alignment = generate_alignment_report()
                report = auto_schedule_research(alignment, progress_callback=lambda msg: send_reply(reply_target, msg))
                send_reply(reply_target, report)
            elif text.strip() in ("重读文档", "reread", "重新导入"):
                import shutil
                processed = Path(".ai-state/inbox/processed")
                inbox = Path(".ai-state/inbox")

                if not processed.exists():
                    send_reply(reply_target, "📂 processed 目录为空，没有需要重读的文档")
                    return

                count = 0
                total_size_mb = 0
                for f in processed.iterdir():
                    if f.is_file() and f.suffix.lower() in ('.pptx', '.pdf', '.docx', '.xlsx', '.txt', '.md', '.csv',
                                                           '.png', '.jpg', '.jpeg', '.gif', '.webp',
                                                           '.mp3', '.wav', '.ogg', '.m4a',
                                                           '.mp4', '.mov'):
                        shutil.move(str(f), str(inbox / f.name))
                        total_size_mb += f.stat().st_size / 1024 / 1024 if (inbox / f.name).exists() else 0
                        count += 1

                if count == 0:
                    send_reply(reply_target, "📂 processed 中没有可重读的文件")
                    return

                send_reply(reply_target, f"📂 已将 {count} 个文件移回 inbox ({total_size_mb:.0f}MB)，开始重新导入...\n这些文件会用更深入的视角重新提炼。")

                # 在后台线程执行，避免阻塞
                def _reread():
                    try:
                        from scripts.doc_importer import scan_and_import
                        report = scan_and_import(progress_callback=lambda msg: send_reply(reply_target, msg))
                        if report:
                            send_reply(reply_target, report)
                        else:
                            send_reply(reply_target, "✅ 重读完成，无新知识产出")
                    except Exception as e:
                        send_reply(reply_target, f"❌ 重读失败: {e}")

                threading.Thread(target=_reread, daemon=True).start()
            elif text.strip() in ("待处理", "pending", "失败文件"):
                import json as _json
                pending_dir = Path(".ai-state/pending_imports")
                if not pending_dir.exists() or not list(pending_dir.glob("*.json")):
                    send_reply(reply_target, "📂 没有待处理的失败文件")
                    return

                lines = ["📂 待重试文件："]
                for f in pending_dir.glob("*.json"):
                    try:
                        data = _json.loads(f.read_text(encoding="utf-8"))
                        lines.append(f"  - {data.get('filename', '?')} ({data.get('size_mb', '?')}MB)")
                        lines.append(f"    失败时间: {data.get('failed_at', '?')}")
                        lines.append(f"    原因: {data.get('reason', '?')[:80]}")
                    except:
                        continue
                send_reply(reply_target, "\n".join(lines))
            elif text.strip() in ("知识库 详细", "知识库详细", "kb detail"):
                from src.tools.knowledge_base import KB_ROOT
                import json as _json

                # 按竞品统计
                competitor_entries = {}
                for f in (KB_ROOT / "competitors").glob("*.json"):
                    try:
                        data = _json.loads(f.read_text(encoding="utf-8"))
                        title = data.get("title", "")
                        content = data.get("content", "")
                        # 提取品牌关键词
                        for brand in ["Shoei", "Forcite", "Motoeye", "Cardo", "Sena", "Livall", "Lumos",
                                       "Meta", "BleeqUP", "ASMAX", "RESO", "Unit 1", "TALI", "CrossHelmet",
                                       "雷鸟", "RayNeo"]:
                            if brand.lower() in title.lower() or brand.lower() in content[:200].lower():
                                if brand not in competitor_entries:
                                    competitor_entries[brand] = []
                                competitor_entries[brand].append({
                                    "title": title[:40],
                                    "depth": len(content)
                                })
                    except Exception:
                        continue

                lines = ["[KB Detail] 竞品知识覆盖率"]

                # 按竞品展示
                for brand in sorted(competitor_entries.keys()):
                    entries = competitor_entries[brand]
                    total_depth = sum(e["depth"] for e in entries)
                    avg_depth = total_depth // len(entries) if entries else 0
                    depth_label = "浅" if avg_depth < 150 else ("中" if avg_depth < 300 else "深")
                    lines.append(f"\n  {brand} ({len(entries)} 条, 深度:{depth_label})")
                    for e in entries[:3]:
                        lines.append(f"    - {e['title']}")
                    if len(entries) > 3:
                        lines.append(f"    ... 及其他 {len(entries)-3} 条")

                # 知识深度分布
                all_depths = []
                for f in KB_ROOT.rglob("*.json"):
                    try:
                        data = _json.loads(f.read_text(encoding="utf-8"))
                        all_depths.append(len(data.get("content", "")))
                    except Exception:
                        continue

                shallow = sum(1 for d in all_depths if d < 150)
                medium_kb = sum(1 for d in all_depths if 150 <= d < 300)
                deep = sum(1 for d in all_depths if d >= 300)

                lines.append(f"\n[Depth] 深度分布:")
                lines.append(f"  浅(<150字): {shallow} 条")
                lines.append(f"  中(150-300字): {medium_kb} 条")
                lines.append(f"  深(>300字): {deep} 条")

                # 来源分布
                source_counts = {}
                for f in KB_ROOT.rglob("*.json"):
                    try:
                        data = _json.loads(f.read_text(encoding="utf-8"))
                        src = data.get("source", "unknown").split(":")[0]
                        source_counts[src] = source_counts.get(src, 0) + 1
                    except Exception:
                        continue

                lines.append(f"\n[Source] 来源分布:")
                for src, cnt in sorted(source_counts.items(), key=lambda x: -x[1]):
                    lines.append(f"  {src}: {cnt} 条")

                send_reply(reply_target, "\n".join(lines))
            elif text.strip() in ("JDM学习", "jdm学习", "供应商学习", "jdm"):
                def _run_jdm():
                    global _long_task_running
                    _long_task_running = True
                    try:
                        from scripts.tonight_jdm_learning import run_jdm_learning
                        report = run_jdm_learning(
                            progress_callback=lambda msg: send_reply(reply_target, msg)
                        )
                        send_reply(reply_target, f"[JDM] {report}")
                    except Exception as e:
                        send_reply(reply_target, f"[JDM] 学习失败: {e}")
                    finally:
                        _long_task_running = False
                threading.Thread(target=_run_jdm, daemon=True).start()
                send_reply(reply_target, "[JDM] 供应商定向学习已启动（约 33 个主题，预计 10-15 分钟）...")
            elif text.strip() in ("深度研究", "deep research", "JDM深度", "jdm深度"):
                def _run_deep():
                    global _long_task_running
                    _long_task_running = True
                    try:
                        from scripts.tonight_deep_research import run_all
                        report_path = run_all(
                            progress_callback=lambda msg: send_reply(reply_target, msg)
                        )
                        send_reply(reply_target, f"[Deep Research] 5 份深度报告已完成！\n报告路径: {report_path}\n发送「日报」查看知识库变化")
                    except Exception as e:
                        send_reply(reply_target, f"[Deep Research] 失败: {e}")
                    finally:
                        _long_task_running = False
                threading.Thread(target=_run_deep, daemon=True).start()
                send_reply(reply_target, "[Deep Research] JDM 供应商选型深度研究已启动（5 个任务，每个 8 源，预计 30-40 分钟）...")
            elif text.strip() in ("今晚研究", "tonight", "研究队列"):
                def _run_tonight():
                    global _long_task_running
                    _long_task_running = True
                    try:
                        from scripts.tonight_research_queue import run_tonight
                        report_path = run_tonight(progress_callback=lambda msg: send_reply(reply_target, msg))
                        send_reply(reply_target, f"[Tonight] All done! Report: {report_path}")
                    except Exception as e:
                        send_reply(reply_target, f"[Tonight] Failed: {e}")
                    finally:
                        _long_task_running = False
                threading.Thread(target=_run_tonight, daemon=True).start()
                send_reply(reply_target, "[Tonight] Deep research queue started (10 tasks, ETA 60-90 min)...")
            # ===== 新增：从文件读取研究任务（支持多文件） =====
            elif "参考文件：" in text or "reference file:" in text.lower():
                import re
                # 提取所有 .md 文件路径（支持单行、多行、带序号等格式）
                matches = re.findall(r'[\w./\\-]+\.md', text)
                if matches:
                    # 区分参考文件(docs/tasks/)和约束文件(docs/specs/)
                    task_files = []
                    constraint_files = []
                    for raw_path in matches:
                        # 支持相对路径和绝对路径
                        if not raw_path.startswith('/') and not raw_path.startswith('D:'):
                            full_path = str(Path(__file__).parent.parent / raw_path)
                        else:
                            full_path = raw_path

                        if Path(full_path).exists():
                            if 'docs/tasks/' in raw_path or 'docs\\tasks' in raw_path:
                                task_files.append((raw_path, full_path))
                            elif 'docs/specs/' in raw_path or 'docs\\specs' in raw_path:
                                constraint_files.append((raw_path, full_path))
                            else:
                                # 其他位置的 .md 默认作为参考文件
                                task_files.append((raw_path, full_path))
                        else:
                            send_reply(reply_target, f"[Research] 文件不存在: {raw_path}")

                    if task_files:
                        def _run_multi_files():
                            global _long_task_running
                            _long_task_running = True
                            try:
                                from scripts.tonight_deep_research import run_research_from_file
                                # 读取约束文件内容作为上下文注入
                                constraint_context = ""
                                for raw_path, full_path in constraint_files:
                                    constraint_context += f"\n\n---\n## 约束文件: {raw_path}\n\n"
                                    constraint_context += Path(full_path).read_text(encoding='utf-8')

                                # 每个参考文件独立执行
                                for idx, (raw_path, full_path) in enumerate(task_files, 1):
                                    send_reply(reply_target, f"🔍 [{idx}/{len(task_files)}] 开始研究: {Path(full_path).name}")
                                    # 如果有约束文件，作为附加参数传入
                                    report_path = run_research_from_file(
                                        full_path,
                                        progress_callback=lambda msg: send_reply(reply_target, msg),
                                        constraint_context=constraint_context if constraint_context else None
                                    )
                                    if report_path:
                                        send_reply(reply_target, f"✅ [{idx}/{len(task_files)}] 完成: {Path(full_path).name}\n报告: {report_path}")
                                    else:
                                        send_reply(reply_target, f"[Research] {Path(full_path).name} 未解析到有效任务")
                            except Exception as e:
                                send_reply(reply_target, f"[Research] 执行失败: {e}")
                            finally:
                                _long_task_running = False

                        threading.Thread(target=_run_multi_files, daemon=True).start()
                        msg = f"[Research] 启动 {len(task_files)} 个研究任务"
                        if constraint_files:
                            msg += f"，附带 {len(constraint_files)} 个约束文件"
                        send_reply(reply_target, msg)
                    elif constraint_files and not task_files:
                        send_reply(reply_target, "[Research] 仅收到约束文件，缺少研究任务文件(docs/tasks/)")
                else:
                    send_reply(reply_target, "格式错误。正确格式：参考文件：docs/tasks/xxx.md")
            # ===== End 新增 =====
            elif text.strip().startswith("研究 "):
                topic = text.strip()[3:].strip()
                if not topic:
                    send_reply(reply_target, "请输入研究主题，如：研究 歌尔JDM合作模式")
                else:
                    def _run_single():
                        global _long_task_running
                        _long_task_running = True
                        try:
                            from scripts.tonight_deep_research import deep_research_one
                            task = {
                                "id": "adhoc_" + topic[:10].replace(" ", "_"),
                                "title": topic,
                                "goal": f"深入研究「{topic}」，给出有具体数据支撑的分析报告",
                                "searches": []  # 空，让发现层自动生成
                            }
                            report = deep_research_one(task, progress_callback=lambda msg: send_reply(reply_target, msg))

                            # 生成摘要
                            from src.utils.model_gateway import get_model_gateway
                            gw = get_model_gateway()
                            summary_prompt = (
                                f"以下是一份研究报告（{len(report)}字）。\n"
                                f"请写一份 500 字以内的执行摘要，包含：\n"
                                f"1. 核心结论（一句话）\n"
                                f"2. 关键数据点（3-5 个最重要的数字）\n"
                                f"3. 行动建议\n\n"
                                f"{report[:5000]}"
                            )
                            sum_result = gw.call_azure_openai("cpo", summary_prompt, "写执行摘要。", "report_summary")

                            if sum_result.get("success"):
                                summary_text = sum_result["response"][:1500]
                            else:
                                summary_text = report[:1500]

                            send_reply(reply_target, f"[Research] Done: {topic}\n\n{summary_text}\n\nReport saved ({len(report)} chars). Send '日报' to check knowledge base.")
                        except Exception as e:
                            send_reply(reply_target, f"[Research] Failed: {e}")
                        finally:
                            _long_task_running = False
                    threading.Thread(target=_run_single, daemon=True).start()
                    send_reply(reply_target, f"[Research] Starting: {topic}...")
            elif text.strip().startswith("设置目标"):
                # 用户可以通过飞书更新产品目标
                goal_text = text.strip()[4:].strip()
                if len(goal_text) < 20:
                    send_reply(reply_target, "请提供更详细的目标描述，至少 20 字。例如：\n设置目标 V1 目标2026年Q4量产，核心卖点HUD导航+4K黑匣子+SOS，面向高端摩旅用户，定价5000-8000元", reply_type)
                else:
                    goal_file = Path(__file__).parent.parent / ".ai-state" / "product_goal.json"
                    import json as _json
                    goal_data = {
                        "goal": goal_text,
                        "updated": datetime.now().isoformat(),
                        "updated_by": open_id
                    }
                    goal_file.write_text(_json.dumps(goal_data, ensure_ascii=False, indent=2), encoding="utf-8")
                    send_reply(reply_target, f"[Goal] Product goal updated. All agent planning will align with this goal.\n\n{goal_text}", reply_type)
            elif text.strip() in ("token", "tokens", "Token统计", "用量"):
                # Token 使用统计
                from src.utils.token_usage_tracker import TokenUsageTracker
                tracker = TokenUsageTracker()
                stats = tracker.get_stats(days=7)

                lines = ["[Token] 7天用量统计"]
                lines.append(f"总调用: {stats['total_calls']} 次")
                lines.append(f"成功: {stats['success_calls']} | 失败: {stats['failed_calls']}")
                lines.append(f"总Tokens: {stats['total_tokens']:,}")
                lines.append(f"  输入: {stats['total_prompt_tokens']:,}")
                lines.append(f"  输出: {stats['total_completion_tokens']:,}")
                lines.append(f"估算成本: ${stats['total_cost']:.4f}")

                # 今日统计
                today = tracker.get_today_stats()
                lines.append(f"\n[今日] {today['total_calls']} 次调用")
                lines.append(f"Tokens: {today['total_tokens']:,} | 成本: ${today['total_cost']:.4f}")

                # 模型排名
                ranking = tracker.get_model_ranking(days=7)
                if ranking:
                    lines.append("\n[模型 Top3]")
                    for i, m in enumerate(ranking[:3], 1):
                        lines.append(f"  {i}. {m['model']}: {m['calls']}次, {m['tokens']:,}tokens")

                send_reply(reply_target, "\n".join(lines))
            elif text.strip() in ("导入记录", "inbox", "导入历史"):
                from src.tools.knowledge_base import KB_ROOT
                import json as _json

                lines = ["[DocImport] 文档导入记录"]

                # 从知识库中找所有 doc_import 来源的条目
                doc_entries = []
                for f in KB_ROOT.rglob("*.json"):
                    try:
                        data = _json.loads(f.read_text(encoding="utf-8"))
                        source = data.get("source", "")
                        if "doc_import" in source:
                            doc_entries.append({
                                "title": data.get("title", ""),
                                "domain": f.parent.name,
                                "source": source,
                                "tags": data.get("tags", []),
                                "confidence": data.get("confidence", ""),
                                "content_preview": data.get("content", "")[:100]
                            })
                    except Exception:
                        continue

                if not doc_entries:
                    lines.append("暂无文档导入记录")
                else:
                    lines.append(f"共导入 {len(doc_entries)} 条知识\n")
                    for i, e in enumerate(doc_entries, 1):
                        filename = e["source"].replace("doc_import:", "")
                        lines.append(f"{i}. [{e['domain']}] {e['title']}")
                        lines.append(f"   文件: {filename}")
                        lines.append(f"   标签: {', '.join(e['tags'][:4])}")
                        lines.append(f"   摘要: {e['content_preview']}...")
                        lines.append("")

                # 检查 processed 目录
                processed_dir = Path(__file__).parent.parent / ".ai-state" / "inbox" / "processed"
                if processed_dir.exists():
                    processed_files = list(processed_dir.iterdir())
                    lines.append(f"--- inbox/processed: {len(processed_files)} 个已处理文件 ---")

                send_reply(reply_target, "\n".join(lines))
            elif text.strip() in ("日报", "学习日报", "daily report"):
                from scripts.daily_learning import generate_daily_report
                report = generate_daily_report()
                send_reply(reply_target, report)
            elif text.strip().startswith("@dev ") or text.strip().startswith("@dev："):
                dev_request = text.strip().replace("@dev ", "").replace("@dev：", "").strip()
                if dev_request:
                    send_reply(reply_target, f"🛠️ 分析开发需求中...\n需求：{dev_request[:100]}")
                    from scripts.dev_assistant import generate_dev_proposal
                    result = generate_dev_proposal(dev_request)
                    if result.get("success"):
                        send_reply(reply_target, result["message"])
                    else:
                        send_reply(reply_target, f"❌ 生成提案失败：{result.get('error', '未知错误')[:300]}")
                else:
                    send_reply(reply_target, "请描述开发需求，如：@dev 给 state_merge 增加日志")
            elif _has_shareable_url(text):
                log(f"触发分享流程, text={text[:80]}")
                # 裸链接自动触发分享入库（后台线程）
                threading.Thread(
                    target=handle_share_content,
                    args=(open_id,),
                    kwargs={"text": text, "reply_target": reply_target},
                    daemon=True
                ).start()
            # === 结构化文档快速通道（PRD/清单/表格不走多Agent）===
            from scripts.feishu_handlers.structured_doc import try_structured_doc_fast_track
            if try_structured_doc_fast_track(text, reply_target, reply_type, open_id, chat_id, send_reply):
                return

            elif is_rd_task(text):
                if _rd_task_running:
                    send_reply(reply_target, "⏳ 上一个研发任务还在执行中，请稍后再试")
                else:
                    send_reply(reply_target, "🚀 检测到研发任务，启动多Agent工作流...")
                    threading.Thread(
                        target=_run_rd_task_background,
                        args=(text, open_id, reply_target),
                        daemon=True
                    ).start()
            else:
                # ============================================
                # 非精确指令 -> LLM 意图识别 + 智能路由
                # ============================================
                from src.utils.conversation_memory import get_conversation_memory
                from src.utils.capability_registry import get_capabilities_for_intent, get_capabilities_summary, CAPABILITIES
                from src.utils.intent_router import classify_intent
                from src.tools.knowledge_base import search_knowledge, format_knowledge_for_prompt, KB_ROOT

                gateway = get_model_gateway()
                mem = get_conversation_memory()
                session_id = chat_id if chat_type == "group" else open_id

                # 记录用户消息
                mem.add_user_message(session_id, text)

                # 获取对话历史
                history = mem.get_history_for_prompt(session_id)

                # 检查是否有等待中的上下文（如：上轮说要出图，等 prompt）
                pending_tool = mem.get_context(session_id, "pending_tool")

                if pending_tool:
                    # 有等待中的工具调用 -> 直接执行
                    intent_result = {
                        "intent": "tool_call",
                        "tool": pending_tool,
                        "needs_more_input": False,
                        "reasoning": "用户提供了上轮等待的输入"
                    }
                    mem.clear_context(session_id, "pending_tool")
                else:
                    # LLM 意图识别
                    caps_desc = get_capabilities_for_intent()
                    intent_result = classify_intent(text, history, caps_desc, gateway)

                intent = intent_result.get("intent", "chat")
                tool = intent_result.get("tool", "none")
                needs_more = intent_result.get("needs_more_input", False)

                log(f"[Intent] {intent}, tool={tool}, needs_more={needs_more}")

                # === 所有对话都先搜知识库（不只是 knowledge_search intent） ===
                kb_entries = search_knowledge(text, limit=8)
                kb_context = format_knowledge_for_prompt(kb_entries) if kb_entries else ""

                # 如果知识库有强匹配（得分高的条目），强制走知识库回答
                if kb_entries and intent != "tool_call":
                    # 检查知识库是否有直接相关的技术档案
                    has_strong_match = any(
                        "技术档案" in e.get("title", "") or
                        "decision_tree" in str(e.get("tags", [])) or
                        "芯片" in e.get("title", "") or
                        "AR1" in str(e.get("content", "")) or
                        "AR2" in str(e.get("content", ""))
                        for e in kb_entries
                    )
                    if has_strong_match:
                        intent = "knowledge_search"  # 强制切到知识库回答
                        log(f"[Intent Override] 知识库有强匹配，切换到 knowledge_search")

                # === 路由执行 ===

                if intent == "tool_call" and tool == "image_generation":
                    if needs_more:
                        mem.set_context(session_id, "pending_tool", "image_generation")
                        reply_text = intent_result.get("what_to_ask", "请发送你想生成的图片描述（prompt），越详细越好。")
                        send_reply(reply_target, reply_text, reply_type)
                        mem.add_bot_message(session_id, reply_text, "ask_input")
                    else:
                        send_reply(reply_target, "正在生成图片...", reply_type)
                        from src.tools.tool_registry import get_tool_registry
                        _registry = get_tool_registry()
                        img_result = _registry.call("image_generation", text)
                        if img_result.get("success") and img_result.get("image_base64"):
                            # 发送图片到飞书
                            import base64
                            img_bytes = base64.b64decode(img_result["image_base64"])
                            send_image_reply(reply_target, img_bytes, reply_type)
                            reply_text = "图片已生成。要调整什么地方吗？"
                            send_reply(reply_target, reply_text, reply_type)
                            mem.set_context(session_id, "last_image_prompt", text)
                        else:
                            reply_text = f"图片生成失败: {img_result.get('error', '未知错误')[:200]}"
                            send_reply(reply_target, reply_text, reply_type)
                        mem.add_bot_message(session_id, reply_text, "image_generation")

                elif intent == "tool_call" and tool == "knowledge_search":
                    # 先快速回复，避免 WebSocket 超时
                    send_reply(reply_target, "正在查阅知识库...", reply_type)

                    def _kb_answer_bg():
                        try:
                            if kb_entries:
                                answer_prompt = (
                                    f"基于以下知识库内容，回答用户的问题。必须引用具体数据（型号、参数、价格）。\n"
                                    f"如果知识库信息不够回答，诚实说明并建议深入研究。\n"
                                    f"知识库中标记了 [speculative] 的条目是推测/假想内容，引用时必须明确标注'这是推测，非官方确认'。\n"
                                    f"优先引用有明确来源的确认数据，推测数据放最后或不引用。\n\n"
                                    f"{kb_context[:4000]}\n\n"
                                    f"用户问题：{text}\n\n"
                                    f"## 回复风格（必须遵守）\n"
                                    f"- 先用 3-5 句话给结论，像合伙人面对面说话\n"
                                    f"- 引用 2-3 个关键数据点（型号、参数、价格），用【】标注来源\n"
                                    f"- 最后一句问'要展开吗？'或给下一步建议\n"
                                    f"- 总长度 300-500 字，绝不超过 800 字\n"
                                    f"- 不用论文格式（禁止 ##、---、大段编号列表）\n"
                                    f"- 短句口语化，像微信聊天不像文档\n"
                                    f"- 如果信息不够回答，诚实说'这块知识库还不够，要不要我深入研究一下？'\n"
                                )
                                answer_result = gateway.call_azure_openai("cpo", answer_prompt,
                                    "你是项目合伙人，简洁回答，300-500字以内，像微信对话。引用知识库数据，不知道就说不知道。", "kb_answer")

                                if answer_result.get("success") and answer_result.get("response"):
                                    reply_text = answer_result["response"]
                                else:
                                    # 降级：直接返回知识库原文摘要
                                    error = answer_result.get("error", "未知")
                                    log(f"[KB Answer] LLM 调用失败: {error[:200]}")
                                    fallback_lines = ["知识库相关信息：\n"]
                                    for e in kb_entries[:5]:
                                        fallback_lines.append(f"**{e.get('title', '')}**")
                                        fallback_lines.append(f"{e.get('content', '')[:300]}\n")
                                    reply_text = "\n".join(fallback_lines)
                            else:
                                reply_text = f"知识库中暂无关于「{text[:20]}」的信息。要我发起一次深度研究吗？"
                                mem.set_context(session_id, "pending_research_topic", text)
                            send_reply(reply_target, reply_text, reply_type)
                            mem.add_bot_message(session_id, reply_text, "knowledge_search")
                        except Exception as e:
                            log(f"[KB Answer] 异常: {e}")
                            send_reply(reply_target, f"查询过程出错: {str(e)[:100]}", reply_type)

                    threading.Thread(target=_kb_answer_bg, daemon=True).start()

                elif intent == "knowledge_search":
                    # Intent override 强制切换的知识库回答
                    send_reply(reply_target, "正在查阅知识库...", reply_type)

                    def _kb_answer_override_bg():
                        try:
                            if kb_entries:
                                answer_prompt = (
                                    f"基于以下知识库内容，回答用户的问题。必须引用具体数据（型号、参数、价格）。\n"
                                    f"如果知识库信息不够回答，诚实说明并建议深入研究。\n"
                                    f"知识库中标记了 [speculative] 的条目是推测/假想内容，引用时必须明确标注'这是推测，非官方确认'。\n"
                                    f"优先引用有明确来源的确认数据，推测数据放最后或不引用。\n\n"
                                    f"{kb_context[:4000]}\n\n"
                                    f"用户问题：{text}\n\n"
                                    f"## 回复风格（必须遵守）\n"
                                    f"- 先用 3-5 句话给结论，像合伙人面对面说话\n"
                                    f"- 引用 2-3 个关键数据点（型号、参数、价格），用【】标注来源\n"
                                    f"- 最后一句问'要展开吗？'或给下一步建议\n"
                                    f"- 总长度 300-500 字，绝不超过 800 字\n"
                                    f"- 不用论文格式（禁止 ##、---、大段编号列表）\n"
                                    f"- 短句口语化，像微信聊天不像文档\n"
                                    f"- 如果信息不够回答，诚实说'这块知识库还不够，要不要我深入研究一下？'\n"
                                )
                                answer_result = gateway.call_azure_openai("cpo", answer_prompt,
                                    "你是项目合伙人，简洁回答，300-500字以内，像微信对话。引用知识库数据，不知道就说不知道。", "kb_answer")

                                if answer_result.get("success") and answer_result.get("response"):
                                    reply_text = answer_result["response"]
                                else:
                                    # 降级：直接返回知识库原文摘要
                                    error = answer_result.get("error", "未知")
                                    log(f"[KB Answer Override] LLM 调用失败: {error[:200]}")
                                    fallback_lines = ["知识库相关信息：\n"]
                                    for e in kb_entries[:5]:
                                        fallback_lines.append(f"**{e.get('title', '')}**")
                                        fallback_lines.append(f"{e.get('content', '')[:300]}\n")
                                    reply_text = "\n".join(fallback_lines)
                            else:
                                reply_text = f"知识库中暂无关于「{text[:20]}」的信息。要我发起一次深度研究吗？"
                                mem.set_context(session_id, "pending_research_topic", text)
                            send_reply(reply_target, reply_text, reply_type)
                            mem.add_bot_message(session_id, reply_text, "knowledge_search")
                        except Exception as e:
                            log(f"[KB Answer Override] 异常: {e}")
                            send_reply(reply_target, f"查询过程出错: {str(e)[:100]}", reply_type)

                    threading.Thread(target=_kb_answer_override_bg, daemon=True).start()

                elif intent == "research":
                    # 触发研发任务（LangGraph）
                    send_reply(reply_target, "检测到研发任务，启动多 Agent 工作流...", reply_type)
                    mem.add_bot_message(session_id, "启动研发任务", "research")
                    # 调用已有的 LangGraph 逻辑
                    if _rd_task_running:
                        send_reply(reply_target, "上一个研发任务还在执行中，请稍后再试", reply_type)
                    else:
                        threading.Thread(
                            target=_run_rd_task_background,
                            args=(text, open_id, reply_target, reply_type),
                            daemon=True
                        ).start()

                elif intent == "chat":
                    # 智能对话：带知识库 + 能力清单 + 对话历史
                    # 先快速回复，避免 WebSocket 超时
                    send_reply(reply_target, "思考中...", reply_type)

                    def _chat_bg():
                        try:
                            # 读取产品锚点
                            product_anchor = ""
                            for f in KB_ROOT.rglob("*.json"):
                                try:
                                    data = json.loads(f.read_text(encoding="utf-8"))
                                    tags = data.get("tags", [])
                                    if "internal" in tags and ("prd" in tags or "product_definition" in tags):
                                        product_anchor = data.get("content", "")[:800]
                                        break
                                except:
                                    continue

                            caps_summary = get_capabilities_summary()

                            chat_prompt = (
                                f"你是智能摩托车全盔项目的 AI 合伙人。\n\n"
                                f"## 重要规则\n"
                                f"- 如果知识库中有相关技术档案，必须引用其中的具体数据回答，不要凭空编造\n"
                                f"- 如果知识库没有相关信息，诚实说'我知识库里暂时没有这个信息，要我深入研究一下吗？'\n"
                                f"- 不要编造不存在的数据或参数\n\n"
                                f"## 对话历史\n{history}\n\n"
                                f"## 我的能力\n{caps_summary}\n\n"
                            )
                            if product_anchor:
                                chat_prompt += f"## 产品定义\n{product_anchor[:500]}\n\n"
                            if kb_context:
                                chat_prompt += f"## 相关知识（来自项目知识库）\n{kb_context[:3000]}\n\n"
                            chat_prompt += f"## 用户消息\n{text}\n\n"
                            chat_prompt += (
                                f"## 回复要求\n"
                                f"- 像合伙人之间对话，简洁有力，不像客服\n"
                                f"- 如果用户问你能不能做什么，对照能力清单诚实回答\n"
                                f"- 如果知识库有相关内容，引用具体数据，标注来源\n"
                                f"- 如果你觉得用户的问题值得深入研究，主动建议\n"
                                f"- 保持上下文连贯，不要每次都像新对话\n"
                            )

                            result = gateway.call_azure_openai("cpo", chat_prompt,
                                "你是项目合伙人Leo's Agent。回复简洁有力，300-500字，像面对面说话。", "smart_chat")

                            if result.get("success") and result.get("response"):
                                reply_text = result["response"]
                            else:
                                error = result.get("error", "未知")
                                log(f"[Chat] LLM 调用失败: {error[:200]}")
                                reply_text = "服务暂时不可用，请稍后重试"

                            send_reply(reply_target, reply_text, reply_type)
                            mem.add_bot_message(session_id, reply_text, "chat")

                            # === Chat 洞察自动存档 ===
                            # 如果用户消息包含行业动态信号词，且机器人回复包含对项目的启发，自动存入知识库
                            try:
                                signal_words = ["发布", "上线", "推出", "宣布", "融资", "合作",
                                                "MCP", "skill", "plugin", "开源", "GitHub",
                                                "对我们", "启示", "启发", "借鉴", "参考"]
                                has_signal = any(w in text for w in signal_words)
                                has_insight = any(w in reply_text for w in ["对我们", "启示", "启发", "建议", "借鉴", "意味着"])

                                if has_signal and has_insight and len(reply_text) > 200:
                                    from src.tools.knowledge_base import add_knowledge
                                    archive_content = f"## 用户分享\n{text[:500]}\n\n## AI 分析\n{reply_text[:2000]}"
                                    add_knowledge(
                                        title=f"[洞察] {text[:40]}",
                                        domain="lessons",
                                        content=archive_content,
                                        tags=["chat_insight", "auto_archive", "industry_signal"],
                                        source="chat_auto_archive",
                                        confidence="medium",
                                        caller="user_share",
                                    )
                                    print(f"[Chat] 洞察已自动存档: {text[:40]}")
                            except Exception as archive_err:
                                print(f"[Chat] 自动存档失败: {archive_err}")
                        except Exception as e:
                            log(f"[Chat] 异常: {e}")
                            send_reply(reply_target, f"回复过程出错: {str(e)[:100]}", reply_type)

                    threading.Thread(target=_chat_bg, daemon=True).start()

                else:
                    # 兜底
                    reply_text = "收到，但我不太确定你想让我做什么。可以说得更具体一些吗？"
                    send_reply(reply_target, reply_text, reply_type)
                    mem.add_bot_message(session_id, reply_text, "fallback")

        elif msg_type == "post":
            # 富文本消息，提取纯文字内容和图片
            try:
                post_content = content.get("content", [])
                text_parts = []
                image_keys = []
                for block in post_content:
                    for element in block:
                        if element.get("tag") == "text":
                            text_parts.append(element.get("text", ""))
                        elif element.get("tag") == "img":
                            img_key = element.get("image_key", "")
                            if img_key:
                                image_keys.append(img_key)
                text = "".join(text_parts).strip()
                if not text:
                    text = str(content)  # fallback
            except Exception:
                text = str(content)
                image_keys = []

            print(f"  [DEBUG] post 路由开始, 长度={len(text)}, 图片数={len(image_keys)}, 前50字: {text[:50]}")
            print(f"  内容(post): {text[:200]}...")

            # === 新增：长文章走知识库导入 ===
            has_url = _has_shareable_url(text)
            if _is_likely_article(text, has_url):
                print(f"  [路由] 检测到长文章 ({len(text)}字)，导入知识库")
                send_reply(reply_target, f"📖 检测到长文章 ({len(text)}字)，正在学习...")

                def _import_article_bg():
                    try:
                        # 提取并理解图片
                        image_descriptions = []
                        if image_keys:
                            image_descriptions = _extract_images_from_post(
                                post_content, message.message_id, open_id
                            )

                        # 导入知识库
                        result = _import_article_to_kb(text, image_descriptions, open_id)
                        send_reply(reply_target, f"✅ {result}")
                    except Exception as e:
                        send_reply(reply_target, f"❌ 导入失败: {e}")

                threading.Thread(target=_import_article_bg, daemon=True).start()
            elif handle_fix_command(text, open_id, reply_target):
                pass
            # 评价匹配：必须是单字母 或 字母+空格+理由，且必须有待评价任务
            elif (lambda _s=text.strip().upper(): len(_s) >= 1 and _s[0] in ("A", "B", "C", "D") and (len(_s) == 1 or (len(_s) > 1 and _s[1] == " ")))() and _last_task_memory.get("memory_dir") and handle_rating(text, open_id, reply_target):
                pass
            elif text.strip().startswith("@dev ") or text.strip().startswith("@dev："):
                dev_request = text.strip().replace("@dev ", "").replace("@dev：", "").strip()
                if dev_request:
                    send_reply(reply_target, f"🛠️ 分析开发需求中...\n需求：{dev_request[:100]}")
                    from scripts.dev_assistant import generate_dev_proposal
                    result = generate_dev_proposal(dev_request)
                    if result.get("success"):
                        send_reply(reply_target, result["message"])
                    else:
                        send_reply(reply_target, f"❌ 生成提案失败：{result.get('error', '未知错误')[:300]}")
                else:
                    send_reply(reply_target, "请描述开发需求，如：@dev 给 state_merge 增加日志")
            elif has_url:
                log(f"触发分享流程, text={text[:80]}")
                # 裸链接自动触发分享入库（后台线程）
                threading.Thread(
                    target=handle_share_content,
                    args=(open_id,),
                    kwargs={"text": text, "reply_target": reply_target},
                    daemon=True
                ).start()
            # === 结构化文档快速通道（PRD/清单/表格不走多Agent）===
            from scripts.feishu_handlers.structured_doc import try_structured_doc_fast_track
            if try_structured_doc_fast_track(text, reply_target, reply_type, open_id, chat_id, send_reply):
                return

            elif is_rd_task(text):
                if _rd_task_running:
                    send_reply(reply_target, "⏳ 上一个研发任务还在执行中，请稍后再试")
                else:
                    send_reply(reply_target, "🚀 检测到研发任务，启动多Agent工作流...")
                    threading.Thread(
                        target=_run_rd_task_background,
                        args=(text, open_id, reply_target),
                        daemon=True
                    ).start()
            else:
                _smart_route_and_reply(text, open_id, chat_id, chat_type, reply_target, reply_type)

        elif msg_type == "audio":
            print(f"  语音消息")
            print(f"  消息内容原始: {raw_content}")
            handle_audio_message(open_id, message.message_id, raw_content, reply_target)

        elif msg_type == "image":
            image_key = content.get("image_key", "")
            handle_image_message(open_id, image_key, message.message_id, reply_target)

        else:
            print(f"  其他类型消息: {msg_type}")
            send_reply(reply_target, f"收到{msg_type}类型消息")

    except Exception as e:
        log(f"handle_message 外层异常: {type(e).__name__}: {e}")
        import traceback
        log(traceback.format_exc())
        try:
            send_reply(reply_target, "处理失败，请稍后重试")
        except Exception:
            pass
    finally:
        # P0 任务结束
        pm.p0_end()
        print(f"{'='*50}\n")


def send_reply(target_id: str = None, text: str = "", id_type: str = None) -> bool:
    """发送回复消息

    如果不传参数，使用 _reply_context 中的默认值（群聊时自动回复到群里）

    Args:
        target_id: open_id（私聊）或 chat_id（群聊），None 时使用上下文
        text: 回复内容
        id_type: "open_id" 或 "chat_id"，None 时使用上下文

    Returns:
        bool: 是否发送成功
    """
    # 使用上下文默认值
    if target_id is None:
        target_id = _reply_context.get("target")
    if id_type is None:
        id_type = _reply_context.get("type", "open_id")

    if not target_id:
        print("  [回复失败: 无目标ID]")
        return False

    token = get_tenant_access_token()
    if not token:
        print("  [回复失败: 无法获取token]")
        return False

    url = f"https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type={id_type}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    data = {
        "receive_id": target_id,
        "msg_type": "text",
        "content": json.dumps({"text": text})
    }
    try:
        response = requests.post(url, headers=headers, json=data, timeout=10)
        result = response.json()
        if result.get("code") == 0:
            print(f"  [回复成功: {id_type}]")
            return True
        else:
            print(f"  [回复失败: {result}]")
            return False
    except Exception as e:
        print(f"  [回复异常: {e}]")
        return False


def send_image_reply(target_id: str, image_bytes: bytes, id_type: str = "open_id") -> None:
    """通过飞书发送图片消息"""
    import base64 as b64
    token = get_tenant_access_token()
    if not token:
        return
    # 先上传图片到飞书
    upload_url = "https://open.feishu.cn/open-apis/im/v1/images"
    headers = {"Authorization": f"Bearer {token}"}
    files = {"image": ("design.png", image_bytes, "image/png")}
    data = {"image_type": "message"}
    try:
        resp = requests.post(upload_url, headers=headers, files=files, data=data, timeout=30)
        result = resp.json()
        if result.get("code") != 0:
            print(f"  [图片上传失败: {result}]")
            return
        image_key = result["data"]["image_key"]
        # 发送图片消息
        msg_url = f"https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type={id_type}"
        msg_data = {
            "receive_id": target_id,
            "msg_type": "image",
            "content": json.dumps({"image_key": image_key})
        }
        msg_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        requests.post(msg_url, headers=msg_headers, json=msg_data, timeout=10)
        print(f"  [图片发送成功: {id_type}]")
    except Exception as e:
        print(f"  [图片发送异常: {e}]")


def main():
    global client

    # 启动时自动重新生成哈希快照，避免配置变更后忘记更新
    try:
        from scripts.regenerate_hashes import regenerate
        regenerate()
    except Exception as e:
        print(f"[WARN] 哈希重建失败: {e}")

    print("=" * 50)
    print("飞书长连接客户端 (SDK版)")
    print("=" * 50)
    print(f"App ID: {APP_ID}")
    print()

    # 创建客户端
    client = lark.Client.builder() \
        .app_id(APP_ID) \
        .app_secret(APP_SECRET) \
        .log_level(lark.LogLevel.INFO) \
        .build()

    # 构建事件处理器
    event_handler = lark.EventDispatcherHandler.builder(
        "", ""
    ).register_p2_im_message_receive_v1(handle_message) \
     .build()

    print("正在启动长连接...")
    print("现在可以在飞书中发送消息给机器人测试")
    print("按 Ctrl+C 停止")
    print("=" * 50 + "\n")

    # 初始化心跳
    _heartbeat_path = Path(__file__).parent.parent / ".ai-state" / "heartbeat.txt"
    _heartbeat_path.parent.mkdir(parents=True, exist_ok=True)
    _heartbeat_path.write_text(datetime.now().isoformat(), encoding="utf-8")
    print("[Heartbeat] 心跳已初始化")

    # 后台线程定期刷新心跳（每 30 秒）
    def _heartbeat_loop():
        while True:
            try:
                _heartbeat_path.write_text(datetime.now().isoformat(), encoding="utf-8")
            except Exception:
                pass
            time.sleep(30)
    threading.Thread(target=_heartbeat_loop, daemon=True).start()

    # 启动每日定时学习
    try:
        from scripts.daily_learning import start_daily_scheduler
        start_daily_scheduler(interval_hours=0.5, feishu_notify=None)
    except Exception as e:
        print(f"[DailyLearning] 定时器启动失败: {e}")

    # 启动 watchdog 子线程
    try:
        from scripts.watchdog import send_webhook, check_heartbeat_alive
        def _watchdog_thread():
            import time as _time
            consecutive_failures = 0
            while True:
                _time.sleep(60)
                try:
                    # 夜间 1:00-5:00 跳过心跳检查（系统可能在深度学习）
                    current_hour = datetime.now().hour
                    if 1 <= current_hour < 5:
                        continue  # 跳过检查，直接进入下一轮

                    # 长任务期间跳过心跳检查
                    if _long_task_running:
                        consecutive_failures = 0
                        continue

                    alive = check_heartbeat_alive()
                    if alive:
                        if consecutive_failures > 0:
                            send_webhook("服务恢复", f"主服务已恢复，之前异常 {consecutive_failures} 次")
                        consecutive_failures = 0
                    else:
                        consecutive_failures += 1
                        if consecutive_failures == 1:
                            send_webhook("心跳异常", "主服务心跳超时，可能假死")
                except Exception as e:
                    print(f"[Watchdog] 检查异常: {e}")
        threading.Thread(target=_watchdog_thread, daemon=True).start()
        print("[Watchdog] 内置监控已启动（60s 间隔）")
    except Exception as e:
        print(f"[Watchdog] 启动失败: {e}")

    # 启动时扫描文档收件箱
    try:
        from scripts.doc_importer import scan_and_import
        def _inbox_scan():
            report = scan_and_import()
            if report:
                print(report)
        threading.Thread(target=_inbox_scan, daemon=True).start()
        print("[DocImporter] 收件箱扫描已启动")
    except Exception as e:
        print(f"[DocImporter] 扫描失败: {e}")

    # 启动热更新文件监听
    try:
        import importlib
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
        print(f"[HotReload] watchdog 导入成功: {Observer}")

        SAFE_RELOAD_MODULES = {
            "intent_router": "src.utils.intent_router",
            "capability_registry": "src.utils.capability_registry",
            "critic_rules": "src.utils.critic_rules",
            "conversation_memory": "src.utils.conversation_memory",
        }

        class HotReloadHandler(FileSystemEventHandler):
            def __init__(self):
                self._last_reload = {}

            def on_modified(self, event):
                if not event.src_path.endswith(".py"):
                    return
                module_stem = Path(event.src_path).stem
                if module_stem not in SAFE_RELOAD_MODULES:
                    return

                # 防抖：同一文件 2 秒内不重复 reload
                import time as _time
                now = _time.time()
                last = self._last_reload.get(module_stem, 0)
                if now - last < 2:
                    return
                self._last_reload[module_stem] = now

                full_module = SAFE_RELOAD_MODULES[module_stem]
                try:
                    if full_module in sys.modules:
                        importlib.reload(sys.modules[full_module])
                        log(f"[HotReload] ✅ {module_stem} 已热更新")
                    else:
                        log(f"[HotReload] {module_stem} 未加载，跳过")
                except Exception as e:
                    log(f"[HotReload] ❌ {module_stem} 热更新失败: {e}")

        watch_path = str(Path(__file__).parent.parent / "src" / "utils")
        observer = Observer()
        observer.schedule(HotReloadHandler(), watch_path, recursive=False)
        observer.daemon = True
        observer.start()
        print(f"[HotReload] 文件监听已启动: {watch_path}")
        print(f"[HotReload] 安全模块: {list(SAFE_RELOAD_MODULES.keys())}")
    except ImportError as e:
        print(f"[HotReload] watchdog 未安装或导入失败: {e}")
    except Exception as e:
        print(f"[HotReload] 启动失败: {e}")

    # 启动健康监控（每 6 小时自检）
    try:
        def _health_monitor_loop():
            import time as _time
            _time.sleep(3600)  # 首次延迟 1 小时
            while True:
                _time.sleep(6 * 3600)  # 每 6 小时
                try:
                    from scripts.self_heal import run_self_heal_cycle
                    run_self_heal_cycle()
                except Exception as e:
                    print(f"[HealthMonitor] 自检异常: {e}")
        threading.Thread(target=_health_monitor_loop, daemon=True).start()
        print("[HealthMonitor] 健康监控已启动（6h 间隔）")
    except Exception as e:
        print(f"[HealthMonitor] 启动失败: {e}")

    # 启动时扫描未处理的 handoff
    try:
        from scripts.handoff_processor import scan_unprocessed, execute_handoff
        _pending_handoffs = scan_unprocessed()
        if _pending_handoffs:
            print(f"[Handoff] 发现 {len(_pending_handoffs)} 个未处理")
            for _hf in _pending_handoffs:
                try:
                    result = execute_handoff(_hf)
                    print(f"  [Handoff] {_hf.name}: {result.get('success', False)}")
                except Exception as _he:
                    print(f"  [Handoff] {_hf.name} 执行失败: {_he}")
    except Exception as _e:
        print(f"[Handoff] 扫描失败: {_e}")

    # 启动长连接
    cli = lark.ws.Client(APP_ID, APP_SECRET, event_handler=event_handler)
    cli.start()


if __name__ == "__main__":
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8")
    main()