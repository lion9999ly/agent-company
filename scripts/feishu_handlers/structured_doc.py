# -*- coding: utf-8 -*-
"""
@description: 结构化文档快速通道 - PRD/清单/表格类需求不走多Agent
@dependencies: model_gateway, knowledge_base, openpyxl
@last_modified: 2026-03-27

功能:
- 检测清单/表格/PRD类需求关键词
- 单次LLM调用生成JSON结构
- 导出Excel并发送到飞书
- 降级发送树形摘要文本
"""

import json
import re
import os
import tempfile
import threading
import yaml
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

# ===== Bug C fix: 清洗 LLM 输出中的控制字符 =====
def _clean_json_text(text: str) -> str:
    """清洗 JSON 文本中的非法控制字符，保留正常换行"""
    if not text:
        return text
    # 先提取 JSON 数组/对象部分
    start = text.find('[')
    end = text.rfind(']')
    if start == -1 or end == -1:
        start = text.find('{')
        end = text.rfind('}')
    if start != -1 and end != -1 and end > start:
        text = text[start:end + 1]

    # 处理 JSON 字符串值内部的非法控制字符
    # 简单有效的方式：直接替换所有控制字符为空格（保留 \t \n \r）
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', ' ', text)

    return cleaned
# ===== End Bug C fix =====

# ===== A2: Anchor 读取逻辑 =====
def _load_anchor():
    """加载产品需求锚点文件，返回合并后的模块列表和配置"""
    anchor_path = Path(__file__).resolve().parent.parent.parent / '.ai-state' / 'product_spec_anchor.yaml'
    if not anchor_path.exists():
        print("[Anchor] 未找到 product_spec_anchor.yaml，跳过锚点加载")
        return None

    try:
        with open(anchor_path, 'r', encoding='utf-8') as f:
            anchor = yaml.safe_load(f)
        print(f"[Anchor] 已加载锚点文件: "
              f"HUD {len(anchor.get('hud_modules', []))} 模块, "
              f"App {len(anchor.get('app_modules', []))} 模块, "
              f"跨端 {len(anchor.get('cross_cutting', []))} 模块")
        return anchor
    except Exception as e:
        print(f"[Anchor] 加载失败: {e}")
        return None


def _merge_anchor_with_prompt(anchor: dict, prompt_modules: list) -> list:
    """
    Layer 1: Anchor 唯一驱动——Prompt 关键词只补充已有模块，不创建新模块。

    anchor 为模块定义的唯一来源。prompt_modules 中的模块名如果不在 anchor 中，
    会尝试归一化映射和模糊匹配，匹配不上的直接丢弃。

    Returns:
        Anchor 原始模块列表（数量 = Anchor 模块数，不会增加）
    """
    anchor_names = set()

    # 从 anchor 提取所有模块名
    for section in ['hud_modules', 'app_modules', 'cross_cutting']:
        for mod in anchor.get(section, []):
            name = mod.get('name', '')
            if name:
                anchor_names.add(name)

    # 获取归一化映射
    normalize_map = anchor.get('module_normalize', {}) if anchor else {}

    # Prompt 中的模块名处理
    prompt_set = set(prompt_modules)
    absorbed = []
    dropped = []

    for kw in prompt_set:
        # 1. 先查 normalize_map
        normalized = normalize_map.get(kw, kw)

        if normalized in anchor_names:
            # 已在 Anchor 中（归一化后），跳过
            absorbed.append(f"{kw} → {normalized}")
            continue

        # 2. 模糊匹配：关键词是否是某个 Anchor 模块的子串
        matched = False
        for anchor_name in anchor_names:
            if kw in anchor_name or anchor_name in kw:
                absorbed.append(f"{kw} → {anchor_name} (fuzzy)")
                matched = True
                break

        if not matched:
            # 3. 不创建新模块，丢弃并警告
            dropped.append(kw)

    if absorbed:
        print(f"[Anchor] Prompt 关键词吸收: {len(absorbed)} 个")
        for a in absorbed[:5]:
            print(f"  ✓ {a}")
        if len(absorbed) > 5:
            print(f"  ...还有 {len(absorbed) - 5} 个")

    if dropped:
        print(f"[Anchor] ⚠️ Prompt 关键词丢弃（未匹配到 Anchor 模块）: {dropped}")

    # 返回 Anchor 原始模块列表，不增加
    print(f"[Anchor] 最终模块数: {len(anchor_names)} 个（Anchor 定义）")
    return list(anchor_names)


def _get_anchor_sub_features(anchor: dict, module_name: str) -> str:
    """
    获取 anchor 中某个模块的 sub_features 和 existing_l2，
    拼接成字符串注入到 _gen_one 的 prompt 中。
    B3: 强化提醒语气
    B4: 拆解子模块过滤 existing_l2 避免撞 4096
    """
    # B4: 检测是否是拆解子模块
    is_split = '-' in module_name
    parent_name = module_name.rsplit('-', 1)[0] if is_split else module_name
    suffix = module_name.rsplit('-', 1)[1] if is_split else ''

    for section in ['hud_modules', 'app_modules', 'cross_cutting']:
        for mod in anchor.get(section, []):
            if mod.get('name') == parent_name or mod.get('name') == module_name:
                parts = []

                sub = mod.get('sub_features', [])
                if sub:
                    # B3: 更强的指令语气
                    parts.append("【⚠️ 以下功能点必须全部出现在输出中，缺一不可 ⚠️】")
                    for i, s in enumerate(sub, 1):
                        parts.append(f'{i}. {s}')
                    parts.append(f"\n以上 {len(sub)} 个功能点中的每一个，至少对应输出 1 条 L2 或 L3 功能。如果某个功能点未出现在你的输出中，视为不合格。")

                existing = mod.get('existing_l2', [])
                if existing:
                    if is_split:
                        # B4: 拆解子模块：只保留相关的 existing_l2（最多 5 条）
                        suffix_keywords = {
                            '核心功能': ['核心', '功能', '规格', '展示', '显示', '控制'],
                            '交互与状态': ['交互', '状态', '反馈', 'HUD', '语音', '按键'],
                            '异常与边界': ['异常', '降级', '故障', '恢复', '边界', '兜底', '低电'],
                        }
                        keywords = suffix_keywords.get(suffix, [])
                        filtered = [e for e in existing if any(kw in e for kw in keywords)]
                        if not filtered:
                            filtered = existing[:5]  # 兜底取前5
                        parts.append(f"\n【相关L2维度（仅 {suffix} 相关）】")
                        parts.append(', '.join(filtered[:5]))
                    else:
                        parts.append("\n【已有L2维度（保留并深化，但优先级低于上述必须功能点）】")
                        parts.append(', '.join(existing))

                notes = mod.get('notes', '')
                if notes:
                    parts.append(f"\n【设计备注】{notes}")

                ref = mod.get('reference', '')
                if ref:
                    parts.append(f"\n【参考】{ref}")

                return '\n'.join(parts)

    return ''


def _get_anchor_config(anchor: dict) -> dict:
    """提取 anchor 中的配置项"""
    config = {
        'hud_columns': anchor.get('hud_table_columns', []),
        'normalize_map': anchor.get('module_normalize', {}),
        'separation_rules': anchor.get('separation_rules', []),
        'languages': [],
    }
    # 提取多语言配置
    for mod in anchor.get('cross_cutting', []):
        if mod.get('name') == '多语言支持':
            config['languages'] = mod.get('languages', [])
            break
    return config

# 全局 anchor 缓存
_ANCHOR_CACHE = None

def get_cached_anchor():
    """获取缓存的 anchor，避免重复加载"""
    global _ANCHOR_CACHE
    if _ANCHOR_CACHE is None:
        _ANCHOR_CACHE = _load_anchor()
    return _ANCHOR_CACHE
# ===== End A2 =====

# ===== A1: 模块归属映射（基于 anchor） =====
def _build_module_placement(anchor: dict) -> dict:
    """
    根据 anchor 的 section 分类，决定每个模块写入 HUD表 还是 App表。
    返回 {module_name: 'hud' | 'app'}
    """
    placement = {}

    # HUD 端模块
    for mod in anchor.get('hud_modules', []):
        name = mod.get('name', '')
        if name:
            placement[name] = 'hud'

    # App 端模块
    for mod in anchor.get('app_modules', []):
        name = mod.get('name', '')
        if name:
            placement[name] = 'app'

    # 跨端模块：根据性质分配
    cross_to_hud = {'AI功能', '语音交互', '视觉交互', '多模态交互', '实体按键交互', '氛围灯交互'}
    cross_to_app = {'多语言支持', '手机系统权限管理', '部件识别与兼容性管理'}

    for mod in anchor.get('cross_cutting', []):
        name = mod.get('name', '')
        if name:
            if name in cross_to_hud:
                placement[name] = 'hud'
            elif name in cross_to_app:
                placement[name] = 'app'
            else:
                # 默认：如果名字含 Tab/App/手机/商城/社区 → app，否则 hud
                if any(kw in name for kw in ['Tab', 'App', '手机', '商城', '社区', '通知', '用户成就', '用户学习', '新手', '身份', '权限']):
                    placement[name] = 'app'
                else:
                    placement[name] = 'hud'

    hud_count = sum(1 for v in placement.values() if v == 'hud')
    app_count = sum(1 for v in placement.values() if v == 'app')
    print(f"[Placement] HUD: {hud_count} 模块, App: {app_count} 模块")

    return placement
# ===== End A1 =====

# ===== A2: 全量归一化（含旧版继承数据） =====
def _normalize_all_rows(rows: list, normalize_map: dict, anchor: dict = None) -> list:
    """对所有行的 module 字段做归一化，然后合并同名模块去重"""
    from collections import defaultdict
    import re

    # Step 1: 归一化 module 名
    for row in rows:
        module = row.get('module', '')
        if module in normalize_map:
            row['module'] = normalize_map[module]

    # Step 2: 按归一化后的 module 分组去重
    groups = defaultdict(list)
    for row in rows:
        groups[row.get('module', 'unknown')].append(row)

    deduped = []

    for module_name, group_rows in groups.items():
        # 按 name (L3功能名) 去重
        seen = {}
        for row in group_rows:
            name = str(row.get('name', ''))
            if name in seen:
                # 保留验收标准更长的
                old_acc = str(seen[name].get('acceptance', ''))
                new_acc = str(row.get('acceptance', ''))
                if len(new_acc) > len(old_acc):
                    seen[name] = row
            else:
                seen[name] = row

        deduped_list = list(seen.values())

        # 修复 2: 去掉硬截断，只去重。超 80 行警告
        if len(deduped_list) > 80:
            print(f"  [Normalize] ⚠️ {module_name}: {len(deduped_list)} 条，内容较多，建议人工审视")

        # 不再截断，直接添加
        deduped.extend(deduped_list)

    original = len(rows)
    final = len(deduped)
    if original != final:
        print(f"[NormalizeAll] {original} → {final} 条 (去重 {original - final})")

    return deduped
# ===== End A2 =====

# ===== 强制归一化函数 — 确保写入时名称正确 =====
def _force_normalize(rows: list, normalize_map: dict) -> list:
    """强制重命名 module 字段，使用模糊匹配（去除空格）"""
    # 构建清洁的映射表（strip + 统一空格）
    clean_map = {}
    for k, v in normalize_map.items():
        clean_key = str(k).strip().replace(' ', '')
        clean_map[clean_key] = str(v).strip()

    renamed_count = 0
    for row in rows:
        # 检查所有可能包含模块名的字段
        for key in ['module', 'L1功能', 'l1', 'name']:
            val = str(row.get(key, '')).strip()
            clean_val = val.replace(' ', '')
            if clean_val and clean_val in clean_map:
                new_val = clean_map[clean_val]
                if new_val != val:
                    row[key] = new_val
                    renamed_count += 1

    if renamed_count:
        print(f"[ForceNormalize] 重命名 {renamed_count} 个字段")
    return rows

# ===== A3: HUD 新 4 列回填 =====
def _backfill_hud_columns(hud_rows: list) -> list:
    """为缺少 HUD 新 4 列的行填充默认值"""
    filled_count = 0

    for row in hud_rows:
        changed = False

        # 显示输出（视觉）
        if not row.get('visual_output') or len(str(row.get('visual_output', ''))) < 3:
            interaction = str(row.get('interaction', ''))
            name = str(row.get('name', ''))
            if 'HUD' in interaction or 'HUD' in name:
                row['visual_output'] = 'HUD卡片/图标显示'
            elif '语音' in interaction:
                row['visual_output'] = '无视觉输出（纯语音）'
            elif '按键' in interaction:
                row['visual_output'] = 'HUD操作反馈'
            elif '灯' in interaction or '灯' in name:
                row['visual_output'] = '灯效反馈'
            else:
                row['visual_output'] = 'HUD状态提示'
            changed = True

        # 显示优先级
        if not row.get('display_priority') or len(str(row.get('display_priority', ''))) < 2:
            priority = str(row.get('priority', 'P2'))
            priority_map = {'P0': 'critical', 'P1': 'high', 'P2': 'medium', 'P3': 'low'}
            row['display_priority'] = priority_map.get(priority, 'medium')
            changed = True

        # 异常与降级
        if not row.get('degradation') or len(str(row.get('degradation', ''))) < 3:
            row['degradation'] = '功能不可用时隐藏对应卡片，不影响其他功能'
            changed = True

        # 显示时长
        if not row.get('display_duration') or len(str(row.get('display_duration', ''))) < 2:
            name = str(row.get('name', ''))
            if any(kw in name for kw in ['预警', '告警', '提醒', '通知', '来电']):
                row['display_duration'] = 'event_5s'
            elif any(kw in name for kw in ['速度', '导航', '电量', '状态']):
                row['display_duration'] = 'permanent'
            else:
                row['display_duration'] = 'user_dismiss'
            changed = True

        if changed:
            filled_count += 1

    print(f"[Backfill] HUD新列回填: {filled_count}/{len(hud_rows)} 行")
    return hud_rows
# ===== End A3 =====

# ===== 修复 1: 版本信息继承上一版 =====
def _get_previous_version_info(prd_versions_dir: str = None) -> dict:
    """读取上一版的版本号和功能统计，用于递增和 changelog"""
    if prd_versions_dir is None:
        prd_versions_dir = str(EXPORT_DIR.parent / "prd_versions")

    if not os.path.exists(prd_versions_dir):
        return {}

    # 找最新的版本快照目录（按时间戳排序）
    try:
        snapshots = sorted(os.listdir(prd_versions_dir), reverse=True)
    except Exception:
        return {}

    if not snapshots:
        return {}

    latest = snapshots[0]  # 如 20260329_010055

    # 尝试读取上一版的版本信息
    version_info_path = os.path.join(prd_versions_dir, latest, 'version_info.json')
    if os.path.exists(version_info_path):
        try:
            with open(version_info_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass

    # 兜底：从快照目录名推断时间，版本号从 V1.0 开始
    return {
        'version': 'V1.0',
        'snapshot': latest,
        'total_features': 0,
    }
# ===== End 修复 1 =====

# ===== 修复 4: 跨 Sheet 一致性引擎 =====
def _cross_sheet_consistency_check(feature_rows: list, sheets_data: dict) -> dict:
    """
    扫描所有功能条目，检查是否需要在其他 Sheet 中有对应条目。
    返回需要补生成的条目列表。
    """
    gaps = {
        'voice': [],   # 缺少的语音指令
        'light': [],   # 缺少的灯效定义
        'test': [],    # 缺少的测试用例
    }

    # 已有的语音指令关键词
    existing_voice = set()
    for v in sheets_data.get('voice', []):
        key = str(v.get('user_says', v.get('user_say', v.get('用户说法', ''))))
        if key:
            existing_voice.add(key[:10])

    # 已有的灯效场景
    existing_light = set()
    for l in sheets_data.get('light', []):
        key = str(l.get('trigger', l.get('触发场景', '')))
        if key:
            existing_light.add(key[:10])

    # 已有的测试用例覆盖的功能名
    existing_test_features = set()
    for t in sheets_data.get('test_cases', []):
        key = str(t.get('feature_name', t.get('feature', t.get('功能名', ''))))
        if key:
            existing_test_features.add(key[:10])

    for row in feature_rows:
        name = str(row.get('name', ''))
        desc = str(row.get('description', ''))
        interaction = str(row.get('interaction', ''))
        module = str(row.get('module', ''))

        # 检查语音指令覆盖
        if any(kw in desc or kw in interaction for kw in ['语音控制', '语音', '声控', '语音唤醒']):
            if not any(name[:6] in v for v in existing_voice):
                gaps['voice'].append({
                    'feature_name': name,
                    'module': module,
                    'context': desc[:50]
                })

        # 检查灯效覆盖
        if any(kw in desc or kw in interaction for kw in ['灯光', '灯效', '闪烁', '常亮']):
            if not any(name[:6] in l for l in existing_light):
                gaps['light'].append({
                    'feature_name': name,
                    'module': module,
                    'context': desc[:50]
                })

        # 检查测试用例覆盖（每个 P0 功能至少应有测试用例）
        priority = str(row.get('priority', ''))
        if priority == 'P0' and not any(name[:6] in t for t in existing_test_features):
            gaps['test'].append({
                'feature_name': name,
                'module': module,
            })

    # 汇报
    for sheet_name, gap_list in gaps.items():
        if gap_list:
            print(f"  [Consistency] {sheet_name}: 发现 {len(gap_list)} 个功能缺少对应条目")

    return gaps
# ===== End 修复 4 =====

PARALLEL_WORKERS = 8  # 并行生成线程数（从4提升到8以提速）

# 导出目录(项目固定路径,避免 tempfile 权限问题)
EXPORT_DIR = Path(__file__).resolve().parent.parent.parent / ".ai-state" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# === 关键词检测 ===
STRUCTURED_DOC_KEYWORDS = ["清单", "表格", "PRD", "Excel", "excel", "列表", "功能列表"]
FULL_SPEC_KEYWORDS = ["规格书", "状态场景", "语音指令", "按键映射", "灯效", "导航场景", "6个Sheet", "完整PRD", "6Sheet", "7Sheet", "用户故事", "测试用例", "页面映射", "开发任务", "全部Sheet"]
MIN_TEXT_LENGTH = 50

# ===== Fix 3: 模块名归一化映射 =====
MODULE_NAME_NORMALIZE = {
    # App 端归一化
    'App-我的': '我的Tab',
    '我的首页': '我的Tab',
    '我的首页与个人中心': '我的Tab',
    '我的首页与账户中心': '我的Tab',
    '我的': '我的Tab',
    'App-社区': '社区',
    '社区Tab': '社区',
    'App-商城': '商城',
    'App-设备': '设备Tab',
    '设备连接与管理': '设备Tab',
    '设备控制与显示设置': '设备Tab',
    '电量续航与固件维护': '设备Tab',
    '设备 Tab': '设备Tab',  # 注意空格
    '设备管理': '设备Tab',
    '身份认证': '我的Tab',
    '身份认证与用户信息采集': '我的Tab',
    # HUD 端归一化
    '简易': '导航',
    '简易路线': '导航',
    '简易导航': '导航',
    '简易导航模式': '导航',
    '路线': '导航',
    '显示队友位置': '组队',
    'Ai语音助手': 'AI语音助手',
    'ai语音助手': 'AI语音助手',
    # 其他归一化
    '高光时刻异常与素材完整性': '相册Tab',
    '高光结果异常提示': '相册Tab',
}

def _normalize_module_name(name: str) -> str:
    """归一化模块名，统一变体"""
    return MODULE_NAME_NORMALIZE.get(name, name)
# ===== End Fix 3 =====

# ===== Layer 3: 模块密度上限配置 =====
# 按复杂度分级：核心复杂模块 60 行，标准模块 35 行，简单模块 25 行
MODULE_DENSITY_LIMITS = {
    # 核心复杂模块：60 行
    "导航": 60, "组队": 60, "设备Tab": 60, "我的Tab": 60,
    "手机系统权限管理": 50,

    # 标准模块：35 行
    "场景模式": 35, "社区": 35, "商城": 35, "摄像状态": 35,
    "信息中岛": 35, "多模态交互": 35, "AI语音助手": 35,
    "主动安全预警提示": 40, "设备互联": 35, "设备配对流程": 35,
    "SOS与紧急救援": 35, "部件识别与兼容性管理": 35,
    "通知中心": 35, "AI Tab": 35, "相册Tab": 35,
    "多语言支持": 35, "AI功能": 35, "恢复出厂设置": 35,
    "佩戴检测与电源管理": 35, "生命体征与疲劳监测": 35,

    # 简单模块：25 行
    "开机动画": 25, "速度": 25, "来电": 25, "消息": 25,
    "音乐": 25, "胎温胎压": 25,
    "氛围灯交互": 25,
    "视觉交互": 25, "语音交互": 25, "实体按键交互": 25,
    "设备状态": 25,
}

DEFAULT_DENSITY_LIMIT = 35
# ===== End Layer 3 Config =====


def is_structured_doc_request(text: str) -> bool:
    """判断是否为结构化文档需求"""
    if len(text) < MIN_TEXT_LENGTH:
        return False
    return any(kw in text for kw in STRUCTURED_DOC_KEYWORDS)


def is_full_spec_request(text: str) -> bool:
    """判断是否需要完整规格书(6 Sheet)"""
    if len(text) < MIN_TEXT_LENGTH:
        return False
    return any(kw in text for kw in FULL_SPEC_KEYWORDS)


def _extract_l1_from_user_text(text: str) -> List[Dict]:
    """从用户文本中提取一级功能列表"""
    features = []

    # HUD 功能(拆分"简易"和"路线"为独立项)
    hud_names = [
        "导航", "来电", "音乐", "消息", "Ai语音助手", "AI语音助手",
        "简易",   # 简易模式/极简显示
        "路线",   # 路线规划/路线概览
        "主动安全预警提示",  # 匹配完整名称优先
        "组队", "摄像状态", "胎温胎压",
        "开机动画", "速度", "设备状态", "显示队友位置"
    ]
    for name in hud_names:
        if name in text:
            features.append({"name": name, "module": "头盔HUD"})

    # App Tab(4 个)- 独立匹配
    app_tabs = ["设备", "社区", "商城", "我的"]
    for tab in app_tabs:
        # 明确的 App-设备 格式
        if f"App-{tab}" in text or f"App的{tab}" in text or f"APP-{tab}" in text:
            features.append({"name": f"App-{tab}", "module": f"App-{tab}"})
        elif tab in text and ("App" in text or "APP" in text or "app" in text):
            # 检查是否有独立的 tab(不在其他词中)
            if tab == "设备":
                # 如果"设备"出现次数 > "设备状态"出现次数,说明有独立的"设备"
                device_count = text.count("设备")
                device_status_count = text.count("设备状态")
                if device_count > device_status_count:
                    features.append({"name": f"App-{tab}", "module": f"App-{tab}"})
            else:
                features.append({"name": f"App-{tab}", "module": f"App-{tab}"})

    # 额外模块(AI 相关拆细)
    extras = [
        ("AI功能", "AI功能"),
        ("语音交互", "AI功能"),
        ("视觉交互", "AI功能"),      # ADAS、追踪
        ("多模态交互", "AI功能"),   # 陀螺仪、心率
        ("实体按键交互", "头盔交互"),
        ("氛围灯交互", "头盔交互"),
        ("身份认证", "系统"),
        ("用户学习", "系统"),
        ("设备互联", "系统"),
        ("产品介绍", "系统"),
        ("设备配对流程", "系统"),
    ]
    for name, module in extras:
        if name in text:
            features.append({"name": name, "module": module})

    # 额外匹配:用户可能分开写"简易"和"路线"
    if "简易" in text and not any(f["name"] == "简易" for f in features):
        features.append({"name": "简易显示模式", "module": "头盔HUD"})
    if "路线" in text and not any(f["name"] == "路线" for f in features):
        features.append({"name": "路线规划与预览", "module": "头盔HUD"})

    # 去重
    seen = set()
    unique_features = []
    for f in features:
        key = f["name"]
        if key not in seen:
            seen.add(key)
            unique_features.append(f)

    # 兜底:如果没有提取到,使用默认
    if not unique_features:
        unique_features = [{"name": "完整功能", "module": "全部"}]

    return unique_features


# === 功能ID生成 ===
def _generate_ids(items: List[Dict]) -> List[Dict]:
    """为功能清单生成层级式功能ID"""
    MODULE_PREFIX = {
        "头盔HUD": "HUD", "导航": "HUD", "来电": "HUD", "音乐": "HUD",
        "消息": "HUD", "AI语音助手": "HUD", "Ai语音助手": "HUD",
        "简易": "HUD", "路线": "HUD", "主动安全预警提示": "HUD",
        "组队": "HUD", "摄像状态": "HUD", "胎温胎压": "HUD",
        "开机动画": "HUD", "速度": "HUD", "设备状态": "HUD",
        "显示队友位置": "HUD", "HUD": "HUD",
        "App-设备": "DEV", "App-社区": "COM", "App-商城": "MAL", "App-我的": "MY",
        "AI功能": "AI", "语音交互": "AI", "视觉交互": "AI", "多模态交互": "AI",
        "实体按键交互": "BTN", "氛围灯交互": "LED",
        "身份认证": "AUTH", "用户学习": "LRN", "设备互联": "IOT", "产品介绍": "INTRO",
        "设备配对流程": "PAIR",
        "系统": "SYS", "头盔交互": "BTN",
    }

    l1_counter = {}
    l2_counter = {}
    l3_counter = {}
    current_l1_id = ""
    current_l2_id = ""

    for item in items:
        level = item.get("level", "")
        module = item.get("module", "")
        prefix = MODULE_PREFIX.get(module, "OTH")

        if level == "L1":
            l1_counter[prefix] = l1_counter.get(prefix, 0) + 1
            current_l1_id = f"{prefix}-{l1_counter[prefix]:03d}"
            item["_id"] = current_l1_id
        elif level == "L2":
            l2_counter[current_l1_id] = l2_counter.get(current_l1_id, 0) + 1
            current_l2_id = f"{current_l1_id}-{l2_counter[current_l1_id]:02d}"
            item["_id"] = current_l2_id
        elif level == "L3":
            l3_counter[current_l2_id] = l3_counter.get(current_l2_id, 0) + 1
            item["_id"] = f"{current_l2_id}-{l3_counter[current_l2_id]:02d}"

    return items


# === 模块名统一 ===
def _normalize_module_names(items: List[Dict]) -> List[Dict]:
    """统一模块名(大小写/变体合并)"""
    name_map = {
        "Ai语音助手": "AI语音助手",
        "ai语音助手": "AI语音助手",
        "头盔HUD": "HUD",
    }
    for item in items:
        m = item.get("module", "")
        if m in name_map:
            item["module"] = name_map[m]
    return items


# === Excel Sheet写入 ===
def _write_sheet(ws, items: List[Dict]):
    """写入单个 Sheet"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    headers = ["功能ID", "L1功能", "L2功能", "L3功能", "优先级",
               "交互方式", "描述", "验收标准", "关联功能", "备注"]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 优先级颜色
    priority_fills = {
        "P0": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "P1": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "P2": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        "P3": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
    }
    priority_fonts = {
        "P0": Font(bold=True, color="FFFFFF"),
        "P1": Font(bold=True, color="FFFFFF"),
        "P2": Font(color="FFFFFF"),
        "P3": Font(color="FFFFFF"),
    }

    # L1 行样式
    l1_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    l1_font = Font(bold=True, size=11)
    l2_font = Font(bold=True, size=10)

    for row_idx, item in enumerate(items, 2):
        level = item.get("level", "")
        name = item.get("name", "")

        # 防止 list/dict 写入 Excel 的辅助函数
        def _safe_val(key):
            v = item.get(key, "")
            if isinstance(v, (list, dict)):
                return ", ".join(str(x) for x in v) if isinstance(v, list) else str(v)
            elif v is None:
                return ""
            return str(v) if v else ""

        # 功能ID
        ws.cell(row=row_idx, column=1, value=_safe_val("_id"))

        # L1/L2/L3 分列填写
        if level == "L1":
            ws.cell(row=row_idx, column=2, value=name)
        elif level == "L2":
            ws.cell(row=row_idx, column=3, value=name)
        elif level == "L3":
            ws.cell(row=row_idx, column=4, value=name)

        # 其余列（统一转 string）
        ws.cell(row=row_idx, column=5, value=_safe_val("priority"))
        ws.cell(row=row_idx, column=6, value=_safe_val("interaction"))
        ws.cell(row=row_idx, column=7, value=_safe_val("description"))
        ws.cell(row=row_idx, column=8, value=_safe_val("acceptance"))
        ws.cell(row=row_idx, column=9, value=_safe_val("dependencies"))
        ws.cell(row=row_idx, column=10, value=_safe_val("note"))

        # 样式
        for col in range(1, 11):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        if level == "L1":
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).fill = l1_fill
                ws.cell(row=row_idx, column=col).font = l1_font
        elif level == "L2":
            ws.cell(row=row_idx, column=3).font = l2_font

        # 优先级颜色
        p = item.get("priority", "").upper()
        if p in priority_fills:
            pc = ws.cell(row=row_idx, column=5)
            pc.fill = priority_fills[p]
            pc.font = priority_fonts[p]
            pc.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    widths = [16, 20, 24, 28, 8, 16, 36, 36, 20, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def _write_hud_sheet(ws, items: List[Dict]):
    """写入 HUD Sheet（12列，含 HUD 专用字段）"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # HUD 专用表头：增加 4 列
    headers = ["功能ID", "L1功能", "L2功能", "L3功能", "优先级",
               "交互方式", "描述", "验收标准", "关联功能", "备注",
               "显示输出（视觉）", "显示优先级", "异常与降级", "显示时长"]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 优先级颜色
    priority_fills = {
        "P0": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "P1": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "P2": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        "P3": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
    }
    priority_fonts = {
        "P0": Font(bold=True, color="FFFFFF"),
        "P1": Font(bold=True, color="FFFFFF"),
        "P2": Font(color="FFFFFF"),
        "P3": Font(color="FFFFFF"),
    }

    # L1 行样式
    l1_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    l1_font = Font(bold=True, size=11)
    l2_font = Font(bold=True, size=10)

    for row_idx, item in enumerate(items, 2):
        level = item.get("level", "")
        name = item.get("name", "")

        # 防止 list/dict 写入 Excel 的辅助函数
        def _safe_val(key):
            v = item.get(key, "")
            if isinstance(v, (list, dict)):
                return ", ".join(str(x) for x in v) if isinstance(v, list) else str(v)
            elif v is None:
                return ""
            return str(v) if v else ""

        # 功能ID
        ws.cell(row=row_idx, column=1, value=_safe_val("_id"))

        # L1/L2/L3 分列填写
        if level == "L1":
            ws.cell(row=row_idx, column=2, value=name)
        elif level == "L2":
            ws.cell(row=row_idx, column=3, value=name)
        elif level == "L3":
            ws.cell(row=row_idx, column=4, value=name)

        # 基础列（1-10）
        ws.cell(row=row_idx, column=5, value=_safe_val("priority"))
        ws.cell(row=row_idx, column=6, value=_safe_val("interaction"))
        ws.cell(row=row_idx, column=7, value=_safe_val("description"))
        ws.cell(row=row_idx, column=8, value=_safe_val("acceptance"))
        ws.cell(row=row_idx, column=9, value=_safe_val("dependencies"))
        ws.cell(row=row_idx, column=10, value=_safe_val("note"))

        # HUD 专用列（11-14）
        ws.cell(row=row_idx, column=11, value=_safe_val("visual_output"))
        ws.cell(row=row_idx, column=12, value=_safe_val("display_priority"))
        ws.cell(row=row_idx, column=13, value=_safe_val("degradation"))
        ws.cell(row=row_idx, column=14, value=_safe_val("display_duration"))

        # 样式
        for col in range(1, 15):  # 14列
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        if level == "L1":
            for col in range(1, 15):
                ws.cell(row=row_idx, column=col).fill = l1_fill
                ws.cell(row=row_idx, column=col).font = l1_font
        elif level == "L2":
            ws.cell(row=row_idx, column=3).font = l2_font

        # 优先级颜色
        p = item.get("priority", "").upper()
        if p in priority_fills:
            pc = ws.cell(row=row_idx, column=5)
            pc.fill = priority_fills[p]
            pc.font = priority_fonts[p]
            pc.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽（14列）
    widths = [16, 20, 24, 28, 8, 16, 36, 36, 20, 16, 24, 12, 20, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# === Mermaid 代码清洗函数 ===
def _clean_mermaid_code(code: str) -> str:
    """清洗 Mermaid 代码，修复常见语法问题"""
    import re

    if not code:
        return 'graph TD\n    A[空流程图] --> B[请重新生成]'

    # 移除 markdown 代码块标记
    code = code.strip()
    if code.startswith('```'):
        code = code.split('\n', 1)[1] if '\n' in code else code
    code = code.replace('```mermaid', '').replace('```', '').strip()

    # 确保以 graph 开头
    if not code.startswith('graph') and not code.startswith('flowchart'):
        code = 'graph TD\n' + code

    # 替换中文标点
    code = code.replace('（', '(').replace('）', ')')
    code = code.replace('【', '[').replace('】', ']')
    code = code.replace('：', ':').replace('；', ';')
    code = code.replace('"', '"').replace('"', '"')
    code = code.replace(''', "'").replace(''', "'")

    # Fix 1: 统一 HTML 转义
    code = code.replace('&lt;br/&gt;', '<br/>')
    code = code.replace('&amp;lt;br/&amp;gt;', '<br/>')

    # Fix 1: 去掉节点标签内的 <br/>[xxx] 标注 → " - xxx"
    code = re.sub(r'<br/>\[([^\]]+)\]', r' - \1', code)
    code = re.sub(r'<br/>\(([^)]+)\)', r' - \1', code)

    # Mermaid 节点中不能有裸的 < > & 字符
    lines = code.split('\n')
    cleaned = []
    for line in lines:
        stripped = line.strip()
        # 跳过纯箭头行和 style 行
        if stripped.startswith('style') or stripped.startswith('classDef'):
            cleaned.append(line)
            continue
        # 在 [] () {} 内的文字中转义
        line = re.sub(r'\[([^\]]*)<([^\]]*)\]', r'[\1&lt;\2]', line)
        line = re.sub(r'\[([^\]]*)>([^\]]*)\]', r'[\1&gt;\2]', line)

        # Fix 1: 修复嵌套方括号 A[xxx[yyy]] → A[xxx - yyy]
        # 匹配节点定义行并修复嵌套方括号
        if re.match(r'\s*\w+\[', stripped):
            # 找到第一个 [ 后的内容
            bracket_match = re.search(r'\[', stripped)
            if bracket_match:
                bracket_start = bracket_match.start()
                content = stripped[bracket_start + 1:]
                # 检查是否有嵌套 [ 且不是最后的 ]
                if '[' in content[:-1]:
                    # 将嵌套的 [] 替换为 ()
                    content = re.sub(r'\[([^\]]*)\]', r'(\1)', content)
                    line = stripped[:bracket_start + 1] + content

        cleaned.append(line)

    code = '\n'.join(cleaned)

    # 移除连续空行
    code = re.sub(r'\n\s*\n', '\n', code)

    # 修复箭头格式
    code = re.sub(r'--\s+>', '-->', code)
    code = re.sub(r'==\s+>', '==>', code)

    return code.strip()


def _generate_flow_diagrams(anchor: dict, gateway) -> list:
    """
    基于 anchor 中的 flow_diagrams 配置，调用 LLM 生成 Mermaid 流程图代码。
    返回 [{name, mermaid_code, description}, ...]
    """
    flow_configs = anchor.get('flow_diagrams', [])
    if not flow_configs:
        print("[FlowDiagram] anchor 中无流程图配置，跳过")
        return []

    print(f"[FlowDiagram] 开始生成 {len(flow_configs)} 个流程图...")

    results = []

    for i, flow in enumerate(flow_configs):
        name = flow.get('name', f'流程{i+1}')
        trigger = flow.get('trigger', '')
        scope = flow.get('scope', '')
        must_include = flow.get('must_include', [])
        exceptions = flow.get('exceptions', [])

        prompt = f"""你是智能骑行头盔产品的交互设计师。请为以下流程生成 Mermaid 流程图代码。

【流程名称】{name}
【触发条件】{trigger}
【涉及范围】{scope}

【必须包含的步骤】
{chr(10).join(f'- {s}' for s in must_include)}

【异常分支（每个异常必须有对应的处理路径）】
{chr(10).join(f'- {e}' for e in exceptions)}

【输出要求】
1. 输出纯 Mermaid flowchart 代码（用 graph TD 纵向布局）
2. 主流程用实线，异常分支用虚线
3. 每个步骤标注属于哪端（HUD/App/头盔/手机），用方括号标注如 [App]
4. 异常处理节点用红色标注：style nodeId fill:#fee,stroke:#c00
5. 成功结束节点用绿色标注：style nodeId fill:#efe,stroke:#0a0
6. 不要输出任何解释，只输出 Mermaid 代码块
7. 节点文字简洁，每个节点最多 15 个字
8. 确保语法正确，可以直接被 Mermaid 渲染引擎解析
"""

        try:
            # B1: 使用正确的 LLM 调用方式
            result_obj = gateway.call_azure_openai(
                "cpo", prompt, "只输出Mermaid代码。", "flow_diagram", max_tokens=3000
            )
            result = result_obj.get("response", "") if result_obj.get("success") else ""

            if result:
                # 提取 mermaid 代码（去掉可能的 ```mermaid ``` 包裹）
                code = result.strip()
                if code.startswith('```'):
                    code = code.split('\n', 1)[1] if '\n' in code else code
                if code.endswith('```'):
                    code = code.rsplit('```', 1)[0]
                code = code.replace('```mermaid', '').replace('```', '').strip()

                results.append({
                    'name': name,
                    'trigger': trigger,
                    'scope': scope,
                    'mermaid_code': _clean_mermaid_code(code),
                    'steps_count': len(must_include),
                    'exceptions_count': len(exceptions),
                })
                print(f"  OK [{i+1}/{len(flow_configs)}] {name}: {len(code)} chars")
            else:
                print(f"  FAIL [{i+1}/{len(flow_configs)}] {name}: LLM returned empty")
        except Exception as e:
            print(f"  FAIL [{i+1}/{len(flow_configs)}] {name}: {e}")

    print(f"[FlowDiagram] Done: {len(results)}/{len(flow_configs)} diagrams")
    return results


def _write_flow_sheet(wb, flow_diagrams: list):
    """写入关键流程 Sheet"""
    if not flow_diagrams:
        return

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet('关键流程')

    headers = ['流程名称', '触发条件', '涉及范围', '步骤数', '异常分支数', 'Mermaid代码']

    header_fill = PatternFill('solid', fgColor='4A5568')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, flow in enumerate(flow_diagrams, 2):
        ws.cell(row=i, column=1, value=flow.get('name', ''))
        ws.cell(row=i, column=2, value=flow.get('trigger', ''))
        ws.cell(row=i, column=3, value=flow.get('scope', ''))
        ws.cell(row=i, column=4, value=flow.get('steps_count', 0))
        ws.cell(row=i, column=5, value=flow.get('exceptions_count', 0))
        cell = ws.cell(row=i, column=6, value=flow.get('mermaid_code', ''))
        cell.alignment = Alignment(wrap_text=True)
        cell.border = thin_border

        for col in range(1, 6):
            ws.cell(row=i, column=col).border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 80

    print(f"  OK Flow Sheet: {len(flow_diagrams)} diagrams")

    ws.freeze_panes = "A2"
    if flow_diagrams:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(flow_diagrams) + 1}"


# ===== B5: 用户旅程和 AI 场景表 =====
def _generate_user_journeys(anchor: dict) -> list:
    """
    基于 anchor 中的 user_journeys 配置，生成 Mermaid journey 图。
    不需要调 LLM — 数据已经在 anchor 中结构化了，直接拼 Mermaid 语法。
    """
    journey_configs = anchor.get('user_journeys', [])
    if not journey_configs:
        print("[UserJourney] anchor 中无用户旅程配置，跳过")
        return []

    print(f"[UserJourney] 开始生成 {len(journey_configs)} 个角色旅程图...")

    results = []

    for j in journey_configs:
        role = j.get('role', '')
        persona = j.get('persona', '')
        stages = j.get('journey', [])

        # 拼 Mermaid journey 语法
        lines = [f'journey', f'    title {role}']

        for stage in stages:
            stage_name = stage.get('stage', '')
            touchpoints = stage.get('touchpoints', [])
            lines.append(f'    section {stage_name}')

            for i, tp in enumerate(touchpoints):
                # journey 语法: 任务名: 满意度(1-5): 角色
                # 用阶段位置模拟满意度递增（前期低=学习成本，后期高=获得价值）
                score = min(5, 3 + (i // 2))
                # 简化触点文本（去掉括号内容，限长）
                tp_short = tp.split('（')[0].split('(')[0][:25]
                lines.append(f'        {tp_short}: {score}: {role.split("（")[0]}')

        mermaid_code = '\n'.join(lines)

        total_touchpoints = sum(len(s.get('touchpoints', [])) for s in stages)

        results.append({
            'role': role,
            'persona': persona,
            'mermaid_code': mermaid_code,
            'stages_count': len(stages),
            'touchpoints_count': total_touchpoints,
        })

        print(f"  ✅ {role}: {len(stages)} 阶段, {total_touchpoints} 触点")

    print(f"[UserJourney] 完成: {len(results)} 个角色旅程")
    return results


def _write_journey_sheet(wb, journeys: list, anchor: dict = None):
    """写入用户旅程 Sheet — 结构化表格，每个触点一行"""
    if not anchor:
        return

    journey_configs = anchor.get('user_journeys', [])
    if not journey_configs:
        print("[UserJourney] anchor 中无用户旅程配置，跳过")
        return

    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet('用户旅程')

    # 表头
    headers = ['角色', '画像', '阶段', '触点', '阶段序号', '触点序号']
    header_fill = PatternFill('solid', fgColor='5A67D8')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 数据行 — 从 anchor 直接读取结构化数据
    row_idx = 2
    for j in journey_configs:
        role = j.get('role', '')
        persona = j.get('persona', '')
        stages = j.get('journey', [])

        for s_idx, stage in enumerate(stages, 1):
            stage_name = stage.get('stage', '')
            touchpoints = stage.get('touchpoints', [])

            for t_idx, tp in enumerate(touchpoints, 1):
                ws.cell(row=row_idx, column=1, value=role)
                ws.cell(row=row_idx, column=2, value=persona)
                ws.cell(row=row_idx, column=3, value=stage_name)
                ws.cell(row=row_idx, column=4, value=tp)
                ws.cell(row=row_idx, column=5, value=s_idx)
                ws.cell(row=row_idx, column=6, value=t_idx)

                # 每行设置对齐
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).alignment = Alignment(vertical='top', wrap_text=True)

                row_idx += 1

    # 列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8

    total_rows = row_idx - 2
    print(f"  ✅ 用户旅程 Sheet: {len(journey_configs)} 角色, {total_rows} 行")

    print(f"  ✅ 用户旅程 Sheet: {len(journeys)} 个角色")


def _write_ai_scenarios_sheet(wb, anchor: dict):
    """写入主动AI场景 Sheet"""
    scenarios = anchor.get('ai_proactive_scenarios', [])
    if not scenarios:
        return

    ws = wb.create_sheet('主动AI场景')

    headers = ['场景名称', '触发条件', '系统动作', '所需数据', '用户控制']
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill('solid', fgColor='38A169')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, s in enumerate(scenarios, 2):
        ws.cell(row=i, column=1, value=s.get('scenario', ''))
        ws.cell(row=i, column=2, value=s.get('trigger', ''))
        ws.cell(row=i, column=3, value=s.get('action', ''))
        data_needed = s.get('data_needed', [])
        ws.cell(row=i, column=4, value=', '.join(data_needed) if isinstance(data_needed, list) else str(data_needed))
        ws.cell(row=i, column=5, value=s.get('user_control', ''))
        # 每列自动换行
        for col in range(1, 6):
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30

    print(f"  ✅ 主动AI场景 Sheet: {len(scenarios)} 个场景")


def _compare_prd_versions(old_path: str, new_path: str) -> str:
    """对比两个版本的 PRD,生成变更清单"""
    old = json.loads(Path(old_path).read_text(encoding="utf-8"))
    new = json.loads(Path(new_path).read_text(encoding="utf-8"))

    old_names = {(i["name"], i["level"]) for i in old.get("items", [])}
    new_names = {(i["name"], i["level"]) for i in new.get("items", [])}

    added = new_names - old_names
    removed = old_names - new_names

    # 优先级变更
    old_priorities = {i["name"]: i.get("priority", "") for i in old.get("items", [])}
    priority_changes = []
    for item in new.get("items", []):
        name = item["name"]
        new_p = item.get("priority", "")
        old_p = old_priorities.get(name, "")
        if old_p and new_p and old_p != new_p:
            priority_changes.append(f"{name}: {old_p} -> {new_p}")

    report = (
        f"PRD 版本对比\n"
        f"旧版: {old.get('version', '?')} ({old.get('total_features', 0)} 条)\n"
        f"新版: {new.get('version', '?')} ({new.get('total_features', 0)} 条)\n\n"
        f"新增功能 ({len(added)} 条):\n" + "\n".join(f"  + {n} [{l}]" for n, l in sorted(added)[:20]) + "\n\n"
        f"移除功能 ({len(removed)} 条):\n" + "\n".join(f"  - {n} [{l}]" for n, l in sorted(removed)[:20]) + "\n\n"
        f"优先级变更 ({len(priority_changes)} 条):\n" + "\n".join(f"  ~ {c}" for c in priority_changes[:20]) + "\n"
    )
    return report


def _write_generic_sheet(ws, items: List[Dict], headers: List[str], keys: List[str], widths: List[int]):
    """通用 Sheet 写入函数"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 优先级颜色
    priority_fills = {
        "P0": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "P1": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "P2": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        "P3": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
    }
    priority_fonts = {
        "P0": Font(bold=True, color="FFFFFF"),
        "P1": Font(bold=True, color="FFFFFF"),
        "P2": Font(color="FFFFFF"),
        "P3": Font(color="FFFFFF"),
    }

    # 写表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 写数据
    for row_idx, item in enumerate(items, 2):
        for col, key in enumerate(keys, 1):
            val = item.get(key, "")
            # 防止 list/dict 写入 Excel
            if isinstance(val, (list, dict)):
                val = ", ".join(str(x) for x in val) if isinstance(val, list) else str(val)
            elif val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val else "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 优先级颜色
            if key == "priority" and str(val).upper() in priority_fills:
                cell.fill = priority_fills[str(val).upper()]
                cell.font = priority_fonts[str(val).upper()]
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # 冻结+筛选
    ws.freeze_panes = "A2"
    if items:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(items) + 1}"


# === 脑图 .xmind 生成（ZIP 格式）===
def _generate_mindmap_xmind(items: List[Dict], title: str = "智能骑行头盔 V1 功能框架") -> bytes:
    """生成 .xmind 格式脑图（ZIP 包含 content.json + metadata.json）"""
    import uuid
    from io import BytesIO
    import zipfile

    HUD_MODULES = {"导航", "来电", "音乐", "消息", "AI语音助手", "Ai语音助手",
        "简易", "简易路线", "路线", "主动安全预警提示", "组队", "摄像状态",
        "胎温胎压", "开机动画", "速度", "设备状态", "显示队友位置", "头盔HUD",
        "实体按键交互", "氛围灯交互", "AI功能", "语音交互", "视觉交互", "多模态交互"}
    APP_MODULES = {"App-设备", "App-社区", "App-商城", "App-我的"}

    # 构建树结构
    groups = {
        "HUD及头盔端": [],
        "App端": [],
        "系统与交互": [],
    }

    current_l1_node = None
    current_l2_node = None

    for item in items:
        level = item.get("level", "")
        name = item.get("name", "")
        module = item.get("module", "")
        priority = item.get("priority", "")

        if module in HUD_MODULES:
            group = "HUD及头盔端"
        elif module in APP_MODULES:
            group = "App端"
        else:
            group = "系统与交互"

        label_text = f"{name} [{priority}]" if priority else name
        node = {"id": str(uuid.uuid4()), "title": label_text, "children": {"attached": []}}

        if level == "L1":
            groups[group].append(node)
            current_l1_node = node
            current_l2_node = None
        elif level == "L2" and current_l1_node:
            current_l1_node["children"]["attached"].append(node)
            current_l2_node = node
        elif level == "L3" and current_l2_node:
            current_l2_node["children"]["attached"].append(node)

    # 构建 xmind content.json
    root_node = {
        "id": str(uuid.uuid4()),
        "class": "topic",
        "title": title,
        "children": {"attached": []}
    }

    for group_name, group_items in groups.items():
        if group_items:
            group_node = {
                "id": str(uuid.uuid4()),
                "title": group_name,
                "children": {"attached": group_items}
            }
            root_node["children"]["attached"].append(group_node)

    content = [{
        "id": str(uuid.uuid4()),
        "class": "sheet",
        "title": title,
        "rootTopic": root_node,
        "topicPositioning": "fixed"
    }]

    metadata = {
        "creator": {"name": "Agent Company", "version": "1.0"},
        "activeSheetId": content[0]["id"]
    }

    # Fix 7: 添加 manifest.json 以兼容 XMind 在线版
    manifest = {
        "file-entries": {
            "content.json": {},
            "metadata.json": {},
            "manifest.json": {}
        }
    }

    # 打包为 ZIP
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    return buf.getvalue()


# === 脑图 .mm 生成(FreeMind XML 格式)===
def _generate_mindmap_mm(items: List[Dict], title: str = "智能骑行头盔 V1 功能框架") -> str:
    """生成 FreeMind .mm 格式脑图,无冗余层"""
    import xml.etree.ElementTree as ET

    root = ET.Element("map", version="1.0.1")
    root_node = ET.SubElement(root, "node", TEXT=title)

    # 按 L1 分组,跳过模块中间层
    current_l1_node = None
    current_l2_node = None

    # 先按大类分组:HUD端 / App端 / 系统交互
    hud_group = ET.SubElement(root_node, "node", TEXT="HUD及头盔端")
    app_group = ET.SubElement(root_node, "node", TEXT="App端")
    system_group = ET.SubElement(root_node, "node", TEXT="系统与交互")

    HUD_MODULES = {"导航", "来电", "音乐", "消息", "AI语音助手", "Ai语音助手",
        "简易", "简易路线", "路线", "主动安全预警提示", "组队", "摄像状态",
        "胎温胎压", "开机动画", "速度", "设备状态", "显示队友位置", "头盔HUD"}
    APP_MODULES = {"App-设备", "App-社区", "App-商城", "App-我的"}

    for item in items:
        level = item.get("level", "")
        name = item.get("name", "")
        module = item.get("module", "")
        priority = item.get("priority", "")
        label = f"{name} [{priority}]" if priority else name

        # 选择父组
        if module in HUD_MODULES:
            parent_group = hud_group
        elif module in APP_MODULES:
            parent_group = app_group
        else:
            parent_group = system_group

        if level == "L1":
            current_l1_node = ET.SubElement(parent_group, "node", TEXT=label)
            # P0 红色,P1 橙色,P2 绿色,P3 灰色
            color = {"P0": "#FF4444", "P1": "#FF8C00", "P2": "#4CAF50", "P3": "#9E9E9E"}.get(priority, "")
            if color:
                current_l1_node.set("COLOR", color)
            current_l2_node = None
        elif level == "L2" and current_l1_node is not None:
            current_l2_node = ET.SubElement(current_l1_node, "node", TEXT=label)
        elif level == "L3" and current_l2_node is not None:
            ET.SubElement(current_l2_node, "node", TEXT=label)

    import io
    tree = ET.ElementTree(root)
    buf = io.BytesIO()
    tree.write(buf, encoding="utf-8", xml_declaration=True)
    return buf.getvalue().decode("utf-8")


def _generate_mindmap_svg(items: List[Dict], title: str = "智能骑行头盔 V1 功能框架") -> str:
    """生成交互式 HTML 脑图(D3.js,可展开折叠、缩放、拖拽)"""

    # P3: 过滤研发条目（脑图只展示用户可见功能）
    filtered_items = [r for r in items if '[研发]' not in str(r.get('note', ''))]
    if len(filtered_items) < len(items):
        print(f"  [Mindmap] 过滤 {len(items) - len(filtered_items)} 条研发条目")
    items = filtered_items

    # 构建树结构
    tree = {"name": title, "children": []}

    HUD_MODULES = {"导航", "来电", "音乐", "消息", "AI语音助手", "Ai语音助手",
        "简易", "简易路线", "路线", "主动安全预警提示", "组队", "摄像状态",
        "胎温胎压", "开机动画", "速度", "设备状态", "显示队友位置", "头盔HUD"}
    APP_MODULES = {"App-设备", "App-社区", "App-商城", "App-我的"}

    groups = {
        "HUD及头盔端": {"name": "HUD及头盔端", "children": []},
        "App端": {"name": "App端", "children": []},
        "系统与交互": {"name": "系统与交互", "children": []},
    }

    current_l1 = None
    current_l2 = None

    for item in items:
        level = item.get("level", "")
        name = item.get("name", "")
        module = item.get("module", "")
        priority = item.get("priority", "")

        if module in HUD_MODULES:
            group = groups["HUD及头盔端"]
        elif module in APP_MODULES:
            group = groups["App端"]
        else:
            group = groups["系统与交互"]

        node = {"name": f"{name} [{priority}]", "priority": priority, "children": []}

        if level == "L1":
            group["children"].append(node)
            current_l1 = node
            current_l2 = None
        elif level == "L2" and current_l1:
            current_l1["children"].append(node)
            current_l2 = node
        elif level == "L3" and current_l2:
            current_l2["children"].append(node)

    tree["children"] = [g for g in groups.values() if g["children"]]

    # 生成 HTML+D3.js 可交互脑图（径向布局）
    import json as _json
    tree_json = _json.dumps(tree, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>{title}</title>
<style>
/* ===== 修复 D: 脑图布局修复 ===== */
html, body {{ margin: 0; padding: 0; width: 100%; height: 100%; overflow: hidden; }}
body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; background: #fafbfc; }}
svg {{ width: 100vw; height: 100vh; display: block; }}
.node circle {{ fill: #fff; stroke: #2F5496; stroke-width: 2px; cursor: pointer; }}
.node text {{ font-size: 11px; fill: #333; pointer-events: none; }}
.link {{ fill: none; stroke: #ccc; stroke-width: 1.5px; }}
#controls {{ position: fixed; top: 10px; left: 10px; z-index: 100; background: rgba(255,255,255,0.95);
    padding: 8px 12px; border-radius: 6px; box-shadow: 0 2px 8px rgba(0,0,0,0.15); }}
button {{ margin: 0 4px; padding: 4px 10px; cursor: pointer; border: 1px solid #ccc;
    border-radius: 4px; background: #f5f5f5; }}
button:hover {{ background: #e0e0e0; }}
</style>
<script src="https://d3js.org/d3.v7.min.js"></script>
</head><body>
<div id="controls">
    <button onclick="expandAll()">全部展开</button>
    <button onclick="collapseAll()">全部折叠</button>
    <button onclick="fitToContent()">重置缩放</button>
    <span style="margin-left:10px;color:#999;font-size:12px;">滚轮缩放 | 拖拽平移 | 点击展开/折叠</span>
</div>
<script>
const treeData = {tree_json};

// Bug 5: 改为水平树状布局
const width = window.innerWidth;
const height = window.innerHeight;

const svg = d3.select("body").append("svg")
    .attr("width", width).attr("height", height)
    .style("background", "#fafbfc");

const zoomBehavior = d3.zoom()
    .scaleExtent([0.05, 10])
    .on("zoom", (e) => g.attr("transform", e.transform));

svg.call(zoomBehavior);

const g = svg.append("g")
    .attr("transform", `translate(100,${{height/2}})`);

// 水平树布局
const tree = d3.tree()
    .nodeSize([30, 250])  // [垂直间距, 水平间距]
    .separation((a, b) => a.parent === b.parent ? 1 : 1.5);

const root = d3.hierarchy(treeData);

// 初始只展开到第 2 层
root.descendants().forEach(d => {{
    if (d.depth >= 2) {{ d._children = d.children; d.children = null; }}
}});

const priorityColor = {{"P0":"#FF4444","P1":"#FF8C00","P2":"#4CAF50","P3":"#9E9E9E"}};

function update(source) {{
    const treeLayout = tree(root);
    const nodes = root.descendants();
    const links = root.links();

    const node = g.selectAll("g.node").data(nodes, d => d.data.name + d.depth);

    const nodeEnter = node.enter().append("g").attr("class", "node")
        .attr("transform", d => `translate(${{d.y}},${{d.x}})`)
        .on("click", (e, d) => {{ e.stopPropagation(); toggle(d); update(d); }});

    nodeEnter.append("circle").attr("r", 5)
        .style("fill", d => {{
            const p = (d.data.priority || "").toUpperCase();
            return priorityColor[p] || (d._children ? "#ddd" : "#fff");
        }});

    nodeEnter.append("text")
        .attr("dy", "0.31em")
        .attr("x", d => d.children ? -8 : 8)
        .attr("text-anchor", d => d.children ? "end" : "start")
        .text(d => d.data.name)
        .style("font-size", d => d.depth === 0 ? "16px" : d.depth === 1 ? "13px" : "11px")
        .style("font-weight", d => d.depth <= 1 ? "bold" : "normal");

    node.merge(nodeEnter).transition().duration(300)
        .attr("transform", d => `translate(${{d.y}},${{d.x}})`);

    node.exit().remove();

    const link = g.selectAll("path.link").data(links, d => d.target.data.name + d.target.depth);

    link.enter().insert("path", "g").attr("class", "link")
        .attr("d", d3.linkHorizontal()
            .x(d => d.y)
            .y(d => d.x));

    link.transition().duration(300)
        .attr("d", d3.linkHorizontal()
            .x(d => d.y)
            .y(d => d.x));

    link.exit().remove();
}}

function toggle(d) {{
    if (d.children) {{ d._children = d.children; d.children = null; }}
    else {{ d.children = d._children; d._children = null; }}
}}

function expandAll() {{
    root.descendants().forEach(d => {{
        if (d._children) {{ d.children = d._children; d._children = null; }}
    }});
    update(root);
}}

function collapseAll() {{
    root.descendants().forEach(d => {{
        if (d.depth >= 1 && d.children) {{ d._children = d.children; d.children = null; }}
    }});
    update(root);
}}

function fitToContent() {{
    const bbox = g.node().getBBox();
    const padding = 80;
    const contentWidth = bbox.width + padding * 2;
    const contentHeight = bbox.height + padding * 2;
    const svgWidth = window.innerWidth;
    const svgHeight = window.innerHeight;

    const scale = Math.min(
        svgWidth / contentWidth,
        svgHeight / contentHeight,
        1.5
    ) * 0.9;

    const cx = bbox.x + bbox.width / 2;
    const cy = bbox.y + bbox.height / 2;
    const translateX = svgWidth / 2 - cx * scale;
    const translateY = svgHeight / 2 - cy * scale;

    svg.transition().duration(750).call(
        zoomBehavior.transform,
        d3.zoomIdentity.translate(translateX, translateY).scale(scale)
    );
}}

update(root);

// 初始化时自动 fit to content
setTimeout(fitToContent, 300);

// 窗口大小变化时重新适配
window.addEventListener('resize', () => {{
    svg.attr("width", window.innerWidth).attr("height", window.innerHeight);
    setTimeout(fitToContent, 100);
}});
</script></body></html>"""

    return html


def _truncate_list(data: List[Dict], max_len: int = 120) -> List[Dict]:
    """压缩数据列表:截断长文本字段,减少 HTML 体积（目标 < 400KB）"""
    result = []
    for item in data:
        truncated = {}
        for k, v in item.items():
            if isinstance(v, str) and len(v) > max_len:
                truncated[k] = v[:max_len] + "..."
            elif isinstance(v, (list, dict)):
                # list/dict 转 string 后截断
                v_str = ", ".join(str(x) for x in v) if isinstance(v, list) else str(v)
                truncated[k] = v_str[:max_len] + "..." if len(v_str) > max_len else v_str
            elif v is None:
                truncated[k] = ""
            else:
                truncated[k] = v
        result.append(truncated)
    return result


def _generate_interactive_prd_html(items: List[Dict], extra_sheets: Dict = None, title: str = "智能骑行头盔 V1 PRD 规格书", anchor: dict = None) -> str:
    """生成交互式 HTML PRD(可搜索、可筛选、可展开)"""
    import json as _json

    # Bug 2 修复: 使用 anchor 的 placement 进行分流
    if anchor:
        module_placement = _build_module_placement(anchor)
    else:
        module_placement = {}

    hud_features = []
    app_features = []
    for item in items:
        module_name = item.get("module", "")
        target = module_placement.get(module_name, '')
        if not target:
            # 兜底推断
            if any(kw in module_name for kw in ['Tab', 'App-', '商城', '社区', '通知', '用户成就', '新手', '身份', '权限', '部件识别', '消息同步', '多语言', '恢复出厂']):
                target = 'app'
            else:
                target = 'hud'
        if target == 'hud':
            hud_features.append(item)
        else:
            app_features.append(item)

    # Bug 3 修复: 分流后归一化（防止分流带入旧名称）
    if anchor:
        normalize_map = anchor.get('module_normalize', {})
    else:
        normalize_map = {}
    normalize_map.update(MODULE_NAME_NORMALIZE)
    hud_features = _normalize_all_rows(hud_features, normalize_map, anchor)
    app_features = _normalize_all_rows(app_features, normalize_map, anchor)
    print(f"[Normalize-HTML] 分流后: HUD {len(hud_features)}, App {len(app_features)}")

    print(f"[HTML] HUD 功能: {len(hud_features)}, App 功能: {len(app_features)}")

    # 准备数据(压缩后减少 HTML体积)
    # B5: 添加用户旅程和 AI 场景数据
    # 修复 4: 用户旅程改为结构化数据
    journey_data = []
    if anchor:
        for j in anchor.get('user_journeys', []):
            role = j.get('role', '')
            persona = j.get('persona', '')
            stages = []
            for stage in j.get('journey', []):
                stages.append({
                    'stage': stage.get('stage', ''),
                    'touchpoints': stage.get('touchpoints', [])
                })
            journey_data.append({
                'role': role,
                'persona': persona,
                'stages': stages
            })
    ai_scenarios = anchor.get('ai_proactive_scenarios', []) if anchor else []

    all_data = {
        "features": _truncate_list(items),
        "hud_features": _truncate_list(hud_features),
        "app_features": _truncate_list(app_features),
        "state_scenarios": _truncate_list(extra_sheets.get("state", [])) if extra_sheets else [],
        "voice_commands": _truncate_list(extra_sheets.get("voice", [])) if extra_sheets else [],
        "button_mapping": _truncate_list(extra_sheets.get("button", [])) if extra_sheets else [],
        "light_effects": _truncate_list(extra_sheets.get("light", [])) if extra_sheets else [],
        "voice_nav": _truncate_list(extra_sheets.get("voice_nav", [])) if extra_sheets else [],
        "user_stories": _truncate_list(extra_sheets.get("user_stories", [])) if extra_sheets else [],
        "test_cases": _truncate_list(extra_sheets.get("test_cases", [])) if extra_sheets else [],
        "page_mapping": _truncate_list(extra_sheets.get("page_mapping", [])) if extra_sheets else [],
        "dev_tasks": _truncate_list(extra_sheets.get("dev_tasks", [])) if extra_sheets else [],
        "flow_diagrams": extra_sheets.get("flow", []) if extra_sheets else [],
        "user_journeys": journey_data,
        "ai_scenarios": ai_scenarios,
    }

    # Fix 4: 一致性校验 - 确保 features = hud_features + app_features
    total = len(all_data.get('features', []))
    hud_count = len(all_data.get('hud_features', []))
    app_count = len(all_data.get('app_features', []))
    if total != hud_count + app_count:
        print(f"[ConsistCheck] ⚠️ features({total}) ≠ hud({hud_count}) + app({app_count}), 差{total - hud_count - app_count}条")
        # 强制修正
        all_data['features'] = all_data['hud_features'] + all_data['app_features']
        print(f"[ConsistCheck] 已修正 features = {len(all_data['features'])}")
    else:
        print(f"[ConsistCheck] ✅ features={total} = hud({hud_count}) + app({app_count})")

    data_json = _json.dumps(all_data, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>{title}</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
/* ===== 修复 A: 固定布局方案 ===== */
html, body {{ margin: 0; padding: 0; height: 100%; overflow: hidden; }}
body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; background: #f5f6fa; color: #333; }}
.page-wrapper {{ display: flex; flex-direction: column; height: 100vh; }}
.header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #fff;
    position: fixed; top: 0; left: 0; right: 0; z-index: 1000; }}
.header h1 {{ font-size: 20px; margin-bottom: 8px; padding: 12px 24px 0; }}
.controls {{ padding: 12px 24px; display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
.controls input {{ padding: 6px 12px; border: none; border-radius: 4px; width: 280px; font-size: 14px; }}
.controls select {{ padding: 6px 10px; border: none; border-radius: 4px; font-size: 13px; }}
.controls button {{ padding: 6px 14px; border: none; border-radius: 4px; cursor: pointer;
    background: rgba(255,255,255,0.2); color: #fff; font-size: 13px; }}
.controls button:hover {{ background: rgba(255,255,255,0.35); }}
.tabs {{ display: flex; background: #fff; padding: 0 24px;
    position: fixed; left: 0; right: 0; z-index: 999;
    border-bottom: 1px solid #e0e0e0; box-shadow: 0 2px 4px rgba(0,0,0,0.05); overflow-x: auto;
    -webkit-overflow-scrolling: touch; }}
.tab {{ padding: 12px 20px; color: #666; cursor: pointer; font-size: 14px;
    border-bottom: 3px solid transparent; flex-shrink: 0; white-space: nowrap; }}
.tab:hover {{ color: #2F5496; background: #f7f8ff; }}
.tab.active {{ color: #2F5496; border-bottom-color: #2F5496; font-weight: 600; }}
.tab .badge {{ background: #e8edf5; padding: 1px 8px; border-radius: 10px;
    font-size: 11px; margin-left: 6px; color: #2F5496; }}
.content {{ padding: 16px 24px; max-width: 1600px; margin-left: auto; margin-right: auto;
    margin-top: 160px; overflow-y: auto; height: calc(100vh - 160px); -webkit-overflow-scrolling: touch; }}
.tab-content {{ display: none; }}
.tab-content.active {{ display: block; }}
.stats {{ display: flex; gap: 15px; margin-bottom: 20px; flex-wrap: wrap; }}
.stat {{ background: #fff; padding: 12px 20px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
.stat .num {{ font-size: 24px; font-weight: bold; color: #2F5496; }}
.stat .label {{ font-size: 12px; color: #999; }}
.tree-item {{ margin-left: 0; }}
.l1 {{ background: #fff; margin-bottom: 8px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); overflow: hidden; }}
.l1-header {{ padding: 12px 16px; cursor: pointer; display: flex; align-items: center; gap: 10px;
    background: #e8edf5; font-weight: bold; font-size: 15px; }}
.l1-header:hover {{ background: #dce3f0; }}
.l1-body {{ padding: 0 16px 12px; display: none; }}
.l1-body.open {{ display: block; }}
.l2 {{ margin: 8px 0 4px 12px; }}
.l2-title {{ font-weight: bold; font-size: 14px; color: #2F5496; padding: 6px 0; cursor: pointer; }}
.l2-body {{ margin-left: 12px; }}
.l3 {{ margin-left: 24px; padding: 6px 0; font-size: 13px; color: #555;
    border-bottom: 1px solid #f0f0f0; display: flex; gap: 12px; align-items: flex-start;
    min-width: 200px; max-width: 100%; flex-wrap: wrap; }}
.l3:last-child {{ border-bottom: none; }}
.l3 .name {{ min-width: 160px; max-width: 220px; flex-shrink: 0; font-weight: 500; color: #333; }}
.l3 .desc {{ flex: 1; min-width: 200px; color: #666; }}
.l3 .acc {{ flex: 1; min-width: 200px; color: #4a9; font-size: 12px; }}
/* B5: HUD 端新增 4 列样式 */
.visual-output {{ flex: 1; min-width: 150px; color: #5a67d8; font-size: 12px; }}
.display-priority {{ min-width: 60px; font-size: 11px; }}
.display-priority.tag-critical {{ color: #c53030; font-weight: 700; }}
.display-priority.tag-high {{ color: #b7791f; }}
.display-priority.tag-medium {{ color: #276749; }}
.display-priority.tag-low {{ color: #718096; }}
.degradation {{ flex: 1; min-width: 150px; color: #e53e3e; font-size: 12px; }}
.display-duration {{ min-width: 80px; color: #718096; font-size: 12px; }}
.tag {{ display: inline-block; padding: 2px 8px; border-radius: 3px; font-size: 11px; color: #fff; }}
.tag-P0 {{ background: #FF4444; }}
.tag-P1 {{ background: #FF8C00; }}
.tag-P2 {{ background: #4CAF50; }}
.tag-P3 {{ background: #9E9E9E; }}
/* ===== 修复 C: 空内容行样式降级 ===== */
.l1.empty-content {{ opacity: 0.6; border-left: 3px solid #ffd700; }}
.l1.empty-content .l1-header::after {{
    content: '待补充'; font-size: 11px; color: #999; background: #fff3cd;
    padding: 2px 8px; border-radius: 4px; margin-left: auto; }}
.l3.empty-desc {{ color: #999; font-style: italic; }}
.l3.empty-desc::after {{ content: '（待补充）'; font-size: 11px; color: #bbb; margin-left: 4px; }}
/* ===== 修复 B: 表格布局统一 ===== */
.table-wrapper {{ width: 100%; overflow-x: auto; border-radius: 8px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08); margin-bottom: 16px; background: #fff; }}
table {{ width: 100%; min-width: 900px; border-collapse: collapse; table-layout: fixed; }}
thead {{ position: sticky; top: 0; z-index: 50; }}
th {{ background: #4a5568; color: #fff; padding: 10px 12px; text-align: left; font-size: 13px;
    font-weight: 600; white-space: nowrap; border-bottom: 2px solid #2d3748; }}
td {{ padding: 8px 12px; border-bottom: 1px solid #edf2f7; font-size: 13px; vertical-align: top;
    word-wrap: break-word; overflow-wrap: break-word; }}
tbody tr:nth-child(odd) {{ background: #fafbfc; }}
tbody tr:hover {{ background: #f0f4ff; }}
/* ===== 修复 E: 全局样式提升 ===== */
.priority-tag {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.5px; }}
.priority-tag.p0 {{ background: #fed7d7; color: #c53030; }}
.priority-tag.p1 {{ background: #fefcbf; color: #b7791f; }}
.priority-tag.p2 {{ background: #c6f6d5; color: #276749; }}
.priority-tag.p3 {{ background: #e2e8f0; color: #4a5568; }}
.l1 {{ transition: box-shadow 0.2s ease, transform 0.2s ease; }}
.l1:hover {{ box-shadow: 0 4px 12px rgba(0,0,0,0.12); transform: translateY(-1px); }}
.l3:nth-child(odd) {{ background: #fafbfc; }}
.l3:hover {{ background: #f0f4ff; }}
.acc-value {{ color: #2b6cb0; font-weight: 600; }}
.search-highlight {{ background: #fff3cd; padding: 0 2px; border-radius: 2px; }}
.stats-bar {{ display: flex; gap: 16px; padding: 8px 24px;
    background: rgba(255,255,255,0.1); font-size: 12px; color: rgba(255,255,255,0.85); }}
.stats-bar .stat {{ display: flex; align-items: center; gap: 4px; }}
.stats-bar .stat-value {{ font-weight: 700; color: #fff; }}
.hidden {{ display: none !important; }}
@media (max-width: 768px) {{
    .header h1 {{ font-size: 16px; }}
    .controls {{ flex-direction: column; }}
    .controls input {{ width: 100%; }}
    .tabs {{ overflow-x: auto; white-space: nowrap; padding: 0 10px; }}
    .tab {{ padding: 10px 14px; font-size: 12px; }}
    .content {{ padding: 10px 12px; }}
    .l3 {{ flex-direction: column; gap: 4px; }}
    .l3 .name {{ max-width: none; min-width: auto; }}
    table {{ font-size: 12px; display: block; overflow-x: auto; }}
    th, td {{ padding: 6px 8px; }}
}}

/* Flow diagram card styles */
.flow-card {{ background: #fff; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); margin-bottom: 16px; overflow: hidden; }}
.flow-header {{ display: flex; align-items: center; padding: 12px 20px; cursor: pointer; background: #f7fafc; border-bottom: 1px solid #e2e8f0; gap: 16px; }}
.flow-header:hover {{ background: #edf2f7; }}
.flow-title {{ font-weight: 600; font-size: 15px; color: #2d3748; min-width: 150px; }}
.flow-meta {{ flex: 1; font-size: 12px; color: #718096; }}
.flow-toggle {{ font-size: 12px; color: #a0aec0; transition: transform 0.2s; }}
.flow-body {{ padding: 20px; overflow-x: auto; }}
.flow-body .mermaid {{ display: flex; justify-content: center; }}
.flow-body svg {{ max-width: 100%; height: auto; }}
</style>
</head><body>

<div class="header">
    <h1>{title}</h1>
    <div class="stats-bar" id="statsBar"></div>
    <div class="controls">
        <input type="text" id="search" placeholder="搜索功能名称..." oninput="filterAll()">
        <select id="priorityFilter" onchange="filterAll()">
            <option value="">全部优先级</option>
            <option value="P0">P0 必做</option>
            <option value="P1">P1 应有</option>
            <option value="P2">P2 规划</option>
            <option value="P3">P3 远期</option>
        </select>
        <button onclick="expandAllTrees()">全部展开</button>
        <button onclick="collapseAllTrees()">全部折叠</button>
    </div>
</div>

<div class="tabs" id="tabBar"></div>
<div class="content" id="contentArea"></div>

<script>
const DATA = {data_json};

const TAB_CONFIG = [
    {{id: "hud", label: "HUD及头盔端", type: "tree"}},
    {{id: "app", label: "App端", type: "tree"}},
    {{id: "state", label: "状态场景对策", type: "table"}},
    {{id: "voice", label: "语音指令表", type: "table"}},
    {{id: "button", label: "按键映射表", type: "table"}},
    {{id: "light", label: "灯效定义表", type: "table"}},
    {{id: "voice_nav", label: "AI语音导航", type: "table"}},
    {{id: "user_stories", label: "用户故事", type: "table"}},
    {{id: "test_cases", label: "测试用例", type: "table"}},
    {{id: "page_mapping", label: "页面映射", type: "table"}},
    {{id: "dev_tasks", label: "开发任务", type: "table"}},
    {{id: "flow", label: "关键流程", type: "flow"}},
    // B5: 新增用户旅程和AI场景Tab
    {{id: "journey", label: "用户旅程", type: "table"}},
    {{id: "ai_scenarios", label: "主动AI场景", type: "table"}},
];

// Bug 2 修复: 使用服务端预分流的数据，替代静态过滤
const hudFeatures = DATA.hud_features || [];
const appFeatures = DATA.app_features || [];

function buildTabs() {{
    const bar = document.getElementById('tabBar');
    const content = document.getElementById('contentArea');

    TAB_CONFIG.forEach((tab, idx) => {{
        const el = document.createElement('div');
        el.className = 'tab' + (idx === 0 ? ' active' : '');
        el.dataset.target = tab.id;

        let count = 0;
        try {{
            if (tab.id === 'hud') count = (hudFeatures || []).length;
            else if (tab.id === 'app') count = (appFeatures || []).length;
            else if (tab.id === 'state') count = (DATA.state_scenarios || []).length;
            else if (tab.id === 'voice') count = (DATA.voice_commands || []).length;
            else if (tab.id === 'button') count = (DATA.button_mapping || []).length;
            else if (tab.id === 'light') count = (DATA.light_effects || []).length;
            else if (tab.id === 'voice_nav') count = (DATA.voice_nav || []).length;
            else if (tab.id === 'user_stories') count = (DATA.user_stories || []).length;
            else if (tab.id === 'test_cases') count = (DATA.test_cases || []).length;
            else if (tab.id === 'page_mapping') count = (DATA.page_mapping || []).length;
            else if (tab.id === 'dev_tasks') count = (DATA.dev_tasks || []).length;
            else if (tab.id === 'journey') count = (DATA.user_journeys || []).length;
            else if (tab.id === 'ai_scenarios') count = (DATA.ai_scenarios || []).length;
            else if (tab.id === 'flow') count = (DATA.flow_diagrams || []).length;
        }} catch(e) {{
            console.error('[PRD] count 计算错误:', tab.id, e);
        }}

        el.innerHTML = `${{tab.label}}<span class="badge">${{count}}</span>`;
        el.onclick = () => switchTab(tab.id);
        bar.appendChild(el);

        const section = document.createElement('div');
        section.id = 'section-' + tab.id;
        section.className = 'tab-content' + (idx === 0 ? ' active' : '');
        section.dataset.tab = tab.id;

        // ★★★ 关键修复：try-catch 包裹，防止单个 Tab 出错导致整页空白 ★★★
        try {{
            if (tab.type === 'tree') {{
                section.innerHTML = buildTreeView(tab.id === 'hud' ? hudFeatures : appFeatures, tab.id === 'hud');
            }} else if (tab.type === 'flow') {{
                section.innerHTML = buildFlowView();
            }} else {{
                section.innerHTML = buildTableView(tab.id);
            }}
        }} catch(e) {{
            console.error('[PRD] Tab "' + tab.id + '" 渲染错误:', e);
            section.innerHTML = '<div style="padding:40px;text-align:center;color:#c00;">⚠️ 渲染错误: ' + (e.message || e) + '<br>请按 F12 查看 Console 详情</div>';
        }}

        content.appendChild(section);
    }});
}}

function switchTab(id) {{
    document.querySelectorAll('.tab').forEach(t => t.classList.toggle('active', t.dataset.target === id));
    document.querySelectorAll('.tab-content').forEach(s => s.classList.toggle('active', s.dataset.tab === id));
    // 滚动到内容区顶部
    const content = document.querySelector('.content');
    if (content) content.scrollTop = 0;
    // 如果是流程图/旅程图 tab，触发 Mermaid 渲染
    if (id === 'flow' || id === 'journey') {{
        if (typeof mermaid !== 'undefined') {{
            setTimeout(() => mermaid.init(undefined, '.mermaid'), 100);
        }}
    }}
}}

function buildTreeView(features, isHud = false) {{
    let html = '<div class="stats">';
    const counts = {{}};
    features.forEach(f => {{ counts[f.priority] = (counts[f.priority]||0) + 1; }});
    Object.entries(counts).sort().forEach(([p,c]) => {{
        html += `<div class="stat"><div class="num">${{c}}</div><div class="label">${{p}}</div></div>`;
    }});
    html += `<div class="stat"><div class="num">${{features.length}}</div><div class="label">总计</div></div></div>`;

    // 检测 L1 模块是否有实质内容（L2/L3 子节点）
    const l1HasChildren = {{}};
    features.forEach(f => {{
        if (f.level === 'L1') l1HasChildren[f.name] = false;
        if (f.level === 'L2' || f.level === 'L3') l1HasChildren[f.name] = true;
    }});

    let currentL1 = null, currentL2 = null;
    let l1Html = '';

    features.forEach(f => {{
        const tag = `<span class="tag tag-${{f.priority}}">${{f.priority}}</span>`;
        if (f.level === 'L1') {{
            if (currentL2) {{ l1Html += '</div>'; currentL2 = null; }}
            if (currentL1) l1Html += '</div></div>';
            // 检查是否有子节点内容
            const isEmpty = !l1HasChildren[f.name];
            const l1Class = isEmpty ? 'l1 empty-content' : 'l1';
            l1Html += `<div class="${{l1Class}}" data-name="${{f.name}}" data-priority="${{f.priority}}">
                <div class="l1-header" onclick="this.nextElementSibling.classList.toggle('open')">
                    ${{tag}} ${{f.name}} <span style="color:#999;font-size:12px;margin-left:auto">${{f.interaction||''}}</span>
                </div><div class="l1-body">`;
            currentL1 = f; currentL2 = null;
        }} else if (f.level === 'L2') {{
            if (currentL2) l1Html += '</div>';
            l1Html += `<div class="l2"><div class="l2-title" data-name="${{f.name}}" data-priority="${{f.priority}}">${{tag}} ${{f.name}}</div>`;
            currentL2 = f;
        }} else if (f.level === 'L3') {{
            // 检查描述是否为空或待生成
            const desc = f.description || '';
            const isDescEmpty = !desc || desc.includes('待生成') || desc.includes('待补充') || desc.length < 5;
            const descClass = isDescEmpty ? 'desc empty-desc' : 'desc';
            const descText = isDescEmpty ? '功能规划中' : desc;

            // B5: HUD 端增加 4 个新字段
            let extraFields = '';
            if (isHud) {{
                const dp = (f.display_priority || 'medium').toLowerCase();
                extraFields = `
                    <span class="visual-output">${{f.visual_output || ''}}</span>
                    <span class="display-priority tag-${{dp}}">${{f.display_priority || ''}}</span>
                    <span class="degradation">${{f.degradation || ''}}</span>
                    <span class="display-duration">${{f.display_duration || ''}}</span>`;
            }}

            l1Html += `<div class="l3" data-name="${{f.name}}" data-priority="${{f.priority}}">
                <span class="name">${{tag}} ${{f.name}}</span>
                <span class="${{descClass}}">${{descText}}</span>
                <span class="acc">${{f.acceptance||''}}</span>${{extraFields}}</div>`;
        }}
    }});
    if (currentL2) l1Html += '</div>';
    if (currentL1) l1Html += '</div></div>';

    return html + l1Html;
}}

function cleanMermaid(code) {{
    code = (code || '').replace(/```mermaid/g, '').replace(/```/g, '');
    code = code.replace(/（/g, '(').replace(/）/g, ')');
    code = code.replace(/【/g, '[').replace(/】/g, ']');
    code = code.replace(/：/g, ':');
    code = code.replace(/"/g, '"').replace(/"/g, '"');
    if (!code.trim().startsWith('graph') && !code.trim().startsWith('flowchart')) {{
        code = 'graph TD\\n' + code;
    }}
    return code;
}}

function buildFlowView() {{
    const flows = DATA.flow_diagrams || [];
    if (!flows.length) return '<p style="padding:20px;color:#999;">暂无流程图数据</p>';

    let html = '';
    flows.forEach((f, i) => {{
        const mermaidCode = cleanMermaid(f.mermaid_code || '');
        const desc = f.description || f.trigger || '';
        html += `
            <div class="flow-card">
                <div class="flow-header" onclick="toggleFlowBody(${{i}})">
                    <span class="flow-title">${{f.name || '流程图' + i}}</span>
                    <span class="flow-meta">${{desc}}</span>
                    <span class="flow-toggle">展开</span>
                </div>
                <div class="flow-body" id="flow-body-${{i}}" style="display:none;">
                    <div class="mermaid">${{mermaidCode}}</div>
                </div>
            </div>`;
    }});
    return html;
}}

function toggleFlowBody(idx) {{
    const body = document.getElementById('flow-body-' + idx);
    const header = body.previousElementSibling;
    const toggle = header.querySelector('.flow-toggle');
    if (body.style.display === 'none') {{
        body.style.display = 'block';
        toggle.textContent = '折叠';
        if (typeof mermaid !== 'undefined') {{
            const mermaidDivs = body.querySelectorAll('.mermaid:not([data-processed])');
            if (mermaidDivs.length > 0) {{ try {{ mermaid.init(undefined, mermaidDivs); }} catch(e) {{}} }}
        }}
    }} else {{
        body.style.display = 'none';
        toggle.textContent = '展开';
    }}
}}

function buildTableView(id) {{
    let data, headers, keys, colWidths;
    if (id === 'state') {{
        data = DATA.state_scenarios || [];
        headers = ['前置状态','场景/操作','执行状态','HUD提示','灯光','语音','App提示','周期'];
        keys = ['pre_state','current','exec_state','hud','light','voice','app','cycle'];
        colWidths = ['12%','14%','12%','14%','10%','14%','14%','10%'];
    }} else if (id === 'voice') {{
        data = DATA.voice_commands || [];
        headers = ['分类','唤醒','用户说法','变体','系统动作','成功反馈','HUD反馈','失败反馈','优先级'];
        keys = ['category','wake','user_says','variants','action','success_voice','success_hud','fail_feedback','priority'];
        colWidths = ['10%','8%','14%','14%','14%','12%','10%','12%','6%'];
    }} else if (id === 'button') {{
        data = DATA.button_mapping || [];
        headers = ['按键','场景','单击','双击','长按','组合键','反馈','备注'];
        keys = ['button','scene','single_click','double_click','long_press','combo','feedback','note'];
        colWidths = ['10%','12%','14%','14%','14%','14%','14%','8%'];
    }} else if (id === 'light') {{
        data = DATA.light_effects || [];
        // B2: 增加 "作用灯区" 列
        headers = ['场景','颜色','模式','频率','时长','灯区','优先级','备注'];
        keys = ['trigger','color','mode','frequency','duration','lamp_zone','priority','note'];
        colWidths = ['12%','8%','10%','8%','10%','12%','8%','22%'];
    }} else if (id === 'voice_nav') {{
        data = DATA.voice_nav || [];
        headers = ['场景','触发','用户输入','AI动作','HUD显示','语音播报','灯光反馈','异常兜底','优先级','备注'];
        keys = ['scene','trigger','user_input','ai_action','hud_display','voice_output','light_effect','fallback','priority','note'];
        colWidths = ['10%','8%','12%','12%','12%','12%','10%','12%','6%','8%'];
    }} else if (id === 'user_stories') {{
        data = DATA.user_stories || [];
        headers = ['功能名','角色','用户故事','验收条件','优先级'];
        keys = ['feature','role','story','acceptance','priority'];
        colWidths = ['16%','12%','30%','28%','14%'];
    }} else if (id === 'test_cases') {{
        data = DATA.test_cases || [];
        headers = ['用例ID','功能名','标题','前置条件','步骤','预期结果','优先级'];
        keys = ['case_id','feature','title','precondition','steps','expected','priority'];
        colWidths = ['10%','14%','16%','14%','20%','18%','8%'];
    }} else if (id === 'page_mapping') {{
        data = DATA.page_mapping || [];
        headers = ['页面','父页面','平台','功能','入口','优先级','备注'];
        keys = ['page','parent','platform','features','entry','priority','note'];
        colWidths = ['14%','12%','10%','20%','14%','8%','22%'];
    }} else if (id === 'dev_tasks') {{
        data = DATA.dev_tasks || [];
        headers = ['任务ID','功能','描述','角色','工时','依赖','迭代','优先级','备注'];
        keys = ['task_id','feature','task','assignee','effort_days','dependency','sprint','priority','note'];
        colWidths = ['10%','12%','20%','10%','8%','12%','10%','6%','22%'];
    }} else if (id === 'journey') {{
        // B5: 用户旅程表
        data = DATA.user_journeys || [];
        headers = ['角色','画像','阶段数','触点数','Mermaid代码'];
        keys = ['role','persona','stages_count','touchpoints_count','mermaid_code'];
        colWidths = ['18%','30%','8%','8%','36%'];
    }} else if (id === 'ai_scenarios') {{
        // B5: AI场景表
        data = DATA.ai_scenarios || [];
        headers = ['场景名称','触发条件','系统动作','所需数据','用户控制'];
        keys = ['scenario','trigger','action','data_needed','user_control'];
        colWidths = ['18%','25%','25%','18%','14%'];
    }} else {{
        data = [];
    }}

    if (!data || data.length === 0) return '<p style="padding:20px;color:#999;">暂无数据</p>';

    // 添加 colgroup 定义列宽
    let colgroupHtml = '<colgroup>';
    (colWidths || []).forEach(w => {{ colgroupHtml += `<col style="width:${{w}}">`; }});
    colgroupHtml += '</colgroup>';

    let html = '<div class="table-wrapper"><table>' + colgroupHtml + '<thead><tr>';
    headers.forEach(h => {{ html += `<th>${{h}}</th>`; }});
    html += '</tr></thead><tbody>';
    data.forEach(row => {{
        html += '<tr>';
        keys.forEach(k => {{
            let val = row[k] || '';
            if (k === 'priority' && val) val = `<span class="tag tag-${{val}}">${{val}}</span>`;
            html += `<td>${{val}}</td>`;
        }});
        html += '</tr>';
    }});
    html += '</tbody></table></div>';
    return html;
}}

function filterAll() {{
    const q = document.getElementById('search').value.toLowerCase();
    const p = document.getElementById('priorityFilter').value;

    document.querySelectorAll('.l1').forEach(el => {{
        const name = (el.dataset.name || '').toLowerCase();
        const priority = el.dataset.priority || '';
        const matchQ = !q || name.includes(q) || el.textContent.toLowerCase().includes(q);
        const matchP = !p || priority === p;
        el.classList.toggle('hidden', !(matchQ && matchP));
    }});

    document.querySelectorAll('.l2, .l3').forEach(el => {{
        const name = (el.dataset.name || '').toLowerCase();
        const priority = el.dataset.priority || '';
        const matchQ = !q || name.includes(q);
        const matchP = !p || priority === p;
        el.classList.toggle('hidden', !(matchQ && matchP));
    }});

    document.querySelectorAll('tbody tr').forEach(tr => {{
        const text = tr.textContent.toLowerCase();
        const matchQ = !q || text.includes(q);
        const tagEl = tr.querySelector('.tag');
        const rowP = tagEl ? tagEl.textContent.trim() : '';
        const matchP = !p || rowP === p;
        tr.classList.toggle('hidden', !(matchQ && matchP));
    }});
}}

function expandAllTrees() {{
    document.querySelectorAll('.l1-body').forEach(b => b.classList.add('open'));
}}
function collapseAllTrees() {{
    document.querySelectorAll('.l1-body').forEach(b => b.classList.remove('open'));
}}

function updateHeaderHeight() {{
    const header = document.querySelector('.header');
    const tabs = document.querySelector('.tabs');
    const content = document.querySelector('.content');

    if (header && tabs && content) {{
        const headerH = header.offsetHeight;
        const tabsH = tabs.offsetHeight;
        const totalH = headerH + tabsH;

        // 设置 CSS 变量
        document.documentElement.style.setProperty('--header-height', headerH + 'px');
        document.documentElement.style.setProperty('--total-header-height', totalH + 'px');

        // 更新 tabs 的 top 位置
        tabs.style.top = headerH + 'px';

        // 更新 content 的 margin-top 和高度
        content.style.marginTop = totalH + 'px';
        content.style.height = `calc(100vh - ${{totalH}}px)`;

        // 更新表格表头的 sticky top
        document.querySelectorAll('th').forEach(th => {{
            th.style.top = '0';
        }});
    }}
}}

function updateStickyPositions() {{
    updateHeaderHeight();
}}

function updateStatsBar() {{
    const allFeatures = DATA.features;
    const total = allFeatures.length;
    const p0 = allFeatures.filter(f => f.priority === 'P0').length;
    const p1 = allFeatures.filter(f => f.priority === 'P1').length;
    const p2 = allFeatures.filter(f => f.priority === 'P2').length;
    const p3 = allFeatures.filter(f => f.priority === 'P3').length;
    const modules = new Set(allFeatures.map(f => f.module)).size;

    const bar = document.getElementById('statsBar');
    if (bar) {{
        bar.innerHTML = `
            <div class="stat">功能总数 <span class="stat-value">${{total}}</span></div>
            <div class="stat">P0 <span class="stat-value">${{p0}}</span></div>
            <div class="stat">P1 <span class="stat-value">${{p1}}</span></div>
            <div class="stat">P2 <span class="stat-value">${{p2}}</span></div>
            <div class="stat">P3 <span class="stat-value">${{p3}}</span></div>
            <div class="stat">模块数 <span class="stat-value">${{modules}}</span></div>
        `;
    }}
}}

buildTabs();
updateStatsBar();
updateHeaderHeight();
// 修复 2: 确保在 DOM 完全加载后执行
document.addEventListener('DOMContentLoaded', function() {{
    setTimeout(updateHeaderHeight, 100);
}});
window.addEventListener('load', function() {{
    setTimeout(updateHeaderHeight, 200);
}});
window.addEventListener('resize', updateHeaderHeight);

const firstBody = document.querySelector('.l1-body');
if (firstBody) firstBody.classList.add('open');
</script>

<!-- Mermaid.js for flow diagrams -->
<script src="https://cdn.jsdelivr.net/npm/mermaid@9.4.3/dist/mermaid.min.js"></script>
<script>
(function() {{
    let mermaidReady = false;
    function tryInit() {{
        if (typeof mermaid !== 'undefined' && !mermaidReady) {{
            mermaid.initialize({{ startOnLoad: false, theme: 'neutral', securityLevel: 'loose', flowchart: {{ useMaxWidth: true }} }});
            mermaidReady = true;
        }}
    }}
    function renderFlows() {{
        tryInit();
        if (mermaidReady) {{
            try {{ mermaid.init(undefined, '.mermaid'); }} catch(e) {{ console.error(e); }}
        }} else {{
            document.querySelectorAll('.mermaid').forEach(el => {{
                el.style.cssText = 'background:#f7f7f7;padding:16px;font-family:monospace;font-size:12px;white-space:pre-wrap;border:1px solid #ddd;border-radius:4px;';
                const tip = document.createElement('div');
                tip.style.cssText = 'color:#999;margin-bottom:8px;';
                tip.textContent = 'Flow diagram needs network (Mermaid.js). Copy code to mermaid.live to render.';
                el.parentNode.insertBefore(tip, el);
            }});
        }}
    }}
    function toggleFlow(idx) {{
        const body = document.getElementById('flow-body-' + idx);
        const hdr = body.previousElementSibling;
        const tog = hdr.querySelector('.flow-toggle');
        if (body.style.display === 'none') {{
            body.style.display = 'block';
            tog.textContent = 'Hide';
            if (mermaidReady) {{
                const mermaidDivs = body.querySelectorAll('.mermaid:not([data-processed])');
                if (mermaidDivs.length > 0) {{ try {{ mermaid.init(undefined, mermaidDivs); }} catch(e) {{}} }}
            }}
        }} else {{
            body.style.display = 'none';
            tog.textContent = 'Show';
        }}
    }}
    const origSwitchTab = window.switchTab;
    window.switchTab = function(id) {{
        origSwitchTab(id);
        if (id === 'flow') {{ setTimeout(renderFlows, 100); }}
    }};
    window.addEventListener('load', tryInit);
}})();
</script>
</body></html>""";

    return html


# === Sheet 3-6 生成函数 ===
def _gen_sheet3_state_scenarios(gateway, kb_context: str = "", goal_text: str = "") -> List[Dict]:
    """生成 Sheet 3: 状态场景对策表"""

    prompt = (
        "为智能摩托车全盔项目生成【状态场景对策表】.\n"
        "输出 JSON 数组,每个元素格式:\n"
        '{"pre_state":"前置状态","current":"当前状态/场景/操作",'
        '"exec_state":"执行状态","hud":"HUD提示内容",'
        '"light":"灯光提示(颜色+模式)","voice":"语音提示内容",'
        '"app":"App提示内容","cycle":"提示周期/时长"}\n\n'
        "必须覆盖以下场景分类(每类至少5条):\n"
        "1. 开关机与启动:开机自检、关机确认、低电自动关机\n"
        "2. 蓝牙连接:首次配对、自动回连、断连、回连成功、回连失败\n"
        "3. 导航:开始导航、转向提醒、偏航重算、到达目的地、导航取消、GPS弱信号\n"
        "4. 来电通话:来电提醒、接听、拒接、挂断、通话中断\n"
        "5. 录制:开始录制、停止录制、存储满、过热降级、录制异常中断\n"
        "6. 组队:创建队伍、加入队伍、掉队提醒、队友离线、退出队伍\n"
        "7. 胎压:正常、低压警告、高温警告、传感器离线\n"
        "8. 电量:充电中、低电量20%、极低电量10%、充电完成\n"
        "9. 安全预警:前向碰撞、侧后来车、盲区占用、急弯减速\n"
        "10. OTA:检测到新版本、下载中、升级中、升级成功、升级失败\n"
        "11. SOS:疑似事故检测、倒计时确认、触发报警、误触取消\n"
        "12. AI语音:唤醒成功、识别中、执行成功、识别失败、网络不可用\n\n"
        "目标 60-80 条.只输出 JSON 数组.\n"
    )

    if kb_context:
        prompt += f"\n参考内部文档:\n{kb_context[:2000]}\n"

    result = gateway.call_azure_openai("cpo", prompt,
        "只输出JSON数组,不要其他文字.", "structured_doc")

    if result.get("success"):
        json_match = re.search(r'\[[\s\S]*\]', result["response"])
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass
    return []


def _gen_sheet4_voice_commands(gateway, kb_context: str = "") -> List[Dict]:
    """生成 Sheet 4: 语音指令表"""

    prompt = (
        "为智能摩托车全盔项目生成【语音指令表】.\n"
        "输出 JSON 数组,每个元素格式:\n"
        '{"category":"指令分类","wake":"唤醒方式(唤醒词/按键/自动)",'
        '"user_says":"用户说法","variants":"常见变体说法(逗号分隔)",'
        '"action":"系统执行动作","success_voice":"成功语音反馈",'
        '"success_hud":"成功HUD反馈","fail_feedback":"失败反馈",'
        '"priority":"P0-P3","note":"备注"}\n\n'
        "必须覆盖以下分类(每类至少5条指令):\n"
        "1. 导航控制:开始导航、取消导航、查询距离/时间、切换路线、回家\n"
        "2. 通话控制:接听、拒接、挂断、回拨、打给XX\n"
        "3. 音乐控制:播放、暂停、上一首、下一首、调大/小音量、静音\n"
        "4. 录制控制:开始录制、停止录制、拍照、标记片段\n"
        "5. 组队控制:创建队伍、加入队伍、退出队伍、呼叫队友\n"
        "6. 设备查询:查电量、查存储、查胎压、查速度、查时间\n"
        "7. 设置控制:调亮度、切模式、开/关降噪、开/关免打扰\n"
        "8. 安全相关:紧急求救、取消报警\n\n"
        "每条指令必须给出至少2个用户常见变体说法.\n"
        "目标 40-60 条.只输出 JSON 数组.\n"
    )

    result = gateway.call_azure_openai("cpo", prompt,
        "只输出JSON数组.", "structured_doc")

    if result.get("success"):
        json_match = re.search(r'\[[\s\S]*\]', result["response"])
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass
    return []


def _gen_sheet5_button_mapping(gateway) -> List[Dict]:
    """生成 Sheet 5: 按键映射表"""

    prompt = (
        "为智能摩托车全盔项目生成【实体按键场景矩阵表】.\n"
        "头盔有以下按键:主按键(侧面)、音量+键、音量-键、功能键(可选).\n"
        "同一个按键在不同场景下动作不同.\n\n"
        "输出 JSON 数组,每个元素代表一个按键在一个场景下的完整映射:\n"
        '{"button":"按键位置","scene":"场景(通用/导航中/通话中/录制中/组队中/音乐播放中/来电响铃中/语音助手激活中)",'
        '"single_click":"单击动作","double_click":"双击动作",'
        '"long_press":"长按动作(>2秒)","combo":"组合键动作",'
        '"feedback":"操作反馈(震动/语音/HUD/灯光)","note":"备注"}\n\n'
        "规则:\n"
        "1. 每个按键必须在所有 8 个场景下都有定义(4 按键 x 8 场景 = 32 条)\n"
        "2. 同一操作在不同场景可以含义不同(如通用单击=播放暂停,导航中单击=确认路线)\n"
        "3. 未定义的动作填'同通用'或'无'\n"
        "4. 每条必须标明操作反馈方式\n"
        "5. 骑行中戴手套操作,动作必须简单明确\n\n"
        "目标 32-40 条.只输出 JSON 数组.\n"
    )

    # B2: 诊断日志
    print(f"[Button-Diag] prompt长度: {len(prompt)}")

    result = gateway.call_azure_openai("cpo", prompt,
        "只输出JSON数组.", "structured_doc")

    # B2: 诊断日志
    print(f"[Button-Diag] 响应长度: {len(result.get('response', '')) if result.get('success') else 0}")
    print(f"[Button-Diag] 响应前200字: {result.get('response', '')[:200] if result.get('success') else 'EMPTY'}")

    if result.get("success"):
        json_match = re.search(r'\[[\s\S]*\]', result["response"])
        if json_match:
            try:
                data = json.loads(json_match.group())
                print(f"[Button-Diag] 解析成功: {len(data)} 条")
                return data
            except json.JSONDecodeError as e:
                print(f"[Button-Diag] JSON解析失败: {e}")
    print("[Button-Diag] 生成失败,返回空列表")
    return []


def _gen_sheet6_light_effects(gateway) -> List[Dict]:
    """生成 Sheet 6: 灯效定义表"""

    # B2: 增加 lamp_zone 字段
    prompt = (
        "为智能摩托车全盔项目生成【氛围灯灯效定义表】.\n"
        "头盔有多区域氛围灯:后脑勺灯(制动/转向/后向告警)、眼下方灯(状态/通知反馈)、隐私灯(录像指示).\n"
        "输出 JSON 数组,每个元素格式:\n"
        '{"trigger":"触发场景","color":"灯光颜色(如红/蓝/绿/白/橙/紫)",'
        '"mode":"闪烁模式(常亮/慢闪/快闪/呼吸/流水/脉冲)",'
        '"frequency":"频率(Hz,常亮填0)","duration":"持续时长",'
        '"priority":"优先级(P0-P3,高优先级覆盖低优先级)",'
        '"lamp_zone":"作用灯区(后脑勺灯/眼下方灯/双灯区/隐私灯)",'
        '"note":"备注"}\n\n'
        "必须覆盖:\n"
        "1. 系统状态:开机、关机、充电中、充电完成、低电量、OTA中\n"
        "2. 连接状态:蓝牙配对中、配对成功、断连、回连\n"
        "3. 安全预警:前向碰撞(红快闪,后脑勺灯)、侧后来车(橙方向闪,后脑勺灯)、盲区(黄,后脑勺灯)\n"
        "4. 通信提醒:来电(蓝呼吸,眼下方灯)、消息(白单闪,眼下方灯)、组队成功(绿,双灯区)\n"
        "5. 录制状态:录制中(红微亮,隐私灯)、录制暂停、录制异常\n"
        "6. 骑行辅助:刹车灯效(红常亮,后脑勺灯)、转向灯效(橙流水,后脑勺灯)、掉队提醒\n"
        "7. 特殊场景:SOS(红蓝交替快闪,双灯区)、夜骑尾灯模式\n\n"
        "灯效优先级规则:安全>通信>系统>装饰.\n"
        "lamp_zone 必须明确标注,不漏.\n"
        "目标 25-35 条.只输出 JSON 数组.\n"
    )

    result = gateway.call_azure_openai("cpo", prompt,
        "只输出JSON数组.", "structured_doc")

    if result.get("success"):
        json_match = re.search(r'\[[\s\S]*\]', result["response"])
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass
    return []


def _gen_sheet7_voice_nav_scenarios(gateway) -> List[Dict]:
    """生成 Sheet 7: AI语音导航场景表"""

    prompt = (
        "为智能摩托车全盔项目生成【AI语音导航场景定义表】.\n"
        "头盔集成AI语音助手,支持语音导航控制、POI查询、路线规划.\n"
        "输出 JSON 数组,每个元素格式:\n"
        '{"scene":"场景名称","trigger":"触发条件","user_input":"用户输入(语音/按键/自动)",'
        '"ai_action":"AI执行动作","hud_display":"HUD显示内容与样式",'
        '"voice_output":"语音播报内容","light_effect":"灯光反馈",'
        '"fallback":"异常兜底策略","priority":"P0-P3","note":"备注"}\n\n'
        "必须覆盖:\n"
        "1. 导航启动:开始导航、去某地、导航到XXX\n"
        "2. 路线控制:重新规划、避开高速、最快路线、最短路线\n"
        "3. POI查询:附近加油站、找餐厅、最近充电站、搜索便利店\n"
        "4. 路况播报:前方路况、拥堵情况、事故提醒、封路信息\n"
        "5. 导航暂停/继续:暂停导航、继续导航、取消导航\n"
        "6. 途经点:添加途经点、绕道去某地、修改目的地\n"
        "7. 特殊场景:偏离路线重新规划、到达提醒、剩余距离查询\n"
        "8. 多模态协同:地图显示切换、HUD投影导航、语音播报开关\n\n"
        "响应要简洁(骑行场景),ai_action要具体可执行.\n"
        "目标 20-30 条.只输出 JSON 数组.\n"
    )

    result = gateway.call_azure_openai("cpo", prompt,
        "只输出JSON数组.", "structured_doc")

    if result.get("success"):
        json_match = re.search(r'\[[\s\S]*\]', result["response"])
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass
    return []


def _gen_user_stories(items: List[Dict], gateway) -> List[Dict]:
    """基于功能清单生成用户故事"""

    # 提取所有 L2 功能
    l2_features = [i for i in items if i.get("level") == "L2"]
    batch_size = 30

    # 构建所有批次
    batches = []
    for start in range(0, len(l2_features), batch_size):
        batches.append((l2_features[start:start + batch_size], start // batch_size))

    def _gen_batch(batch_info):
        batch, batch_idx = batch_info
        features_text = "\n".join(
            f"- {f.get('name', '')}({f.get('module', '')}): {f.get('description', '')}"
            for f in batch
        )
        prompt = (
            f"为以下功能生成用户故事(User Story).\n"
            f"输出 JSON 数组,每个元素:\n"
            f'{{"feature":"功能名","role":"用户角色","story":"作为[角色],我想要[功能],以便[价值]",'
            f'"acceptance":"验收条件","priority":"P0-P3"}}\n\n'
            f"角色库(选最合适的):通勤骑手、摩旅骑手、团骑领队、内容创作骑手、新手骑手、后台管理员\n\n"
            f"功能列表:\n{features_text}\n\n"
            f"每个功能一条故事.只输出 JSON."
        )
        result = gateway.call_azure_openai("cpo", prompt,
            "生成用户故事.只输出JSON.", "structured_doc")
        if result.get("success"):
            json_match = re.search(r'\[[\s\S]*\]', result["response"])
            if json_match:
                try:
                    return json.loads(json_match.group())
                except json.JSONDecodeError:
                    pass
        return []

    # 8 路并行处理所有批次
    all_stories = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futs = {pool.submit(_gen_batch, b): i for i, b in enumerate(batches)}
        for f in as_completed(futs):
            all_stories.extend(f.result())

    print(f"[UserStory] 生成 {len(all_stories)} 条用户故事")
    return all_stories


def _gen_test_cases(items: List[Dict], gateway) -> List[Dict]:
    """基于 L3 验收标准生成测试用例"""

    l3_features = [i for i in items if i.get("level") == "L3" and i.get("acceptance")]
    batch_size = 30

    # 构建所有批次
    batches = []
    for start in range(0, len(l3_features), batch_size):
        batches.append((l3_features[start:start + batch_size], start // batch_size))

    def _gen_batch(batch_info):
        batch, batch_idx = batch_info
        features_text = "\n".join(
            f"- {f.get('name', '')}: 验收={f.get('acceptance', '')}"
            for f in batch
        )
        prompt = (
            f"为以下功能的验收标准生成测试用例.\n"
            f"输出 JSON 数组,每个元素:\n"
            f'{{"case_id":"TC-001","feature":"功能名","title":"用例标题",'
            f'"precondition":"前置条件","steps":"操作步骤(分号分隔)",'
            f'"expected":"预期结果","priority":"P0-P3"}}\n\n'
            f"规则:\n"
            f"- 每个功能 1-2 条用例(正常流程 + 异常流程)\n"
            f"- 操作步骤要具体可执行\n"
            f"- 预期结果要包含验收标准中的具体数字\n\n"
            f"功能列表:\n{features_text}\n\n"
            f"只输出 JSON."
        )
        result = gateway.call_azure_openai("cpo", prompt,
            "生成测试用例.只输出JSON.", "structured_doc")
        if result.get("success"):
            json_match = re.search(r'\[[\s\S]*\]', result["response"])
            if json_match:
                try:
                    cases = json.loads(json_match.group())
                    for idx, c in enumerate(cases):
                        c["case_id"] = f"TC-{batch_idx * batch_size + idx + 1:04d}"
                    return cases
                except json.JSONDecodeError:
                    pass
        return []

    # 8 路并行处理所有批次
    all_cases = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futs = {pool.submit(_gen_batch, b): i for i, b in enumerate(batches)}
        for f in as_completed(futs):
            all_cases.extend(f.result())

    print(f"[TestCase] 生成 {len(all_cases)} 条测试用例")
    return all_cases


def _gen_page_mapping(items: List[Dict], gateway) -> List[Dict]:
    """从功能清单生成页面→功能映射表"""

    # 提取所有 note 中的页面信息
    pages_mentioned = set()
    for item in items:
        note = item.get("note", "")
        if note:
            # 提取页面关键词
            for page in re.findall(r'(骑行主界面|App-\S+|首次配对|设置页|导航页|录制页|组队页|商城\S*|社区\S*)', note):
                pages_mentioned.add(page)

    # 如果 note 中没有足够页面信息,让 LLM 生成
    l1_l2 = [i for i in items if i.get("level") in ("L1", "L2")]
    features_text = "\n".join(f"- {f.get('name', '')}({f.get('module', '')})" for f in l1_l2[:60])

    # 修复 7: 扩充页面映射表 prompt
    prompt = (
        f"基于以下智能摩托车全盔的功能列表,生成页面→功能映射表.\n"
        f"输出 JSON 数组,每个元素代表一个页面:\n"
        f'{{"page":"页面名","parent":"父页面(顶级填空)","platform":"HUD/App/系统",'
        f'"features":"该页面包含的功能(逗号分隔)","entry":"入口方式(Tab/按钮/自动/语音)",'
        f'"priority":"P0-P3","note":"备注"}}\n\n'
        f"【App 端页面要求】每个 Tab 至少有：首页 + 2-3 个二级页面 + 关键三级页面\n"
        f"- 设备Tab: 设备首页/蓝牙配对页/设备详情页/骑行轨迹列表/轨迹详情/行车记录列表/视频播放页/摄像参数设置/部件信息页/HUD布局设置/场景模式设置\n"
        f"- 相册Tab: 相册首页/照片详情/视频播放/水印编辑/批量导出\n"
        f"- 社区Tab: 社区首页/帖子详情/发布页/个人主页/粉丝列表/等级说明\n"
        f"- AI Tab: AI首页/AI对话/AI剪片编辑/AI内容搜索结果/AI旅行总结/AI骑行摘要\n"
        f"- 我的Tab: 我的首页/账号信息编辑/通用设置/通知设置/帮助FAQ/反馈提交/关于设备/隐私政策/数据管理/多语言设置\n"
        f"- 通知中心: 通知列表/通知详情\n"
        f"- 商城: 商城首页/商品详情/购物车/订单确认/支付/订单列表/订单详情/报修工单提交/工单详情/工单列表\n"
        f"- 其他: 登录注册页/权限申请引导页/首次配对引导页/新手教学流程(多步)/用户成就页/成就详情\n\n"
        f"【HUD 端页面要求】HUD 不是传统页面，但有不同的'显示状态/卡片组合':\n"
        f"- 主驾驶视图/导航视图/来电视图/音乐视图/组队视图/消息视图/简易模式视图\n"
        f"- 信息中岛各状态\n"
        f"- 各场景模式的 HUD 布局\n\n"
        f"每行包含: 页面名、父页面、平台(App/HUD)、包含功能、入口方式、优先级、备注\n\n"
        f"目标: App 至少 50 个页面 + HUD 至少 10 个视图 = 60+ 条\n\n"
        f"功能模块列表:\n{features_text}\n\n"
        f"只输出 JSON."
    )

    result = gateway.call_azure_openai("cpo", prompt,
        "生成页面映射.只输出JSON.", "structured_doc")

    if result.get("success"):
        json_match = re.search(r'\[[\s\S]*\]', result["response"])
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass
    return []


def _gen_dev_tasks(items: List[Dict], gateway) -> List[Dict]:
    """基于功能清单生成开发任务清单(含工时估算)"""

    l2_features = [i for i in items if i.get("level") == "L2"]
    batch_size = 30

    # 构建所有批次
    batches = []
    for start in range(0, len(l2_features), batch_size):
        batches.append((l2_features[start:start + batch_size], start // batch_size))

    def _gen_batch(batch_info):
        batch, batch_idx = batch_info
        features_text = "\n".join(
            f"- {f.get('name', '')}({f.get('module', '')}): {f.get('description', '')[:60]}"
            for f in batch
        )
        prompt = (
            f"为以下功能生成开发任务清单(含工时估算).\n"
            f"输出 JSON 数组,每个元素:\n"
            f'{{"task_id":"T-001","feature":"功能名","task":"任务描述",'
            f'"assignee":"负责角色(前端/后端/嵌入式/算法/测试/设计)",'
            f'"effort_days":"预估工时(天)","dependency":"前置依赖任务",'
            f'"sprint":"建议迭代(Sprint1-MVP/Sprint2-增强/Sprint3-优化)",'
            f'"priority":"P0-P3","note":"备注"}}\n\n'
            f"工时估算规则:\n"
            f"- 简单UI展示: 0.5-1天\n"
            f"- 标准功能开发: 1-3天\n"
            f"- 复杂交互/算法: 3-5天\n"
            f"- 跨端联调: 2-3天\n\n"
            f"【Bug 5 修复: 任务数量要求】\n"
            f"- 每个 L2 功能拆成 2-4 个开发任务（不是1-3个）\n"
            f"- 每个 L1 模块至少生成 5 条开发任务\n"
            f"- 优先级为 P0/P1 的功能必须拆成 3 个以上任务\n"
            f"- 总任务数目标：不少于 700 条（功能总数约 500，开发任务应为 1.5 倍）\n\n"
            f"功能列表:\n{features_text}\n\n"
            f"只输出 JSON."
        )
        result = gateway.call_azure_openai("cpo", prompt,
            "生成开发任务.只输出JSON.", "structured_doc")
        if result.get("success"):
            json_match = re.search(r'\[[\s\S]*\]', result["response"])
            if json_match:
                try:
                    tasks = json.loads(json_match.group())
                    for idx, t in enumerate(tasks):
                        t["task_id"] = f"T-{batch_idx * batch_size + idx + 1:04d}"
                    return tasks
                except json.JSONDecodeError:
                    pass
        return []

    # 8 路并行处理所有批次
    all_tasks = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futs = {pool.submit(_gen_batch, b): i for i, b in enumerate(batches)}
        for f in as_completed(futs):
            all_tasks.extend(f.result())

    print(f"[DevTask] 生成 {len(all_tasks)} 条开发任务")
    return all_tasks


def _cross_validate(all_rows: List[Dict], sheets_data: Dict) -> List[str]:
    """跨 Sheet 交叉验证，检查覆盖率和一致性"""
    issues = []

    # 1. P0 功能 → 测试用例覆盖
    p0_names = set(r.get('name', '') for r in all_rows if r.get('priority') == 'P0')
    tc_names = set()
    for tc in sheets_data.get('test_cases', []):
        tc_names.add(str(tc.get('功能名', tc.get('feature', tc.get('name', '')))))
    uncovered_p0 = p0_names - tc_names
    if uncovered_p0 and len(uncovered_p0) > 3:
        issues.append(f"[P0-测试] {len(uncovered_p0)} 个 P0 功能无测试用例覆盖")

    # 2. 功能模块 → 开发任务对应
    all_modules = set(r.get('module', r.get('L1功能', r.get('name', ''))) for r in all_rows if r.get('level') in ('L1', 'L2'))
    dev_modules = set()
    for dt in sheets_data.get('dev_tasks', []):
        dev_modules.add(str(dt.get('功能名', dt.get('feature', dt.get('module', '')))))
    uncovered_modules = all_modules - dev_modules
    if uncovered_modules and len(uncovered_modules) > 5:
        issues.append(f"[功能-开发] {len(uncovered_modules)} 个模块无开发任务")

    # 3. 用户故事 vs 功能数量检查
    story_count = len(sheets_data.get('user_stories', []))
    feature_count = len(all_rows)
    ratio = story_count / max(feature_count, 1)
    if ratio < 0.2:
        issues.append(f"[用户故事] 仅 {story_count} 条（功能的 {ratio:.0%}），偏少")

    if issues:
        print(f"\n[CrossValidate] 发现 {len(issues)} 个交叉问题:")
        for issue in issues:
            print(f"  ! {issue}")
    else:
        print(f"\n[CrossValidate] ✅ 跨 Sheet 一致性检查通过")

    return issues


def _extract_data_points(kb_text: str, max_points: int = 15) -> str:
    """从知识库全文中提取关键数据点（纯正则，不走 LLM）"""
    import re as _re

    points = set()

    # 提取含数字+单位的句子片段
    num_patterns = _re.findall(
        r'[^。\n]{0,40}?\d+\.?\d*\s*(?:mm|cm|g|kg|mAh|W|V|Hz|dB|美元|元|USD|\$|%|nits|lux|fps|°|μm|TOPS|nm|GHz|MHz|MB|GB|TB|ms|秒|分钟|小时|天|个月|台|件|条|款|代|层|路|位|倍|次)[^。\n]{0,20}',
        kb_text
    )
    for match in num_patterns[:20]:
        clean = match.strip().strip('，,、；;：:')
        if len(clean) > 10:
            points.add(clean)

    # 提取含型号的句子片段
    model_patterns = _re.findall(
        r'[^。\n]{0,30}?(?:[A-Z]{2,}\d{2,}|IMX\d+|QCC\d+|BES\d+|nRF\d+|AR[12]\s*Gen|BMI\d+|ICM-\d+|STM32|ESP32|MT\d{4}|BCM\d+|CS\d{4}|WM\d{4})[^。\n]{0,30}',
        kb_text
    )
    for match in model_patterns[:10]:
        clean = match.strip().strip('，,、；;：:')
        if len(clean) > 8:
            points.add(clean)

    # 提取含品牌名的句子片段
    brand_patterns = _re.findall(
        r'[^。\n]{0,20}?(?:歌尔|立讯|舜宇|索尼|高通|联发科|博世|Qualcomm|Sony|Bosch|Nordic|Himax|JBD|Sena|Cardo|Forcite|LIVALL|EyeRide|CrossHelmet|GoPro|Insta360|TÜV|DEKRA|SGS)[^。\n]{0,30}',
        kb_text
    )
    for match in brand_patterns[:10]:
        clean = match.strip().strip('，,、；;：:')
        if len(clean) > 8:
            points.add(clean)

    # 提取含认证标准的
    cert_patterns = _re.findall(
        r'[^。\n]{0,20}?(?:ECE\s*22\.0[56]|DOT\s*FMVSS|GB\s*811|FCC|CE|IP\d{2}|MIL-STD|Qi|BQB|SRRC)[^。\n]{0,30}',
        kb_text
    )
    for match in cert_patterns[:5]:
        clean = match.strip().strip('，,、；;：:')
        if len(clean) > 8:
            points.add(clean)

    # 去重并截断
    result = list(points)[:max_points]

    if not result:
        return ""

    return "关键数据点：\n" + "\n".join(f"- {p}" for p in result)


import random

def _sample_compare(old_rows: list, new_rows: list, sample_size: int = 5) -> str:
    """
    抽样比质量：从新旧版各抽 N 条同名 L3 功能，逐条比较验收标准质量。
    返回 'new' / 'keep' / 'merge'
    """

    def _ensure_str(val):
        if isinstance(val, list):
            return ', '.join(str(v) for v in val)
        if isinstance(val, dict):
            import json
            return json.dumps(val, ensure_ascii=False)
        if val is None:
            return ''
        return str(val)

    def _quality_score_single(row: dict) -> float:
        """单条功能的质量评分"""
        score = 0

        acc = _ensure_str(row.get('acceptance', ''))
        desc = _ensure_str(row.get('description', ''))

        # 验收标准质量（权重 60%）
        import re
        numbers = re.findall(r'\d+\.?\d*', acc)
        if len(numbers) >= 2:
            score += 6  # 有 2 个以上具体数字
        elif len(numbers) >= 1:
            score += 3

        if any(unit in acc for unit in ['ms', 's', 'Hz', 'fps', 'dB', '%', '°', 'km', 'mAh', 'lux']):
            score += 2  # 有工程单位

        if '待验证' in acc or '待确认' in acc:
            score -= 1

        if len(acc) > 50:
            score += 1  # 足够详细

        # 描述质量（权重 20%）
        if len(desc) > 30:
            score += 2

        # 优先级合理性（权重 10%）
        priority = str(row.get('priority', ''))
        if priority in ['P0', 'P1', 'P2', 'P3']:
            score += 1

        return score

    # 构建同名功能映射
    old_map = {}
    for r in old_rows:
        name = _ensure_str(r.get('name', ''))
        if name:
            old_map[name] = r

    new_map = {}
    for r in new_rows:
        name = _ensure_str(r.get('name', ''))
        if name:
            new_map[name] = r

    # 找到两版都有的同名功能
    common_names = list(set(old_map.keys()) & set(new_map.keys()))

    if not common_names:
        # 没有重叠 → 比总条目数和平均质量
        old_avg = sum(_quality_score_single(r) for r in old_rows) / max(len(old_rows), 1)
        new_avg = sum(_quality_score_single(r) for r in new_rows) / max(len(new_rows), 1)
        if new_avg > old_avg + 1:
            return 'new'
        elif old_avg > new_avg + 1:
            return 'keep'
        return 'merge'

    # 抽样比较
    sample = random.sample(common_names, min(sample_size, len(common_names)))

    new_wins = 0
    old_wins = 0

    for name in sample:
        old_score = _quality_score_single(old_map[name])
        new_score = _quality_score_single(new_map[name])

        if new_score > old_score:
            new_wins += 1
        elif old_score > new_score:
            old_wins += 1
        # 平局不计

    # 判定
    total_compared = new_wins + old_wins
    if total_compared == 0:
        return 'merge'  # 全部平局

    if new_wins >= total_compared * 0.6:
        return 'new'
    elif old_wins >= total_compared * 0.6:
        return 'keep'
    else:
        return 'merge'


def _merge_best_of_both(old_rows: list, new_rows: list) -> list:
    """逐条取验收标准更好的版本"""

    def _ensure_str(val):
        if isinstance(val, list):
            return ', '.join(str(v) for v in val)
        if isinstance(val, dict):
            import json
            return json.dumps(val, ensure_ascii=False)
        if val is None:
            return ''
        return str(val)

    old_map = {_ensure_str(r.get('name', '')): r for r in old_rows if r.get('name')}
    new_map = {_ensure_str(r.get('name', '')): r for r in new_rows if r.get('name')}

    merged = {}
    all_names = set(old_map.keys()) | set(new_map.keys())

    for name in all_names:
        old_r = old_map.get(name)
        new_r = new_map.get(name)

        if old_r and not new_r:
            merged[name] = old_r
        elif new_r and not old_r:
            merged[name] = new_r
        else:
            # 两者都有 → 比验收标准长度和具体度
            old_acc = _ensure_str(old_r.get('acceptance', ''))
            new_acc = _ensure_str(new_r.get('acceptance', ''))
            import re
            old_nums = len(re.findall(r'\d+', old_acc))
            new_nums = len(re.findall(r'\d+', new_acc))
            if new_nums > old_nums:
                merged[name] = new_r
            elif old_nums > new_nums:
                merged[name] = old_r
            else:
                # 数字一样多 → 取更长的
                merged[name] = new_r if len(new_acc) >= len(old_acc) else old_r

    return list(merged.values())


def _score_module(items: list) -> int:
    """给一个模块的内容打分，用于对比两版取优（质量优先）"""
    if not items:
        return 0

    # ===== Bug A fix: 确保所有待正则匹配的文本字段是 string =====
    def _ensure_str(val):
        if isinstance(val, list):
            return ', '.join(str(v) for v in val)
        if isinstance(val, dict):
            import json
            return json.dumps(val, ensure_ascii=False)
        if val is None:
            return ''
        return str(val)

    for row in items:
        for key in ('acceptance', 'description', 'dependencies', 'interaction', 'name', 'priority'):
            if key in row:
                row[key] = _ensure_str(row[key])
    # ===== End Bug A fix =====

    l1 = sum(1 for i in items if i.get("level") == "L1")
    l2 = sum(1 for i in items if i.get("level") == "L2")
    l3 = sum(1 for i in items if i.get("level") == "L3")

    # 空壳：有 L1 但没有 L2
    if l1 > 0 and l2 == 0:
        return 1

    score = 0

    # 结构完整性（权重 30%）
    structure = min(l2 * 2 + l3 * 0.5, 30)
    score += structure

    # 验收标准质量（权重 50%）— 这是最重要的
    for item in items:
        acc = item.get("acceptance", "")
        desc = item.get("description", "")

        if not acc or acc == "[待生成]":
            score -= 3  # 空验收严重扣分
        elif "[待验证]" in acc:
            score += 1  # 标注待验证比编造好
        else:
            # 检查是否有真实数据（不是编的整数）
            import re
            has_specific = bool(re.search(
                r'\d+\.?\d*\s*(?:mm|cm|g|kg|mAh|W|V|Hz|dB|ms|秒|nits|lux|fps|°|GHz|MHz|Mbps|MB|GB)',
                acc
            ))
            has_model = bool(re.search(r'[A-Z]{2,}\d{2,}|ECE|DOT|GB\s*\d+', acc))

            if has_specific or has_model:
                score += 3  # 有具体参数/认证引用
            elif any(c.isdigit() for c in acc):
                score += 1.5  # 有数字但可能是编的
            else:
                score += 0.5  # 有文字但无数据

    # 描述完整性（权重 20%）
    has_desc = sum(1 for i in items if len(i.get("description", "")) > 15)
    score += min(has_desc * 0.5, 20)

    return max(int(score), 0)


# ===== Fix 4: 比较决策函数（改为抽样比质量）=====
def _compare_decision(module_name: str, old_rows: list, new_rows: list, old_score: int, new_score: int) -> str:
    """比较决策：使用抽样比质量替代总分比较

    Returns:
        'new': 使用新版
        'keep': 保留旧版
        'merge': 合并两版精华
        None: 需要重新生成
    """
    # 两者都是空/失败 → 都不要
    if old_score <= 1 and new_score <= 1:
        print(f"  SKIP {module_name}: 新旧都是空壳 (旧{old_score}, 新{new_score})")
        return None

    # 新版是空/失败 → 保留旧版
    if new_score <= 1:
        print(f"  KEEP {module_name}: 新版空壳 (旧{old_score})")
        return 'keep'

    # 旧版是空/失败 → 用新版
    if old_score <= 1:
        print(f"  OK {module_name}: 旧版空壳 (新{new_score})")
        return 'new'

    # 都有内容 → 使用抽样比质量
    decision = _sample_compare(old_rows, new_rows)

    if decision == 'new':
        print(f"  OK {module_name}: 新版质量更好 (抽样)")
        return 'new'
    elif decision == 'keep':
        print(f"  KEEP {module_name}: 旧版质量更好 (抽样)")
        return 'keep'
    else:
        print(f"  MERGE {module_name}: 逐条取优")
        return 'merge'


# ===== Layer 2: 替换模式比较函数 =====
def _compare_module_replace_mode(module_name: str, old_rows: list, new_rows: list, old_score: int, new_score: int) -> str:
    """模块级对比——择优替换，不做并集

    每轮生成是替换而非累加。新版更好就整体替换，旧版更好就保留不动。
    不吸收对方独有功能。

    Returns:
        'replace': 整体替换为新版
        'keep': 整体保留旧版
        None: 需要重新生成（两者都是空壳）
    """
    old_count = len(old_rows)
    new_count = len(new_rows)

    # Case 1: 旧版空壳（< 3 条），无条件用新版
    if old_count < 3:
        print(f"  [Compare] {module_name}: 旧版空壳 → 用新版 ({new_count}条)")
        return 'replace'

    # Case 2: 新版空壳或生成失败
    if new_count < 3:
        print(f"  [Compare] {module_name}: 新版空壳 → 保留旧版 ({old_count}条)")
        return 'keep'

    # Case 3: 都有内容，抽样对比质量
    winner = _sample_compare(old_rows, new_rows)

    if winner == "new":
        return 'replace'
    elif winner == "old" or winner == "keep":
        return 'keep'
    else:
        # 质量相当（merge），稳定优先，保留旧版
        print(f"  [Compare] {module_name}: 质量相当 → 稳定优先保留旧版 ({old_count}条)")
        return 'keep'
# ===== End Layer 2 =====


# ===== Layer 3: 模块压缩函数 =====
def _compress_module(module_name: str, features: list, limit: int, gateway=None) -> list:
    """对超限模块调用 LLM 合并精简

    Args:
        module_name: 模块名
        features: 当前功能列表
        limit: 行数上限
        gateway: LLM 网关（可选）
    Returns:
        精简后的功能列表（≤ limit 条）
    """
    if len(features) <= limit:
        return features

    print(f"  [Compress] {module_name}: {len(features)} 条 → 目标 ≤{limit} 条")

    # 分离 P0（不可删除）和非 P0
    p0_features = [f for f in features if f.get('priority') == 'P0']
    other_features = [f for f in features if f.get('priority') != 'P0']

    # 如果仅 P0 就超限，只精简非 P0 为 0
    if len(p0_features) >= limit:
        print(f"  [Compress] ⚠️ {module_name} 仅 P0 就有 {len(p0_features)} 条，超过上限 {limit}")
        return p0_features[:limit]

    remaining_budget = limit - len(p0_features)

    # 如果非 P0 不需要精简
    if len(other_features) <= remaining_budget:
        return features

    # 用 LLM 精简非 P0 部分
    compress_prompt = f"""你是 PRD 精简专家。以下模块「{module_name}」有 {len(other_features)} 条非P0功能，需要精简到 {remaining_budget} 条以内。

精简规则：
1. 同一件事的不同说法（如"蓝牙断连重连"和"蓝牙连接中断恢复"）→ 合并为一条，保留描述最完整的
2. L3 粒度过细的功能 → 合并到对应 L2，不需要每个细节都独立成行
3. 纯异常/边界场景如果跟正常功能重复 → 合并到正常功能的验收标准中
4. 保留所有独立功能点，只删除冗余表述
5. 输出精简后的 JSON 数组，格式与输入相同

当前功能列表（{len(other_features)} 条）：
{json.dumps(other_features, ensure_ascii=False, indent=1)[:8000]}

只输出 JSON 数组，不要其他内容。精简到 {remaining_budget} 条以内。"""

    try:
        if gateway:
            result = gateway.call(compress_prompt, model_tier="fast")
        else:
            result = _call_llm_fallback(compress_prompt)

        if result and result.get("success"):
            resp = result.get("response", "").strip()
            # 清理 markdown 代码块
            resp = re.sub(r'^```json\s*', '', resp)
            resp = re.sub(r'^```\s*', '', resp)
            resp = re.sub(r'\s*```$', '', resp)
            compressed = json.loads(resp)

            if isinstance(compressed, list) and len(compressed) <= remaining_budget:
                final = p0_features + compressed
                print(f"  [Compress] ✅ {module_name}: {len(features)} → {len(final)} 条")
                return final
            elif isinstance(compressed, list):
                print(f"  [Compress] ⚠️ LLM 返回 {len(compressed)} 条，超过预算 {remaining_budget}")
                # 截断兜底
                return p0_features + compressed[:remaining_budget]
    except Exception as e:
        print(f"  [Compress] ❌ LLM 精简失败: {e}")

    # LLM 失败兜底：按优先级截断
    print(f"  [Compress] 降级：按优先级截断")
    priority_order = {"P1": 0, "P2": 1, "P3": 2}
    other_features.sort(key=lambda f: priority_order.get(f.get('priority', 'P3'), 3))
    return p0_features + other_features[:remaining_budget]


def _apply_density_limits(items: list, anchor: dict, gateway=None) -> list:
    """对所有模块应用密度上限检查和精简

    Args:
        items: 功能列表
        anchor: 锚点配置
        gateway: LLM 网关
    Returns:
        精简后的功能列表
    """
    # 按模块分组
    by_module = {}
    for item in items:
        mod = item.get('module', item.get('L1功能', 'unknown'))
        if mod not in by_module:
            by_module[mod] = []
        by_module[mod].append(item)

    total_before = len(items)
    total_after = 0
    compressed_modules = []

    for module_name, features in by_module.items():
        limit = MODULE_DENSITY_LIMITS.get(module_name, DEFAULT_DENSITY_LIMIT)
        if len(features) > limit:
            compressed = _compress_module(module_name, features, limit, gateway)
            by_module[module_name] = compressed
            compressed_modules.append(f"{module_name}({len(features)}→{len(compressed)})")
            total_after += len(compressed)
        else:
            total_after += len(features)

    # 重新组合
    result = []
    for features in by_module.values():
        result.extend(features)

    if compressed_modules:
        print(f"[DensityCap] {total_before} → {total_after} 条 (精简 {total_before - total_after})")
        print(f"[DensityCap] 精简模块: {', '.join(compressed_modules[:5])}")
        if len(compressed_modules) > 5:
            print(f"  ...还有 {len(compressed_modules) - 5} 个模块")

    return result
# ===== End Layer 3 =====


# ===== Layer 4: 模块级微循环 QA =====
def _module_qa_check(module_name: str, features: list) -> dict:
    """模块级质量检查——纯规则 + 轻量检查

    Returns:
        {"pass": bool, "issues": [...], "auto_fixable": [...]}
    """
    issues = []
    auto_fixable = []

    for f in features:
        fid = f.get('功能ID', f.get('id', f.get('_id', '')))
        name = f.get('name', f.get('L2功能', f.get('L3功能', '')))
        level = f.get('level', '')

        # Rule 1: 空描述
        desc = f.get('description', f.get('描述', ''))
        if not desc or len(str(desc).strip()) < 10:
            issues.append(f"[空描述] {fid} {name}")

        # Rule 2: 空验收标准
        acc = f.get('acceptance', f.get('验收标准', ''))
        if not acc or len(str(acc).strip()) < 10:
            issues.append(f"[空验收] {fid} {name}")

        # Rule 3: L1/L2/L3 层级错误（L1 有值但不等于 module_name）
        l1 = f.get('module', f.get('L1功能', ''))
        if level == 'L1' and l1 and l1 != module_name:
            issues.append(f"[L1错位] {fid} L1={l1} 应为 {module_name}")
            auto_fixable.append(('fix_l1', fid, module_name))

        # Rule 4: 优先级缺失
        pri = f.get('priority', f.get('优先级', ''))
        if not pri or pri not in ('P0', 'P1', 'P2', 'P3'):
            issues.append(f"[无优先级] {fid} {name}")
            auto_fixable.append(('fix_priority', fid, 'P2'))

    # Rule 5: 模块内 L2 功能名重复
    l2_names = []
    for f in features:
        if f.get('level') in ('L2', ''):
            l2_name = f.get('L2功能', f.get('name', ''))
            if l2_name:
                l2_names.append(l2_name)

    seen = set()
    for n in l2_names:
        if n in seen:
            issues.append(f"[L2重复] {module_name} 有重复 L2: {n}")
        seen.add(n)

    passed = len(issues) == 0
    if not passed:
        print(f"  [QA] {module_name}: {len(issues)} 个问题")
        for iss in issues[:5]:
            print(f"    {iss}")
        if len(issues) > 5:
            print(f"    ...还有 {len(issues) - 5} 个")
    else:
        print(f"  [QA] ✅ {module_name}: 通过")

    return {"pass": passed, "issues": issues, "auto_fixable": auto_fixable}


def _module_qa_autofix(features: list, auto_fixable: list) -> list:
    """自动修复可修复的问题"""
    fix_map = {}
    for fix_type, fid, value in auto_fixable:
        fix_map[(fix_type, fid)] = value

    for f in features:
        fid = f.get('功能ID', f.get('id', f.get('_id', '')))
        if ('fix_l1', fid) in fix_map:
            f['L1功能'] = fix_map[('fix_l1', fid)]
            f['module'] = fix_map[('fix_l1', fid)]
        if ('fix_priority', fid) in fix_map:
            f['优先级'] = fix_map[('fix_priority', fid)]
            f['priority'] = fix_map[('fix_priority', fid)]

    return features


def _apply_module_qa(items: list) -> list:
    """对所有模块执行 QA 检查并自动修复

    Args:
        items: 功能列表
    Returns:
        检查并修复后的功能列表
    """
    # 按模块分组
    by_module = {}
    for item in items:
        mod = item.get('module', item.get('L1功能', 'unknown'))
        if mod not in by_module:
            by_module[mod] = []
        by_module[mod].append(item)

    total_issues = 0
    total_fixed = 0

    for module_name, features in by_module.items():
        qa_result = _module_qa_check(module_name, features)
        if not qa_result["pass"]:
            total_issues += len(qa_result["issues"])
            if qa_result["auto_fixable"]:
                by_module[module_name] = _module_qa_autofix(features, qa_result["auto_fixable"])
                total_fixed += len(qa_result["auto_fixable"])

    if total_issues > 0:
        print(f"[QA] 共发现 {total_issues} 个问题，自动修复 {total_fixed} 个")

    # 重新组合
    result = []
    for features in by_module.values():
        result.extend(features)

    return result
# ===== End Layer 4 =====


def _merge_two_versions(old_rows: list, new_rows: list) -> list:
    """合并两个版本，按 L3 功能名逐条取优"""
    merged = []
    seen_names = {}

    # 先处理旧版
    for row in old_rows:
        name = row.get('name', '')
        if name and name not in seen_names:
            seen_names[name] = row

    # 再处理新版，取优
    for row in new_rows:
        name = row.get('name', '')
        if name in seen_names:
            old_row = seen_names[name]
            # 比较验收标准长度，取更完整的
            if len(str(row.get('acceptance', ''))) > len(str(old_row.get('acceptance', ''))):
                seen_names[name] = row
        else:
            seen_names[name] = row

    return list(seen_names.values())
# ===== End Fix 4 =====


def _gen_one_with_gemini(feature: Dict, gateway) -> List[Dict]:
    """降级到 Gemini 生成单个模块"""
    name = feature["name"]
    module = feature["module"]

    try:
        from src.tools.knowledge_base import search_knowledge, format_knowledge_for_prompt
    except ImportError:
        search_knowledge = None
        format_knowledge_for_prompt = None

    # 使用数据点提取模式
    kb_raw = ""
    if search_knowledge:
        for q in [f"{name} 技术参数", f"{name} 竞品", f"{name} 标准"]:
            try:
                entries = search_knowledge(q, limit=3)
                if entries:
                    kb_raw += format_knowledge_for_prompt(entries)[:1500] + "\n"
            except Exception:
                pass

    # 从全文提取数据点（~200字），不灌全文
    kb_data_points = _extract_data_points(kb_raw)
    kb_inject = ""
    if kb_data_points:
        kb_inject = f"\n\n参考数据：\n{kb_data_points}\n"

    print(f"  [KB-Gemini] {name}: 知识库原文 {len(kb_raw)} 字 -> 数据点 {len(kb_data_points)} 字")

    extra_hint = ""
    if "我的" in name:
        extra_hint = "\n必须包含:帮助与反馈、关于设备、隐私与协议、数据管理。\n"

    batch_prompt = (
        f"为智能摩托车全盔项目生成'{name}'模块的功能清单.\n"
        f"模块归属:{module}\n"
        f"{extra_hint}{kb_inject}\n"
        f"输出 JSON 数组,每个元素格式:\n"
        f'{{"module":"{module}","level":"L1或L2或L3","parent":"父功能名",'
        f'"name":"功能名称","priority":"P0-P3",'
        f'"interaction":"交互方式","description":"描述",'
        f'"acceptance":"验收标准(含数字)","dependencies":"关联功能","note":"备注"}}\n\n'
        f"规则:第一条L1,下至少3个L2每个至少2个L3.验收标准基于知识库,无数据标[待验证].\n"
        f"P0<=30% P1约40% P2约25%.只输出JSON."
    )

    try:
        result = gateway.call_gemini("gemini_2_5_flash", batch_prompt,
            "只输出JSON数组.", "structured_doc_fallback")

        if not result.get("success"):
            error_msg = result.get("error", "未知错误")
            print(f"  [GenOne-Gemini] {name} Gemini 返回失败: {error_msg}")
            return []

        response = result.get("response", "")
        if len(response) < 50:
            print(f"  [GenOne-Gemini] {name} 响应太短: {len(response)} 字")
            return []

        json_match = re.search(r'\[[\s\S]*\]', response)
        if not json_match:
            print(f"  [GenOne-Gemini] {name} 无法提取 JSON，响应前200字: {response[:200]}")
            return []

        try:
            # Bug C: 清洗控制字符
            cleaned_json = _clean_json_text(json_match.group())
            items = json.loads(cleaned_json)
            # 防御：确保所有字段都是 string
            for item in items:
                for key in item:
                    if isinstance(item[key], (list, dict)):
                        item[key] = ", ".join(str(v) for v in item[key]) if isinstance(item[key], list) else str(item[key])
                    elif item[key] is None:
                        item[key] = ""
            return items
        except json.JSONDecodeError as je:
            print(f"  [GenOne-Gemini] {name} JSON 解析失败: {je}")
            print(f"  [GenOne-Gemini] {name} JSON 片段: {cleaned_json[:300]}")
            return []

    except Exception as e:
        print(f"  [GenOne-Gemini] {name} 异常: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return []


def _gen_one_minimal(feature: Dict, gateway) -> List[Dict]:
    """精简模式：最短 prompt，但仍包含知识库数据点和 anchor 核心功能点"""
    name = feature["name"]
    module = feature["module"]

    # 仍然搜知识库，但只取数据点
    kb_raw = ""
    try:
        from src.tools.knowledge_base import search_knowledge, format_knowledge_for_prompt
        entries = search_knowledge(name, limit=3)
        if entries:
            kb_raw = format_knowledge_for_prompt(entries)[:1500]
    except Exception:
        pass

    data_points = _extract_data_points(kb_raw)
    kb_line = f"\n参考数据：{data_points}\n" if data_points else ""

    # Bug 4 Layer 2: 精简模式也注入 anchor 的必须功能点
    anchor_hint = ""
    anchor_obj = get_cached_anchor()
    if anchor_obj:
        hint = _get_anchor_sub_features(anchor_obj, name)
        if hint:
            # 只取功能名（去掉括号里的说明），限 200 字
            names_only = re.sub(r'（[^）]+）|\([^)]+\)', '', hint)
            names_only = names_only[:200]
            anchor_hint = f"\n必须包含：{names_only}"

    prompt = (
        f"生成「{name}」功能清单。模块：{module}。{kb_line}{anchor_hint}"
        f"输出JSON数组：module/level(L1,L2,L3)/parent/name/priority/interaction/description/acceptance/dependencies/note。"
        f"至少3个L2每个2个L3。只输出JSON。"
    )

    try:
        result = gateway.call_azure_openai("cpo", prompt, "只输出JSON。", "structured_doc_minimal")

        if result.get("success"):
            json_match = re.search(r'\[[\s\S]*\]', result["response"])
            if json_match:
                try:
                    # Bug C: 清洗控制字符
                    cleaned_json = _clean_json_text(json_match.group())
                    items = json.loads(cleaned_json)
                    # 防御：确保所有字段都是 string
                    for item in items:
                        for key in item:
                            if isinstance(item[key], (list, dict)):
                                item[key] = ", ".join(str(v) for v in item[key]) if isinstance(item[key], list) else str(item[key])
                            elif item[key] is None:
                                item[key] = ""
                    return items
                except json.JSONDecodeError as je:
                    print(f"  [GenOne-Minimal] {name} JSON 解析失败: {je}")
        print(f"  [GenOne-Minimal] {name} 失败")
    except Exception as e:
        print(f"  [GenOne-Minimal] {name} 异常: {e}")
    return []


def _calibrate_priorities(all_items: List[Dict]) -> None:
    """全局优先级校准:确保 P0<=25%"""
    total = len(all_items)
    if total == 0:
        return

    p0_count = sum(1 for i in all_items if i.get("priority") == "P0")
    p0_ratio = p0_count / total

    if p0_ratio <= 0.25:
        print(f"[Calibrate] P0 占比 {p0_ratio:.0%},无需校准")
        return

    print(f"[Calibrate] P0 占比 {p0_ratio:.0%} > 25%,开始校准...")

    # Round 5 Fix 6: 豁免核心安全模块（不参与降级）
    P0_EXEMPT_MODULES = {
        '主动安全预警提示',  # 安全核心
        'SOS与紧急救援',     # 安全核心
        '佩戴检测与电源管理', # 硬件安全
        '导航',              # 产品核心卖点
        '组队',              # 产品核心卖点
        '设备状态',           # 基础功能
        'AI语音助手', 'Ai语音助手',  # AI 核心
        '摄像状态',          # 核心功能
    }

    p0_l3 = [i for i in all_items if i.get("priority") == "P0" and i.get("level") == "L3"]
    target_demote = p0_count - int(total * 0.25)
    demoted = 0

    for item in p0_l3:
        if demoted >= target_demote:
            break
        if item.get("module", "") not in P0_EXEMPT_MODULES:
            item["priority"] = "P1"
            demoted += 1

    new_p0 = sum(1 for i in all_items if i.get("priority") == "P0")
    print(f"[Calibrate] 降级 {demoted} 个 L3: P0 {p0_count}->{new_p0} ({new_p0/total:.0%})")


def _cross_module_audit(all_items: List[Dict], extra_sheets: Dict) -> List[Dict]:
    """跨模块一致性审计"""
    issues = []

    # 1. 只检查名称或描述明确包含"语音控制/语音指令/语音操作"的 L2
    voice_commands = extra_sheets.get("voice", [])
    voice_texts = " ".join(str(vc) for vc in voice_commands).lower()

    for item in all_items:
        name = item.get("name", "")
        desc = item.get("description", "")

        is_voice_control = any(kw in name + desc for kw in
            ["语音控制", "语音指令", "语音操作", "语音发起", "语音查询", "语音切换"])

        if is_voice_control and item.get("level") == "L2":
            has_match = any(
                name[:4].lower() in str(vc).lower()
                for vc in voice_commands
            )
            if not has_match:
                issues.append({
                    "type": "功能-语音不一致",
                    "issue": f"「{name}」含语音控制功能，但语音指令表无对应",
                    "module": item.get("module", "")
                })

    # 2. 状态场景 vs 灯效交叉检查
    light_triggers = {le.get("trigger", "").lower() for le in extra_sheets.get("light", [])}
    for sc in extra_sheets.get("state", []):
        light = sc.get("light", "")
        scene = sc.get("current", "")
        if light and light != "无" and scene[:6].lower() not in " ".join(light_triggers):
            issues.append({"type": "场景-灯效不一致",
                "issue": f"状态场景 '{scene[:30]}' 有灯光提示但灯效表无对应",
                "module": "灯效"})

    # 3. 模块级 P0 占比过高
    p0_by_mod = {}
    total_by_mod = {}
    for item in all_items:
        mod = item.get("module", "")
        total_by_mod[mod] = total_by_mod.get(mod, 0) + 1
        if item.get("priority") == "P0":
            p0_by_mod[mod] = p0_by_mod.get(mod, 0) + 1
    for mod, p0 in p0_by_mod.items():
        total = total_by_mod.get(mod, 1)
        if p0 / total > 0.5:
            issues.append({"type": "优先级失衡",
                "issue": f"'{mod}' P0 占比 {p0/total:.0%} ({p0}/{total}),建议降级",
                "module": mod})

    # 4. [待生成][待验证] 统计
    pending_gen = sum(1 for i in all_items if "[待生成]" in i.get("description", ""))
    pending_verify = sum(1 for i in all_items if "[待验证]" in i.get("acceptance", ""))
    if pending_gen > 0:
        issues.append({"type": "生成不完整",
            "issue": f"{pending_gen} 个模块标记[待生成],需重新触发", "module": "全局"})
    if pending_verify > 5:
        issues.append({"type": "验收待验证",
            "issue": f"{pending_verify} 个功能标记[待验证],建议补充知识库后重新生成", "module": "全局"})

    if issues:
        print(f"\n[Audit] 发现 {len(issues)} 个一致性问题:")
        for iss in issues[:15]:
            print(f"  ! [{iss['type']}] {iss['issue'][:60]}")
    else:
        print("[Audit] 一致性检查通过 OK")

    # 保存供下次迭代使用
    try:
        audit_path = EXPORT_DIR.parent / "prd_audit_issues.json"
        audit_path.parent.mkdir(parents=True, exist_ok=True)
        audit_path.write_text(json.dumps(issues, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

    return issues


# ===== Fix 6: 一致性审计自我迭代闭环 =====
def _auto_fix_audit_issues(audit_issues: list, sheets_data: dict, gateway) -> list:
    """针对可自动修复的审计问题，自动补生成缺失条目"""
    if not audit_issues:
        return []

    fixable = []
    manual = []
    for issue in audit_issues:
        issue_type = issue.get('type', '')
        if '功能-语音不一致' in issue_type or '场景-灯效不一致' in issue_type:
            fixable.append(issue)
        else:
            manual.append(issue)

    if not fixable:
        print(f"  [Audit-Fix] 无可自动修复的问题")
        return manual

    print(f"  [Audit-Fix] 发现 {len(fixable)} 个可自动修复问题")

    voice_gaps = [i for i in fixable if '语音' in i.get('type', '')]
    light_gaps = [i for i in fixable if '灯效' in i.get('type', '')]

    if voice_gaps and 'voice' in sheets_data:
        missing_features = [i.get('feature_name', '') for i in voice_gaps if i.get('feature_name')]
        if missing_features:
            prompt = f"为以下功能生成语音指令表条目(JSON数组): {', '.join(missing_features[:5])}。字段:category,wake,user_says,action,success_voice。"
            try:
                result = gateway.call_azure_openai("cpo", prompt, "只输出JSON数组。", "audit_fix", max_tokens=1500)
                if result.get("success"):
                    json_match = re.search(r'\[[\s\S]*\]', result.get("response", ""))
                    if json_match:
                        new_rows = json.loads(_clean_json_text(json_match.group()))
                        if new_rows:
                            sheets_data['voice'].extend(new_rows)
                            print(f"  [Audit-Fix] 语音表补充 {len(new_rows)} 条")
            except Exception as e:
                print(f"  [Audit-Fix] 语音补充失败: {e}")

    if light_gaps and 'light' in sheets_data:
        missing_scenes = [i.get('scene_name', '') for i in light_gaps if i.get('scene_name')]
        if missing_scenes:
            prompt = f"为以下场景生成灯效定义条目(JSON数组): {', '.join(missing_scenes[:5])}。字段:trigger,color,mode,frequency,duration,priority。"
            try:
                result = gateway.call_azure_openai("cpo", prompt, "只输出JSON数组。", "audit_fix", max_tokens=1500)
                if result.get("success"):
                    json_match = re.search(r'\[[\s\S]*\]', result.get("response", ""))
                    if json_match:
                        new_rows = json.loads(_clean_json_text(json_match.group()))
                        if new_rows:
                            sheets_data['light'].extend(new_rows)
                            print(f"  [Audit-Fix] 灯效表补充 {len(new_rows)} 条")
            except Exception as e:
                print(f"  [Audit-Fix] 灯效补充失败: {e}")

    print(f"  [Audit-Fix] 完成。剩余 {len(manual)} 个需人工确认")
    return manual
# ===== End Fix 6 =====


# ===== Fix 2: 移除[待生成]占位行 =====
def _remove_placeholder_rows(rows: list, prev_version_map: dict = None) -> list:
    """移除生成失败的占位行，用上一版数据或最小骨架替代"""
    if prev_version_map is None:
        prev_version_map = {}

    cleaned = []
    for row in rows:
        desc = str(row.get('description', ''))
        if '待生成' in desc or '生成失败' in desc:
            # 尝试从上一版获取该模块的数据
            module_name = row.get('module') or row.get('name', '')
            prev_rows = prev_version_map.get(module_name, [])
            if prev_rows:
                cleaned.extend(prev_rows)
                print(f"  [Fallback] {module_name}: 用上一版 {len(prev_rows)} 条替代")
            else:
                # 上一版也没有 → 生成最小骨架
                skeleton = {
                    **row,
                    'description': f'{module_name}模块（待详细展开）',
                    'acceptance': '待下一版本迭代补充',
                    'level': 'L1',
                }
                cleaned.append(skeleton)
                print(f"  [Skeleton] {module_name}: 生成最小骨架")
        else:
            cleaned.append(row)
    return cleaned
# ===== End Fix 2 =====


# === Excel 导出 ===
def _export_to_excel(items: List[Dict], filename_prefix: str, title_hint: str, extra_sheets: Dict = None, audit_issues: List[Dict] = None, version_info: Dict = None, flow_diagrams: list = None, anchor: dict = None) -> str:
    """导出功能清单到 Excel 文件(支持 6 个 Sheet)

    Args:
        version_info: 可选的版本信息字典，包含 prev_version, new_modules, updated_modules 等
        anchor: B5: 用于生成用户旅程和 AI 场景表
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[FastTrack] openpyxl 未安装,降级使用 CSV")
        csv_path = EXPORT_DIR / f"{filename_prefix}.csv"
        with open(csv_path, "w", encoding="utf-8-sig") as f:
            headers = ["功能ID", "L1功能", "L2功能", "L3功能", "优先级",
                       "交互方式", "描述", "验收标准", "关联功能", "备注"]
            f.write(",".join(headers) + "\n")
            for item in items:
                level = item.get("level", "")
                name = item.get("name", "")
                row = [
                    item.get("_id", ""),
                    name if level == "L1" else "",
                    name if level == "L2" else "",
                    name if level == "L3" else "",
                    item.get("priority", ""),
                    item.get("interaction", ""),
                    item.get("description", ""),
                    item.get("acceptance", ""),
                    item.get("dependencies", ""),
                    item.get("note", "")
                ]
                f.write(",".join(row) + "\n")
        return str(csv_path)

    # 模块名统一
    items = _normalize_module_names(items)

    # Fix 2: 移除[待生成]占位行
    items = _remove_placeholder_rows(items)

    # 生成功能ID
    items = _generate_ids(items)

    # 按 anchor 动态分流（替代静态列表）
    anchor = get_cached_anchor()

    # A2: 全量归一化（含旧版继承数据）
    if anchor:
        normalize_map = anchor.get('module_normalize', {})
    else:
        normalize_map = {}
    normalize_map.update(MODULE_NAME_NORMALIZE)  # 合并代码中的硬编码映射
    items = _normalize_all_rows(items, normalize_map, anchor)

    # A1: 构建模块归属映射
    if anchor:
        module_placement = _build_module_placement(anchor)
    else:
        module_placement = {}

    hud_items = []
    app_items = []

    for i in items:
        module_name = i.get("module", "")
        # 查归属，未找到的默认按名字推断
        target = module_placement.get(module_name, '')
        if not target:
            # 兜底推断
            if any(kw in module_name for kw in ['Tab', 'App-', '商城', '社区', '通知', '用户成就', '新手', '身份', '权限', '部件识别', '消息同步', '多语言', '恢复出厂']):
                target = 'app'
            else:
                target = 'hud'

        if target == 'hud':
            hud_items.append(i)
        else:
            app_items.append(i)

    print(f"[Placement] 分流结果: HUD {len(hud_items)} 条, App {len(app_items)} 条")

    # === Fix B: 幽灵模块过滤 ===
    # 过滤掉不在 Anchor 定义中的 L1 模块
    if anchor:
        valid_hud_modules = set()
        valid_app_modules = set()

        for mod in anchor.get('hud_modules', []):
            name = mod.get('name', '')
            if name:
                valid_hud_modules.add(name)

        for mod in anchor.get('app_modules', []):
            name = mod.get('name', '')
            if name:
                valid_app_modules.add(name)

        # 跨端模块也加入有效集合
        for mod in anchor.get('cross_cutting', []):
            name = mod.get('name', '')
            if name:
                # 根据之前的 cross_to_hud/cross_to_app 逻辑分配
                cross_to_hud = {'AI功能', '语音交互', '视觉交互', '多模态交互', '实体按键交互', '氛围灯交互'}
                if name in cross_to_hud:
                    valid_hud_modules.add(name)
                else:
                    valid_app_modules.add(name)

        original_hud_count = len(hud_items)
        original_app_count = len(app_items)

        # 过滤幽灵模块
        hud_items = [f for f in hud_items if f.get('module') in valid_hud_modules]
        app_items = [f for f in app_items if f.get('module') in valid_app_modules]

        filtered_hud = original_hud_count - len(hud_items)
        filtered_app = original_app_count - len(app_items)

        if filtered_hud > 0 or filtered_app > 0:
            print(f"[GhostFilter] 过滤 HUD {filtered_hud} 条, App {filtered_app} 条不在 Anchor 中的功能")
    # === End Fix B ===

    # 修复 1: 强制归一化，确保写入时名称正确
    hud_items = _force_normalize(hud_items, normalize_map)
    app_items = _force_normalize(app_items, normalize_map)

    # 验证归一化结果
    hud_modules = set(r.get('module', r.get('L1功能', '')) for r in hud_items)
    app_modules = set(r.get('module', r.get('L1功能', '')) for r in app_items)
    print(f"[ForceNormalize] HUD 模块: {sorted(hud_modules)}")
    print(f"[ForceNormalize] App 模块: {sorted(app_modules)}")

    # A3: HUD 新 4 列回填
    hud_items = _backfill_hud_columns(hud_items)

    wb = Workbook()

    # Sheet 1: HUD 及头盔端功能
    ws_hud = wb.active
    ws_hud.title = "HUD及头盔端"
    if hud_items:
        _write_hud_sheet(ws_hud, hud_items)

    # Sheet 2: App 功能
    if app_items:
        ws_app = wb.create_sheet("App端")
        _write_sheet(ws_app, app_items)

    # Sheet 3-6: 额外 Sheet
    if extra_sheets:
        if extra_sheets.get("state"):
            ws3 = wb.create_sheet("状态场景对策")
            _write_generic_sheet(ws3, extra_sheets["state"],
                ["前置状态", "当前状态/场景/操作", "执行状态", "HUD提示", "灯光提示", "语音提示", "App提示", "提示周期"],
                ["pre_state", "current", "exec_state", "hud", "light", "voice", "app", "cycle"],
                [16, 24, 16, 28, 20, 28, 28, 14])

        if extra_sheets.get("voice"):
            ws4 = wb.create_sheet("语音指令表")
            _write_generic_sheet(ws4, extra_sheets["voice"],
                ["指令分类", "唤醒方式", "用户说法", "常见变体", "系统动作", "成功语音反馈", "成功HUD反馈", "失败反馈", "优先级", "备注"],
                ["category", "wake", "user_says", "variants", "action", "success_voice", "success_hud", "fail_feedback", "priority", "note"],
                [14, 12, 20, 24, 24, 24, 20, 20, 8, 16])

        if extra_sheets.get("button"):
            ws5 = wb.create_sheet("按键映射表")
            _write_generic_sheet(ws5, extra_sheets["button"],
                ["按键位置", "场景", "单击", "双击", "长按", "组合键", "操作反馈", "备注"],
                ["button", "scene", "single_click", "double_click", "long_press", "combo", "feedback", "note"],
                [14, 16, 20, 20, 20, 20, 18, 16])

        if extra_sheets.get("light"):
            ws6 = wb.create_sheet("灯效定义表")
            # B2: 增加 "作用灯区" 列
            _write_generic_sheet(ws6, extra_sheets["light"],
                ["触发场景", "灯光颜色", "闪烁模式", "频率Hz", "持续时长", "作用灯区", "优先级", "备注"],
                ["trigger", "color", "mode", "frequency", "duration", "lamp_zone", "priority", "note"],
                [24, 12, 16, 10, 14, 14, 8, 20])

        # Sheet 7: AI语音导航场景
        if extra_sheets.get("voice_nav"):
            ws7 = wb.create_sheet("AI语音导航场景")
            _write_generic_sheet(ws7, extra_sheets["voice_nav"],
                ["场景名称", "触发条件", "用户输入", "AI执行动作", "HUD显示", "语音播报", "灯光反馈", "异常兜底", "优先级", "备注"],
                ["scene", "trigger", "user_input", "ai_action", "hud_display", "voice_output", "light_effect", "fallback", "priority", "note"],
                [18, 20, 18, 24, 24, 24, 16, 24, 8, 16])

        # Sheet 8: 用户故事
        if extra_sheets.get("user_stories"):
            ws_story = wb.create_sheet("用户故事")
            _write_generic_sheet(ws_story, extra_sheets["user_stories"],
                ["功能名", "用户角色", "用户故事", "验收条件", "优先级"],
                ["feature", "role", "story", "acceptance", "priority"],
                [24, 14, 50, 36, 8])

        # Sheet 9: 测试用例
        if extra_sheets.get("test_cases"):
            ws_tc = wb.create_sheet("测试用例")
            _write_generic_sheet(ws_tc, extra_sheets["test_cases"],
                ["用例ID", "功能名", "用例标题", "前置条件", "操作步骤", "预期结果", "优先级"],
                ["case_id", "feature", "title", "precondition", "steps", "expected", "priority"],
                [12, 20, 24, 24, 36, 30, 8])

        # Sheet 10: 页面映射表
        if extra_sheets.get("page_mapping"):
            ws_page = wb.create_sheet("页面映射表")
            _write_generic_sheet(ws_page, extra_sheets["page_mapping"],
                ["页面名", "父页面", "平台", "包含功能", "入口方式", "优先级", "备注"],
                ["page", "parent", "platform", "features", "entry", "priority", "note"],
                [20, 16, 10, 40, 16, 8, 20])

        # Sheet 11: 开发任务
        if extra_sheets.get("dev_tasks"):
            ws_dev = wb.create_sheet("开发任务")
            _write_generic_sheet(ws_dev, extra_sheets["dev_tasks"],
                ["任务ID", "功能名", "任务描述", "负责角色", "预估工时(天)", "前置依赖", "建议迭代", "优先级", "备注"],
                ["task_id", "feature", "task", "assignee", "effort_days", "dependency", "sprint", "priority", "note"],
                [10, 20, 30, 14, 12, 20, 16, 8, 16])

    # 版本信息 Sheet
    ws_ver = wb.create_sheet("版本信息")

    # ===== Fix 5: 版本信息自动升版 =====
    prev_version = version_info.get("prev_version", "") if version_info else ""
    # 自动递增版本号
    if prev_version:
        # 从 "V1.0" 提取数字，+0.1
        match = re.search(r'V?(\d+)\.(\d+)', prev_version)
        if match:
            major, minor = int(match.group(1)), int(match.group(2))
            new_version = f"V{major}.{minor + 1}"
        else:
            new_version = "V1.1"
    else:
        new_version = "V1.0"

    # 生成 changelog
    changelog_items = []
    if version_info:
        if version_info.get('new_modules'):
            changelog_items.append(f"新增模块: {', '.join(version_info['new_modules'][:5])}")
        if version_info.get('updated_modules'):
            changelog_items.append(f"更新模块: {', '.join(version_info['updated_modules'][:5])}")
    changelog_items.append(f"功能总数: {len(items)} 条")
    if audit_issues:
        changelog_items.append(f"一致性问题: {len(audit_issues)} 个")
    changelog_text = '; '.join(changelog_items) if changelog_items else '首版生成'
    # ===== End Fix 5 =====

    ver_data = [
        ["项目", "智能骑行头盔 V1"],
        ["版本", new_version],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["功能总数", len(items)],
        ["HUD端功能", len(hud_items)],
        ["App端功能", len(app_items)],
        ["P0功能数", sum(1 for i in items if i.get("priority") == "P0")],
        ["P1功能数", sum(1 for i in items if i.get("priority") == "P1")],
        ["P2功能数", sum(1 for i in items if i.get("priority") == "P2")],
        ["P3功能数", sum(1 for i in items if i.get("priority") == "P3")],
        ["状态场景数", len(extra_sheets.get("state", [])) if extra_sheets else 0],
        ["语音指令数", len(extra_sheets.get("voice", [])) if extra_sheets else 0],
        ["按键映射数", len(extra_sheets.get("button", [])) if extra_sheets else 0],
        ["灯效定义数", len(extra_sheets.get("light", [])) if extra_sheets else 0],
        ["导航场景数", len(extra_sheets.get("voice_nav", [])) if extra_sheets else 0],
        [],
        ["本次主要改动", changelog_text],
    ]
    for row_idx, row_data in enumerate(ver_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_ver.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.font = Font(bold=True)
    ws_ver.column_dimensions['A'].width = 16
    ws_ver.column_dimensions['B'].width = 30

    # Sheet 13: 一致性审计(如果有)
    if audit_issues:
        ws_audit = wb.create_sheet("一致性审计")
        _write_generic_sheet(ws_audit, audit_issues,
            ["问题类型", "问题描述", "关联模块"],
            ["type", "issue", "module"],
            [16, 60, 16])

    # Sheet 14: 关键流程(如果有)
    if flow_diagrams:
        _write_flow_sheet(wb, flow_diagrams)

    # B5: Sheet 15/16: 用户旅程和 AI 场景(如果有 anchor)
    if anchor:
        _write_journey_sheet(wb, [], anchor)  # 修复 4: 直接从 anchor 读取结构化数据
        _write_ai_scenarios_sheet(wb, anchor)

    # 保存
    xlsx_path = EXPORT_DIR / f"{filename_prefix}.xlsx"
    wb.save(xlsx_path)
    print(f"[Export] Excel {wb.sheetnames}: {xlsx_path}")
    return str(xlsx_path)


# === 飞书文件发送 ===
def _send_file_to_feishu(reply_target: str, file_path: str, reply_type: str) -> bool:
    """用 requests 直接调飞书 HTTP API 发送文件"""
    import requests

    file_path = Path(file_path)
    if not file_path.exists():
        print(f"[File] 文件不存在: {file_path}")
        return False

    try:
        # 获取配置
        env_path = Path(__file__).resolve().parent.parent.parent / ".env"
        if not env_path.exists():
            env_path = Path(__file__).resolve().parent.parent / ".env"

        from dotenv import dotenv_values
        env = dotenv_values(str(env_path))
        app_id = env.get("FEISHU_APP_ID", "") or os.getenv("FEISHU_APP_ID", "")
        app_secret = env.get("FEISHU_APP_SECRET", "") or os.getenv("FEISHU_APP_SECRET", "")

        print(f"[File] env路径: {env_path} (存在: {env_path.exists()})")
        print(f"[File] app_id: {app_id[:10] if app_id else '空'}...")
        print(f"[File] secret: {'有' if app_secret else '空'}")
        print(f"[File] 文件: {file_path.name} ({file_path.stat().st_size} bytes)")

        if not app_id or not app_secret:
            print("[File] 缺少 APP_ID 或 APP_SECRET")
            return False

        # 获取 tenant_access_token
        token_resp = requests.post(
            "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
            json={"app_id": app_id, "app_secret": app_secret}
        )
        token_data = token_resp.json()

        if token_data.get("code") != 0:
            print(f"[File] token 获取失败: {token_data}")
            return False

        token = token_data["tenant_access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        # 上传文件
        print(f"[File] 上传中...")
        with open(file_path, 'rb') as f:
            upload_resp = requests.post(
                "https://open.feishu.cn/open-apis/im/v1/files",
                headers=headers,
                data={"file_type": "stream", "file_name": file_path.name},
                files={"file": (file_path.name, f)}
            )

        upload_data = upload_resp.json()
        print(f"[File] 上传响应: code={upload_data.get('code')} msg={upload_data.get('msg', '')}")

        if upload_data.get("code") != 0:
            return False

        file_key = upload_data["data"]["file_key"]
        print(f"[File] file_key: {file_key[:30]}...")

        # 发送文件消息
        # 转换 reply_type:open_id/chat_id
        id_type = "chat_id" if reply_type == "chat_id" else "open_id"

        send_resp = requests.post(
            f"https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type={id_type}",
            headers={**headers, "Content-Type": "application/json"},
            json={
                "receive_id": reply_target,
                "msg_type": "file",
                "content": json.dumps({"file_key": file_key})
            }
        )

        send_data = send_resp.json()
        print(f"[File] 发送响应: code={send_data.get('code')} msg={send_data.get('msg', '')}")

        if send_data.get("code") == 0:
            print(f"[File] ✅ 文件发送成功")
            return True
        else:
            print(f"[File] ❌ 发送失败")
            return False

    except Exception as e:
        print(f"[File] 异常: {e}")
        import traceback
        traceback.print_exc()
        return False


# === 树形摘要格式化 ===
def _format_items_as_tree(items: List[Dict]) -> str:
    """将功能列表格式化为树形摘要"""
    # 按 L1 -> L2 -> L3 分组
    tree: Dict[str, Dict[str, List[Dict]]] = {}

    for item in items:
        level = item.get("level", "L1")
        module = item.get("module", "未分类")
        parent = item.get("parent", "")
        name = item.get("name", "")
        priority = item.get("priority", "P2")

        if level == "L1":
            if module not in tree:
                tree[module] = {"_self": item, "_children": {}}
        elif level == "L2":
            if parent and parent in tree:
                if "_children" not in tree[parent]:
                    tree[parent]["_children"] = {}
                tree[parent]["_children"][name] = {"_self": item, "_children": []}
        elif level == "L3":
            # 找到对应的 L2 parent
            for l1_mod, l1_data in tree.items():
                if parent in l1_data.get("_children", {}):
                    l1_data["_children"][parent]["_children"].append(item)
                    break

    # 格式化输出
    lines = []
    for l1_mod, l1_data in tree.items():
        l1_item = l1_data.get("_self", {})
        l1_name = l1_item.get("name", l1_mod)
        l1_priority = l1_item.get("priority", "P1")
        lines.append(f"📁 {l1_name} [{l1_priority}]")

        for l2_name, l2_data in l1_data.get("_children", {}).items():
            l2_item = l2_data.get("_self", {})
            l2_priority = l2_item.get("priority", "P2")
            lines.append(f"  📂 {l2_name} [{l2_priority}]")

            for l3_item in l2_data.get("_children", []):
                l3_name = l3_item.get("name", "")
                l3_priority = l3_item.get("priority", "P3")
                l3_desc = l3_item.get("description", "")[:30]
                lines.append(f"    📄 {l3_name} [{l3_priority}] - {l3_desc}")

    return "\n".join(lines)


# === 主函数 ===
def try_structured_doc_fast_track(
    text: str,
    reply_target: str,
    reply_type: str,
    open_id: str,
    chat_id: str,
    send_reply_func: Optional[callable] = None
) -> bool:
    """
    结构化文档快速通道

    Args:
        text: 用户消息文本
        reply_target: 回复目标 ID
        reply_type: 回复类型 (open_id / chat_id)
        open_id: 用户 Open ID
        chat_id: 群聊 Chat ID
        send_reply_func: 发送回复的函数(可选,默认使用飞书 API)

    Returns:
        bool: True 表示走了快速通道,False 表示继续走原有逻辑
    """
    if not is_structured_doc_request(text):
        return False

    print(f"[FastTrack] 检测到结构化文档需求,走快速通道")

    # 发送回复的函数
    def _send_reply(msg: str, rt: str = None, rtype: str = None):
        rt = rt or reply_target
        rtype = rtype or reply_type
        if send_reply_func:
            send_reply_func(rt, msg, rtype)
        else:
            # 默认使用飞书 API
            try:
                import lark_oapi as lark
                import os
                APP_ID = os.getenv("FEISHU_APP_ID", "cli_a9326fa6ba389cc5")
                APP_SECRET = os.getenv("FEISHU_APP_SECRET", "")

                client = lark.Client.builder() \
                    .app_id(APP_ID) \
                    .app_secret(APP_SECRET) \
                    .log_level(lark.LogLevel.ERROR) \
                    .build()

                from lark_oapi.api.im.v1 import CreateMessage, CreateMessageRequestBody
                content = json.dumps({"text": msg})

                request = CreateMessage.builder() \
                    .receive_id_type(rtype) \
                    .request_body(CreateMessageRequestBody.builder()
                        .receive_id(rt)
                        .msg_type("text")
                        .content(content)
                        .build()) \
                    .build()

                client.im.v1.message.create(request)
            except Exception as e:
                print(f"[FastTrack] 发送回复失败: {e}")

    _send_reply("📋 检测到结构化文档需求,正在生成...")

    def _process_in_background():
        try:
            from src.utils.model_gateway import get_model_gateway
            from src.tools.knowledge_base import search_knowledge, format_knowledge_for_prompt, KB_ROOT

            gw = get_model_gateway()

            # 检测是否需要完整规格书(6 Sheet)
            is_full_spec = is_full_spec_request(text)
            if is_full_spec:
                print("[FastTrack] 完整规格书模式(6 Sheet)")

            # 搜知识库找内部文档
            kb_entries = search_knowledge(text, limit=10)
            kb_context = format_knowledge_for_prompt(kb_entries) if kb_entries else ""

            # 读产品目标
            goal_text = ""
            goal_file = KB_ROOT.parent.parent / "product_goal.json"
            if goal_file.exists():
                try:
                    goal_text = json.loads(goal_file.read_text(encoding="utf-8")).get("goal", "")
                except:
                    pass

            # 分批 LLM 调用:按一级功能逐个生成
            l1_features = _extract_l1_from_user_text(text)
            total = len(l1_features)

            # ===== Fix 8: KB 驱动功能补全检查 =====
            KNOWN_FEATURES_FROM_KB = [
                '视频水印',  # 自定义水印（品牌/时间/GPS/速度叠加）
                '照片水印',
            ]
            module_names = {f['name'] for f in l1_features}
            for feature in KNOWN_FEATURES_FROM_KB:
                if not any(feature in m for m in module_names):
                    # 检查知识库中是否有相关条目
                    try:
                        from src.tools.knowledge_base import search_knowledge
                        kb_hits = search_knowledge(feature, limit=1)
                        if kb_hits:
                            # 追加到相关模块（如摄像状态）
                            if '摄像' in module_names or '摄像状态' in module_names:
                                print(f"  [AutoComplete] 知识库有 {feature}，将追加到摄像状态模块")
                            else:
                                print(f"  [AutoComplete] 知识库有 {feature}，建议在 prompt 中提及")
                    except Exception:
                        pass
            # ===== End Fix 8 =====

            # ===== A3: 加载 anchor 并合并模块列表 =====
            anchor = get_cached_anchor()
            if anchor:
                prompt_module_names = [f['name'] for f in l1_features]
                merged_names = _merge_anchor_with_prompt(anchor, prompt_module_names)

                # 用 anchor 的归一化映射更新全局映射表
                anchor_config = _get_anchor_config(anchor)
                MODULE_NAME_NORMALIZE.update(anchor_config.get('normalize_map', {}))

                # 更新 l1_features 为合并后的列表
                l1_features = [{'name': name, 'module': name} for name in merged_names]
                total = len(l1_features)
                print(f"[FastTrack] 合并 anchor 后共 {total} 个一级功能,开始并行生成 ({PARALLEL_WORKERS} 路)")
            else:
                print(f"[FastTrack] 检测到 {total} 个一级功能,开始并行生成 ({PARALLEL_WORKERS} 路)")
            # ===== End A3 =====

            # === 读取上一版版本快照 ===
            prev_items = []
            prev_by_module = {}  # {module_name: [items]}

            versions_dir = EXPORT_DIR.parent / "prd_versions"
            if versions_dir.exists():
                latest_files = sorted(versions_dir.glob("prd_v_*.json"), reverse=True)
                if latest_files:
                    try:
                        prev = json.loads(latest_files[0].read_text(encoding="utf-8"))
                        prev_items = prev.get("items", [])
                        prev_version = prev.get("version", "")

                        # 按 L1 模块分组上一版数据
                        current_l1 = ""
                        for item in prev_items:
                            if item.get("level") == "L1":
                                current_l1 = item.get("name", "")
                            if current_l1:
                                if current_l1 not in prev_by_module:
                                    prev_by_module[current_l1] = []
                                prev_by_module[current_l1].append(item)

                        print(f"[SmartIterate] 上一版: {prev_version} ({len(prev_items)} 条, {len(prev_by_module)} 模块)")

                        # 修复 8: 全量重生成机制提示
                        if len(prev_items) > 1200:
                            print(f"[SmartIterate] ⚠️ 上一版 {len(prev_items)} 条，建议做一次全量重生成以清理历史残留")
                        # 检测用户 prompt 是否请求全量重生成
                        if '全量' in text or 'clean' in text.lower() or '重生成' in text:
                            print(f"[SmartIterate] 检测到全量重生成指令，跳过 Compare，使用纯新版")
                            prev_items = []
                            prev_by_module = {}

                        # Bug 3 修复: 对旧版数据归一化
                        if prev_by_module and anchor:
                            normalize_map = anchor.get('module_normalize', {})
                            normalize_map.update(MODULE_NAME_NORMALIZE)
                            normalized_prev = {}
                            for mod_name, rows in prev_by_module.items():
                                normalized_name = normalize_map.get(mod_name, mod_name)
                                for row in rows:
                                    row['module'] = normalize_map.get(row.get('module', ''), row.get('module', ''))
                                if normalized_name not in normalized_prev:
                                    normalized_prev[normalized_name] = []
                                normalized_prev[normalized_name].extend(rows)
                            prev_by_module = normalized_prev
                            print(f"[Normalize-Pre] 旧版数据归一化: {len(prev_by_module)} 模块")
                    except Exception as e:
                        print(f"[SmartIterate] 读取上一版失败: {e}")

            all_items = []
            batch_system_prompt = (
                "你是智能摩托车全盔项目的产品经理.你必须且只输出一个 JSON 数组.\n"
                "不要输出任何 markdown、解释文字、标题、分隔线.\n"
                "只输出以 [ 开头、以 ] 结尾的 JSON 数组.\n\n"
                "每个元素格式:\n"
                '{"module":"模块名","level":"L1或L2或L3","parent":"父功能名(L1填空)","name":"功能名称",'
                '"priority":"P0或P1或P2或P3","interaction":"HUD/语音/按键/App/灯光",'
                '"description":"一句话描述","acceptance":"可测试验收标准(含数字)",'
                '"dependencies":"关联功能","note":"备注"}\n\n'
                "规则:\n"
                "1. 第一条是该模块的 L1\n"
                "2. L1 下至少 3 个 L2,每个 L2 至少 2 个 L3\n"
                "3. 验收标准必须可测试,含具体数字\n"
                "4. 你补充的功能在 note 标注[补充]\n"
                "5. 优先级只用 P0/P1/P2/P3\n"
                "6. 优先级分布:P0≤30%,P1占30-40%,P2占20-30%,P3占5-10%\n"
                "7. P0=不做就不能发售.P1=发售应有但可OTA补.P2=V2规划.P3=远期愿景\n"
                "8. 不要把所有功能都标P0,只有真正阻碍发售的才是P0\n"
            )

            # ===== Bug B: 复杂模块自动拆解 =====
            def _estimate_complexity(module_name: str, kb_data_points: str, prev_version_rows: list = None) -> list:
                """
                判断模块是否需要拆解。返回子模块名列表。
                如果不需要拆解，返回 [module_name] 本身（单元素列表）。
                """
                reasons = []

                # Bug 4 Layer 1: depth:deep 的模块强制拆解
                is_deep = False
                if anchor:
                    for section in ['hud_modules', 'app_modules', 'cross_cutting']:
                        for mod in anchor.get(section, []):
                            if mod.get('name') == module_name and mod.get('depth') == 'deep':
                                is_deep = True
                                reasons.append('depth_deep')
                                break
                        if is_deep:
                            break

                if is_deep:
                    # 深度模块直接拆解，不需要其他信号
                    default_splits = [
                        f"{module_name}-核心功能",
                        f"{module_name}-交互与状态",
                        f"{module_name}-异常与边界",
                    ]
                    print(f"  [AutoSplit] {module_name} 是深度模块，主动拆解")
                    return default_splits

                # 信号1: 知识库数据点超过 700 字 → 说明涉及面广
                if len(kb_data_points) >= 700:
                    reasons.append('kb_dense')

                # 信号2: 上一版该模块条目数 >= 25
                if prev_version_rows and len(prev_version_rows) >= 25:
                    reasons.append('prev_large')

                # 信号3: 模块名本身暗示复杂性
                COMPLEX_MODULES = {"AI语音助手", "主动安全预警提示", "导航", "组队", "设备状态"}
                if module_name in COMPLEX_MODULES:
                    reasons.append('known_complex')

                # 需要至少 2 个信号才触发拆解
                if len(reasons) < 2:
                    return [module_name]

                # === 自动拆解逻辑 ===
                # 按通用维度拆 3 份
                default_splits = [
                    f"{module_name}-核心功能",
                    f"{module_name}-交互与状态",
                    f"{module_name}-异常与边界",
                ]
                print(f"  [AutoSplit] {module_name} 拆解为 {len(default_splits)} 个子模块: {reasons}")
                return default_splits

            def _merge_sub_modules(parent_name: str, sub_results: list) -> list:
                """
                将多个子模块的生成结果合并回父模块，去重。
                """
                merged = []
                seen_names = set()

                for sub_rows in sub_results:
                    if not sub_rows:
                        continue
                    for row in sub_rows:
                        # 统一挂回父模块
                        if row.get('level') == 'L1':
                            row['name'] = parent_name
                            row['module'] = parent_name
                        else:
                            row['module'] = parent_name

                        # 按 name+level 去重
                        name = row.get('name', '')
                        level = row.get('level', '')
                        key = (name, level)
                        if key not in seen_names:
                            seen_names.add(key)
                            merged.append(row)

                print(f"  [Merge] {parent_name}: {sum(len(r) for r in sub_results if r)} 条合并去重为 {len(merged)} 条")
                return merged
            # ===== End Bug B =====

            # ===== Fix 1 Step 3: KB 子模块差异化检索 =====
            def _build_kb_query(module_name: str) -> str:
                """为拆解子模块构建差异化的 KB 检索关键词"""
                if '-' not in module_name:
                    return module_name

                parent, suffix = module_name.rsplit('-', 1)

                # 子维度到检索关键词的映射
                suffix_keywords = {
                    '核心功能': f'{parent} 核心 功能 规格 参数 指标',
                    '核心': f'{parent} 核心 功能 规格 参数 指标',
                    '交互与状态': f'{parent} 交互 HUD 语音 按键 状态 反馈 显示',
                    '交互': f'{parent} 交互 HUD 语音 按键 状态 反馈 显示',
                    '异常与边界': f'{parent} 异常 断连 降级 故障 恢复 兜底 边界 低电量',
                    '异常': f'{parent} 异常 断连 降级 故障 恢复 兜底 边界 低电量',
                    '边界': f'{parent} 异常 断连 降级 故障 恢复 兜底 边界 低电量',
                }

                return suffix_keywords.get(suffix, module_name)
            # ===== End Fix 1 Step 3 =====

            # 单模块生成函数
            def _gen_one(feature: Dict) -> List[Dict]:
                """生成单个 L1 模块的功能清单"""
                name = feature['name']
                module = feature['module']

                # === 知识库注入：提取数据点模式 ===
                # Fix 1 Step 3: 对拆解子模块使用差异化检索
                kb_base_query = _build_kb_query(name)
                kb_queries = [
                    f"{kb_base_query} 技术参数 方案",
                    f"{kb_base_query} 竞品",
                    f"{kb_base_query} 认证 标准",
                ]

                CORE_MODULES = {"导航", "主动安全预警提示", "AI语音助手", "Ai语音助手",
                                "组队", "摄像状态", "胎温胎压", "设备状态"}

                if name in CORE_MODULES:
                    kb_queries.extend([f"{name} 用户需求", f"{name} 功耗 续航"])

                kb_raw = ""
                try:
                    from src.tools.knowledge_base import search_knowledge, format_knowledge_for_prompt
                    for q in kb_queries:
                        try:
                            entries = search_knowledge(q, limit=3)
                            if entries:
                                kb_raw += format_knowledge_for_prompt(entries)[:2000] + "\n"
                        except Exception:
                            pass
                except ImportError:
                    pass

                # 从全文提取数据点（~200字），不灌全文（~3000字）
                kb_data_points = _extract_data_points(kb_raw)

                kb_inject = ""
                if kb_data_points:
                    kb_inject = (
                        f"\n\n## 项目知识库数据（验收标准必须参考这些真实数据）\n"
                        f"{kb_data_points}\n"
                        f"基于以上数据填写验收标准。知识库无相关数据的标注[待验证]。\n"
                    )

                print(f"  [KB] {name}: 知识库原文 {len(kb_raw)} 字 -> 数据点 {len(kb_data_points)} 字")

                # "我的"Tab 额外指示
                extra_hint = ""
                if "我的" in name:
                    extra_hint = (
                        "\n该模块除了账号与设置外,必须包含以下 L2:\n"
                        "- 帮助与反馈(FAQ、在线客服、反馈提交)\n"
                        "- 关于设备(SN、固件版本、保修状态、使用时长统计)\n"
                        "- 隐私与协议(隐私政策、用户协议、数据授权管理)\n"
                        "- 数据管理(骑行数据导出、视频批量导出、账号注销)\n"
                    )

                # ===== A4: 注入 anchor 的 sub_features =====
                anchor_hint = ""
                if anchor:
                    anchor_hint = _get_anchor_sub_features(anchor, name)
                    # 如果是拆解子模块，用父模块名查
                    if not anchor_hint and '-' in name:
                        parent_name = name.rsplit('-', 1)[0]
                        anchor_hint = _get_anchor_sub_features(anchor, parent_name)

                    if anchor_hint:
                        anchor_hint = f"""

【产品锚点要求 — 此模块必须覆盖以下功能点，不可遗漏】
{anchor_hint}
"""
                # ===== End A4 =====

                is_core = name in ["导航", "主动安全预警提示", "AI语音助手", "组队"]

                batch_user_prompt = (
                    f"为智能摩托车全盔项目生成'{name}'模块的功能清单.\n"
                    f"模块归属:{module}\n"
                    f"{extra_hint}"
                    f"{kb_inject}\n"
                    f"{anchor_hint}\n"
                )
                if kb_context:
                    batch_user_prompt += f"内部产品文档参考:\n{kb_context[:2000]}\n\n"
                if goal_text:
                    batch_user_prompt += f"产品目标:\n{goal_text[:300]}\n\n"

                # 添加优化规则
                batch_user_prompt += (
                    f"规则:\n"
                    f"- 模块名必须和'{name}'完全一致,不允许大小写变体或缩写\n"
                    f"- 生成功能时考虑关联页面/场景,在note列标注(如'骑行主界面'、'App-设备Tab')\n"
                    f"- 验收标准必须基于知识库中的真实参数,没有数据的标注[待验证]\n"
                )
                if is_core:
                    batch_user_prompt += f"- 这是核心卖点模块,至少展开5个L2,每个L2至少3个L3\n"

                # 修复 5: 根据 depth 调整输出约束
                module_depth = 'normal'
                parent_name = name.rsplit('-', 1)[0] if '-' in name else name
                if anchor:
                    for section in ['hud_modules', 'app_modules', 'cross_cutting']:
                        for mod in anchor.get(section, []):
                            if mod.get('name') == name or mod.get('name') == parent_name:
                                module_depth = mod.get('depth', 'normal')
                                break

                if module_depth == 'deep':
                    output_limit = """
【输出要求 — 深度模块】
- 本模块最多输出 30 条功能（含 L1/L2/L3 所有层级）
- 每条验收标准必须包含至少 1 个具体数字指标（时间ms/s、百分比%、距离m/km、次数、大小MB等）
- 禁止使用"应支持"、"需提供"等模糊表述，必须量化
- 如果无法确定具体数值，标注为 [需实测:预估值XXX] 而非 [待验证]
"""
                else:
                    output_limit = """
【输出数量硬约束】
- 本模块最多输出 18 条功能（含 L1/L2/L3 所有层级合计）
- 如果功能点超过 18 条，优先保留 P0 和 P1，P2/P3 可精简合并
- 每条功能的 description 字段限 80 字以内
- 每条功能的 acceptance 字段限 120 字以内
- 严格遵守此上限，超出部分不会被系统采纳
"""
                batch_user_prompt += output_limit
                # ===== End 修复 5 =====

                # ===== A5: HUD 端模块的分离规则和新增字段 =====
                HUD_MODULE_SET = {
                    "导航", "来电", "音乐", "消息", "AI语音助手", "Ai语音助手",
                    "简易", "简易路线", "路线", "主动安全预警提示", "组队", "摄像状态",
                    "胎温胎压", "开机动画", "速度", "设备状态", "显示队友位置", "头盔HUD",
                    "实体按键交互", "氛围灯交互", "AI功能", "语音交互", "视觉交互", "多模态交互",
                    "信息中岛", "自定义HUD显示", "SOS与紧急救援", "生命体征与疲劳监测", "场景模式",
                }
                is_hud_module = name in HUD_MODULE_SET or any(h in name for h in HUD_MODULE_SET)

                if is_hud_module:
                    batch_user_prompt += """

【输出规则 — HUD端模块严格遵守】
1. 纯App操作（历史数据查看、复杂设置、内容管理）不放此表，仅放App表
2. 按键操作细节不在此表展开，用"支持按键操作"概述，具体放按键映射表
3. 灯光效果不在此表展开，用"灯光联动提示"概述，具体放灯效定义表
4. 语音指令不在此表展开，用"支持语音控制"概述，具体放语音指令表
5. 本模块最多输出 18 条功能（L1+L2+L3 合计），P0/P1 优先

【额外输出字段 — 每条L3功能必须包含】
- visual_output: 该功能在HUD上的视觉呈现描述（如"左下角箭头卡片+距离数字"）
- display_priority: 显示优先级（critical/high/medium/low），决定屏幕空间争抢时的排序
- degradation: 异常时的降级方案（如"断连时显示最后缓存数据+灰色蒙层"）
- display_duration: 显示时长（permanent/event_Ns/user_dismiss/auto_5s 等）
"""
                # ===== End A5 =====

                # P2: 注入 anchor 的 separation_rules
                if anchor and anchor.get('separation_rules'):
                    rules_text = '\n'.join([f"- {r['rule']}：{r.get('detail','')}" for r in anchor['separation_rules']])
                    batch_user_prompt += f"\n\n【分离规则 — 必须遵守】\n{rules_text}"

                # P2: 注入模块 notes（边界说明）
                if anchor:
                    for section in ['hud_modules', 'app_modules', 'cross_cutting']:
                        for mod in anchor.get(section, []):
                            if mod.get('name') == name and mod.get('notes'):
                                batch_user_prompt += f"\n\n【本模块边界说明】{mod['notes']}"
                                break

                # P4: note 字段标注角色
                batch_user_prompt += """

【角色标注规则 — 每条功能的 note 字段必须遵守】
每条功能的 note 字段必须以 [设计] 或 [研发] 开头：
- [设计] = 用户可见的功能、界面、交互（产品/设计师关注）
- [研发] = 降级策略、异常处理、规则引擎、性能边界（研发工程师关注）
示例:
  "note": "[设计] HUD显示转弯箭头和距离"
  "note": "[研发] 信号丢失时切换离线缓存路线"
"""

                batch_user_prompt += "\n只输出 JSON 数组."

                try:
                    result = gw.call_azure_openai(
                        "cpo", batch_user_prompt, batch_system_prompt,
                        "structured_doc", max_tokens=4096
                    )

                    if not result.get("success"):
                        error_msg = result.get("error", "未知错误")
                        print(f"  [GenOne] {name} LLM 返回失败: {error_msg}")
                        print(f"  [GenOne] {name} prompt 长度: {len(batch_user_prompt)} 字")
                        return []

                    # Fix 3: 截断检测
                    finish_reason = result.get("finish_reason", "")
                    if finish_reason == "length":
                        usage = result.get("usage", {})
                        print(f"  [TruncGuard] ⚠️ {name} 输出被截断 (tokens={usage.get('completion_tokens', '?')})")
                        # Round 8 Fix 6: 先尝试精简 prompt 重试，失败才走 AutoSplit
                        print(f"  [TruncGuard] {name}: 尝试精简模式重试...")
                        minimal_prompt = (
                            f"生成「{name}」功能清单。模块：{name}。\n"
                            f"输出 JSON 数组：module/level(L1,L2,L3)/parent/name/priority/interaction/description/acceptance/dependencies/note。\n"
                            f"至少 2 个 L2，每个 L2 至少 2 个 L3。只输出 JSON。"
                        )
                        try:
                            minimal_result = gw.call_azure_openai(
                                "cpo", minimal_prompt, "只输出JSON。",
                                "structured_doc_minimal", max_tokens=4096
                            )
                            if minimal_result.get("success"):
                                minimal_resp = minimal_result.get("response", "")
                                json_match = re.search(r'\[[\s\S]*\]', minimal_resp)
                                if json_match:
                                    try:
                                        cleaned_json = _clean_json_text(json_match.group())
                                        items = json.loads(cleaned_json)
                                        if items and len(items) > 1:
                                            for item in items:
                                                item["note"] = (item.get("note", "") + " [精简重试]").strip()
                                            print(f"  [TruncGuard] {name}: 精简重试成功 → {len(items)} 条")
                                            return items
                                    except Exception as e:
                                        print(f"  [TruncGuard] {name}: 精简重试 JSON 解析失败: {e}")
                        except Exception as e:
                            print(f"  [TruncGuard] {name}: 精简重试失败: {e}")
                        # 精简重试失败，返回空触发 AutoSplit
                        print(f"  [TruncGuard] {name}: 精简重试失败，触发 AutoSplit")
                        return []

                    response = result.get("response", "")
                    if len(response) < 50:
                        print(f"  [GenOne] {name} 响应太短: {len(response)} 字")
                        return []

                    json_match = re.search(r'\[[\s\S]*\]', response)
                    if not json_match:
                        print(f"  [GenOne] {name} 无法提取 JSON，响应前200字: {response[:200]}")
                        return []

                    try:
                        # Bug C: 清洗控制字符
                        cleaned_json = _clean_json_text(json_match.group())
                        items = json.loads(cleaned_json)
                        # 防御：确保所有字段都是 string
                        for item in items:
                            for key in item:
                                if isinstance(item[key], (list, dict)):
                                    item[key] = ", ".join(str(v) for v in item[key]) if isinstance(item[key], list) else str(item[key])
                                elif item[key] is None:
                                    item[key] = ""
                        return items
                    except json.JSONDecodeError as je:
                        print(f"  [GenOne] {name} JSON 解析失败: {je}")
                        # 二次尝试: strict=False 模式
                        try:
                            items = json.loads(cleaned_json, strict=False)
                            for item in items:
                                for key in item:
                                    if isinstance(item[key], (list, dict)):
                                        item[key] = ", ".join(str(v) for v in item[key]) if isinstance(item[key], list) else str(item[key])
                                    elif item[key] is None:
                                        item[key] = ""
                            print(f"  [GenOne] {name} JSON 解析成功(strict=False)")
                            return items
                        except json.JSONDecodeError as e2:
                            print(f"  [GenOne] {name} JSON 二次解析失败: {e2}")
                            print(f"  [GenOne] {name} JSON 片段: {cleaned_json[:300]}")
                            return []

                except Exception as e:
                    print(f"  [GenOne] {name} 异常: {type(e).__name__}: {e}")
                    print(f"  [GenOne] {name} prompt 长度: {len(batch_user_prompt)} 字")
                    import traceback
                    traceback.print_exc()
                    return []

            # 并行生成
            done_count = 0

            # ===== Speed 2: 提前启动独立 Sheet（不依赖功能表）=====
            independent_futures = {}
            independent_pool = ThreadPoolExecutor(max_workers=6)
            independent_futures['light'] = independent_pool.submit(_gen_sheet6_light_effects, gw)
            independent_futures['button'] = independent_pool.submit(_gen_sheet5_button_mapping, gw)
            independent_futures['voice'] = independent_pool.submit(_gen_sheet4_voice_commands, gw, kb_context)
            independent_futures['voice_nav'] = independent_pool.submit(_gen_sheet7_voice_nav_scenarios, gw)
            if anchor and anchor.get('flow_diagrams'):
                independent_futures['flow'] = independent_pool.submit(_generate_flow_diagrams, anchor, gw)
            if anchor and anchor.get('user_journeys'):
                independent_futures['journey'] = independent_pool.submit(_generate_user_journeys, anchor)
            print(f"[FastTrack] 已提前启动 {len(independent_futures)} 个独立 Sheet")

            # Bug B: 复杂模块自动拆解
            def _gen_one_with_split(feature: Dict) -> tuple:
                """生成单个模块，复杂模块自动拆解"""
                name = feature['name']
                prev_rows = prev_by_module.get(name, [])

                # 计算知识库数据点（用于复杂度判断）
                kb_dp = ""
                try:
                    from src.tools.knowledge_base import search_knowledge
                    entries = search_knowledge(name, limit=5)
                    if entries:
                        from src.tools.knowledge_base import format_knowledge_for_prompt
                        kb_raw = format_knowledge_for_prompt(entries)[:3000]
                        kb_dp = _extract_data_points(kb_raw)
                except Exception:
                    pass

                # 判断是否需要拆解
                sub_modules = _estimate_complexity(name, kb_dp, prev_rows)

                if len(sub_modules) == 1:
                    # 不需要拆解，正常生成
                    result = _gen_one(feature)
                    return name, result, False
                else:
                    # 拆解模式：为每个子模块创建 feature 并行生成
                    sub_results = []
                    sub_features = [{"name": sub_name, "module": sub_name} for sub_name in sub_modules]

                    with ThreadPoolExecutor(max_workers=3) as sub_pool:
                        sub_futures = {sub_pool.submit(_gen_one, sf): sf for sf in sub_features}
                        for sf in as_completed(sub_futures):
                            sf_info = sub_futures[sf]
                            try:
                                sub_result = sf.result()
                                if sub_result:
                                    sub_results.append(sub_result)
                            except Exception as e:
                                print(f"  [AutoSplit] 子模块 {sf_info['name']} 失败: {e}")

                    if sub_results:
                        merged = _merge_sub_modules(name, sub_results)
                        return name, merged, True
                    else:
                        # 拆解失败，退回普通生成
                        result = _gen_one(feature)
                        return name, result, False

            with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as pool:
                futures = {pool.submit(_gen_one_with_split, f): f for f in l1_features}
                for future in as_completed(futures):
                    feature = futures[future]
                    try:
                        name, batch, was_split = future.result()
                        done_count += 1
                        if batch:
                            all_items.extend(batch)
                            split_tag = " [拆解]" if was_split else ""
                            print(f"  ✅ [{done_count}/{total}] {name}: +{len(batch)} 条{split_tag}")
                        else:
                            print(f"  ❌ [{done_count}/{total}] {feature['name']}: 失败")
                    except Exception as e:
                        done_count += 1
                        print(f"  ❌ [{done_count}/{total}] {feature['name']}: {e}")

            print(f"[FastTrack] 完成: {len(all_items)} 条功能 ({done_count} 个模块)")

            # 检测哪些 L1 没有生成成功（同时检查 name 和 module 字段）
            generated_names = set()
            for item in all_items:
                if item.get("level") == "L1":
                    generated_names.add(item.get("name", ""))
                    generated_names.add(item.get("module", ""))

            failed_features = [
                f for f in l1_features
                if f["name"] not in generated_names
                and f.get("module", "") not in generated_names
            ]

            print(f"[Retry] 已生成标识: {generated_names}")
            if failed_features:
                print(f"[Retry] 真正失败: {[f['name'] for f in failed_features]}")

            if failed_features:
                print(f"[Retry] {len(failed_features)} 个模块失败，并行重试...")
                import time as _time

                def _retry_one(feature):
                    name = feature["name"]

                    # ===== Round 5 Fix 3: 失败先拆分再精简 =====
                    # 策略 0：先尝试 AutoSplit（保留完整上下文，比精简质量高）
                    print(f"  [Retry] {name}: 拆分重试...")
                    sub_modules = [
                        f"{name}-核心功能",
                        f"{name}-交互与状态",
                        f"{name}-异常与边界",
                    ]
                    sub_results = []
                    split_success = True
                    for sub_name in sub_modules:
                        sub_feature = {"name": sub_name, "module": sub_name}
                        try:
                            # 使用 _gen_one（完整 prompt），而非精简
                            sub_batch = _gen_one(sub_feature)
                            if sub_batch and len(sub_batch) > 1:
                                sub_results.append(sub_batch)
                                print(f"    [Split] {sub_name}: +{len(sub_batch)} 条")
                            else:
                                print(f"    [Split] {sub_name}: 空结果")
                                split_success = False
                        except Exception as e:
                            print(f"    [Split] {sub_name} 失败: {e}")
                            split_success = False

                    if sub_results:
                        merged = _merge_sub_modules(name, sub_results)
                        for item in merged:
                            item["note"] = (item.get("note", "") + " [拆解重试]").strip()
                        print(f"  [Retry] {name}: 拆分成功 → {len(merged)} 条")
                        return name, merged
                    else:
                        print(f"  [Retry] {name}: 拆分失败，降级精简模式...")
                    # ===== End Round 5 Fix 3 =====

                    # 策略 1：拆分失败后，用精简 prompt（完整 prompt 已经失败过了）
                    print(f"  [Retry] {name}: 精简模式...")
                    try:
                        batch = _gen_one_minimal(feature, gw)
                        if batch and len(batch) > 1:
                            for item in batch:
                                item["note"] = (item.get("note", "") + " [精简生成]").strip()
                            return name, batch
                    except Exception as e:
                        print(f"  [Retry] {name} 精简失败: {e}")

                    # 策略 2：等 10 秒后完整 prompt 重试一次
                    _time.sleep(10)
                    print(f"  [Retry] {name}: 完整模式重试...")
                    try:
                        batch = _gen_one(feature)
                        if batch and len(batch) > 1:
                            return name, batch
                    except Exception as e:
                        print(f"  [Retry] {name} 完整重试失败: {e}")

                    # 策略 3：Gemini 降级
                    print(f"  [Retry] {name}: Gemini 降级...")
                    try:
                        batch = _gen_one_with_gemini(feature, gw)
                        if batch and len(batch) > 1:
                            return name, batch
                    except Exception as e:
                        print(f"  [Retry] {name} Gemini 失败: {e}")

                    return name, None

                with ThreadPoolExecutor(max_workers=8) as pool:
                    futs = {pool.submit(_retry_one, f): f for f in failed_features}
                    for future in as_completed(futs):
                        feature = futs[future]
                        name, batch = future.result()
                        if batch:
                            all_items.extend(batch)
                            print(f"  ✅ [Retry] {name}: +{len(batch)} 条")
                        elif name in prev_by_module and len(prev_by_module[name]) > 1:
                            all_items.extend(prev_by_module[name])
                            print(f"  ⏪ [兜底] {name}: 保留上一版 ({len(prev_by_module[name])} 条)")
                        else:
                            all_items.append({
                                "module": feature["module"] if feature["name"] == name else name,
                                "level": "L1", "parent": "", "name": name,
                                "priority": "P0", "interaction": "",
                                "description": "[待生成] 多次尝试失败",
                                "acceptance": "[待生成]", "dependencies": "", "note": ""
                            })
                            print(f"  ❌ [放弃] {name}: 标记为[待生成]")

            # 去重:同名+同层级只保留第一条(不管模块名)
            seen_names = set()
            unique_items = []
            for item in all_items:
                key = (item.get("name", "").strip(), item.get("level", ""))
                if key not in seen_names:
                    seen_names.add(key)
                    unique_items.append(item)
                else:
                    print(f"  [去重] {item.get('name')}")

            if len(unique_items) < len(all_items):
                print(f"[FastTrack] 去重: {len(all_items)} -> {len(unique_items)}")
            all_items = unique_items

            # === 空壳 L1 检测与补救 ===
            l1_child_count = {}  # {l1_name: child_count}
            current_l1 = ""
            for item in all_items:
                if item.get("level") == "L1":
                    current_l1 = item.get("name", "")
                    if current_l1 not in l1_child_count:
                        l1_child_count[current_l1] = 0
                elif item.get("level") in ("L2", "L3") and current_l1:
                    l1_child_count[current_l1] = l1_child_count.get(current_l1, 0) + 1

            empty_l1s = [name for name, count in l1_child_count.items() if count == 0]

            if empty_l1s:
                print(f"\n[QA] 发现 {len(empty_l1s)} 个空壳 L1: {empty_l1s}")

                for l1_name in empty_l1s:
                    feature = next((f for f in l1_features if f["name"] == l1_name), None)
                    if not feature:
                        continue

                    # 先检查上一版有没有
                    if l1_name in prev_by_module and len(prev_by_module[l1_name]) > 1:
                        # 上一版有内容，直接用上一版
                        all_items = [i for i in all_items if not (i.get("level") == "L1" and i.get("name") == l1_name)]
                        all_items.extend(prev_by_module[l1_name])
                        print(f"  [空壳补救] {l1_name}: 用上一版 ({len(prev_by_module[l1_name])} 条)")
                    else:
                        # 上一版也没有，用精简模式生成
                        print(f"  [空壳补救] {l1_name}: 精简模式生成...")
                        batch = _gen_one_minimal(feature, gw)
                        if batch and len(batch) > 1:
                            all_items = [i for i in all_items if not (i.get("level") == "L1" and i.get("name") == l1_name)]
                            all_items.extend(batch)
                            print(f"  OK [空壳补救] {l1_name}: +{len(batch)} 条")
                        else:
                            print(f"  X [空壳补救] {l1_name}: 仍然失败")

            # === 逐模块对比取优 ===
            if prev_by_module:
                print(f"\n[Compare] 逐模块对比取优...")

                # ===== Fix 3 Step 2: 按归一化名称分组 =====
                from collections import defaultdict
                normalized_groups = defaultdict(list)
                for item in all_items:
                    module = item.get('module') or item.get('name', '')
                    norm_name = _normalize_module_name(module)
                    item['module'] = norm_name  # 统一模块名
                    normalized_groups[norm_name].append(item)

                # 同名模块合并去重
                all_items = []
                for norm_name, group_rows in normalized_groups.items():
                    if len(group_rows) <= 1:
                        all_items.extend(group_rows)
                        continue

                    # 按 name(L3功能名) 去重，保留验收标准更完整的那条
                    seen = {}
                    for row in group_rows:
                        key = row.get('name', '')
                        if key in seen:
                            # 比较质量，保留更好的
                            old_acc = str(seen[key].get('acceptance', ''))
                            new_acc = str(row.get('acceptance', ''))
                            if len(new_acc) > len(old_acc):
                                seen[key] = row
                        else:
                            seen[key] = row

                    deduped = list(seen.values())
                    print(f"  [Normalize] {norm_name}: {len(group_rows)} 条合并去重为 {len(deduped)} 条")
                    all_items.extend(deduped)
                # ===== End Fix 3 Step 2 =====

                # 按 L1 分组新版数据
                new_by_module = {}
                current_l1 = ""
                for item in all_items:
                    if item.get("level") == "L1":
                        current_l1 = item.get("name", "")
                    if current_l1:
                        if current_l1 not in new_by_module:
                            new_by_module[current_l1] = []
                        new_by_module[current_l1].append(item)

                # 逐模块对比
                final_items = []
                all_module_names = set(list(prev_by_module.keys()) + list(new_by_module.keys()))

                for mod_name in all_module_names:
                    old = prev_by_module.get(mod_name, [])
                    new = new_by_module.get(mod_name, [])

                    try:
                        old_score = _score_module(old)
                    except Exception as e:
                        print(f"  [Score] 旧版评分异常: {e}, 默认0分")
                        old_score = 0

                    try:
                        new_score = _score_module(new)
                    except Exception as e:
                        print(f"  [Score] 新版评分异常: {e}, 默认0分")
                        new_score = 0

                    # Layer 2: 使用替换模式对比——择优替换，不做并集
                    # decision = 'replace' / 'keep'（不再有 merge 或吸收逻辑）
                    decision = _compare_module_replace_mode(mod_name, old, new, old_score, new_score)

                    if decision == 'replace':
                        # 整体替换为新版，不保留旧版独有功能
                        final_items.extend(new)
                        print(f"  REPLACE {mod_name}: 新版质量更好 → 整体替换 (新{len(new)}条，弃旧{len(old)}条)")
                    elif decision == 'keep':
                        # 整体保留旧版，不吸收新版独有功能
                        final_items.extend(old)
                        print(f"  KEEP {mod_name}: 旧版质量更好 → 保留旧版 ({len(old)}条)")
                    # decision is None → 跳过该模块

                all_items = final_items

                # 修复 6: Compare 后再归一化一次（处理 KEEP 带入的旧名称）
                if anchor:
                    normalize_map = anchor.get('module_normalize', {})
                else:
                    normalize_map = {}
                normalize_map.update(MODULE_NAME_NORMALIZE)
                all_items = _normalize_all_rows(all_items, normalize_map, anchor)
                print(f"[Normalize-Post] Compare 后归一化: {len(all_items)} 条")

                print(f"[Compare] 最终: {len(all_items)} 条")

                # === Layer 3: 密度控制 ===
                print(f"\n[DensityCap] 检查模块密度上限...")
                all_items = _apply_density_limits(all_items, anchor, gw)
                print(f"[DensityCap] 最终: {len(all_items)} 条")

                # === Layer 4: 模块级 QA ===
                print(f"\n[QA] 模块级质量检查...")
                all_items = _apply_module_qa(all_items)
                print(f"[QA] 最终: {len(all_items)} 条")

            if not all_items:
                _send_reply("生成失败:所有批次均未生成有效内容")
                return

            # 全局优先级校准
            _calibrate_priorities(all_items)

            items = all_items

            # ========== 如果是完整规格书,并行生成 Sheet 3-6 ==========
            extra_sheets = None
            early_mindmap_result = {"mm": None, "svg": None}  # 预初始化供非完整模式使用

            if is_full_spec:
                print("[FastTrack] 完整规格书模式:收集提前启动的 Sheet + 生成依赖型 Sheet")
                _send_reply("📋 功能清单已完成,正在生成状态对策/语音/按键/灯效/导航场景/用户故事/测试用例/页面映射/开发任务...")

                extra_sheets = {}

                # 收集提前启动的独立 Sheet 结果
                for name, future in independent_futures.items():
                    try:
                        data = future.result(timeout=300)
                        extra_sheets[name] = data
                        print(f"  ✅ {name} (提前): {len(data) if data else 0} 条")
                    except Exception as e:
                        print(f"  ❌ {name} (提前): {e}")
                        extra_sheets[name] = []

                # 关闭独立 Sheet 线程池
                independent_pool.shutdown(wait=False)

                # 脑图早期生成线程(与第一批并行)
                # early_mindmap_result 已在外层初始化

                def _early_gen_mindmap():
                    try:
                        early_mindmap_result["mm"] = _generate_mindmap_mm(items, "智能骑行头盔 V1 功能框架")
                        early_mindmap_result["svg"] = _generate_mindmap_svg(items, "智能骑行头盔 V1 功能框架")
                        print("[FastTrack] 脑图早期生成完成")
                    except Exception as e:
                        print(f"[FastTrack] 脑图早期生成失败: {e}")

                mindmap_thread = threading.Thread(target=_early_gen_mindmap, daemon=True)
                mindmap_thread.start()

                # 只启动依赖功能表的 Sheet: state, user_stories, test_cases, page_mapping, dev_tasks
                with ThreadPoolExecutor(max_workers=5) as pool:
                    futures = {
                        pool.submit(_gen_sheet3_state_scenarios, gw, kb_context, goal_text): "state",
                        pool.submit(_gen_user_stories, items, gw): "user_stories",
                        pool.submit(_gen_test_cases, items, gw): "test_cases",
                        pool.submit(_gen_page_mapping, items, gw): "page_mapping",
                        pool.submit(_gen_dev_tasks, items, gw): "dev_tasks",
                    }
                    for future in as_completed(futures):
                        name = futures[future]
                        data = future.result()
                        extra_sheets[name] = data
                        print(f"  ✅ {name}: {len(data)} 条")

                # 等待脑图早期生成完成
                mindmap_thread.join(timeout=30)

                total_extra = sum(len(v) for v in extra_sheets.values())
                print(f"[FastTrack] 额外 Sheet 完成: {total_extra} 条")

                # Round 5 Fix 5: 跨 Sheet 交叉验证
                cross_issues = _cross_validate(items, extra_sheets)

            # 跨模块一致性审计
            audit_issues = _cross_module_audit(items, extra_sheets if extra_sheets else {})

            # B4: 一致性审计闭环 — 自动修复
            if audit_issues:
                print(f"[Audit] 发现 {len(audit_issues)} 个一致性问题")
                try:
                    audit_issues = _auto_fix_audit_issues(audit_issues, extra_sheets if extra_sheets else {}, gw)
                except Exception as e:
                    print(f"[Audit-Fix] 自动修复异常: {e}，跳过")

            # 并行生成文件
            try:
                task_id = f"{hash(text) % 100000:05d}"

                # === 并行生成 4 个文件（Excel、xmind、mindmap HTML、PRD HTML）===
                file_results = {}

                # 获取早期生成的脑图结果（如果有）
                early_mm = early_mindmap_result.get("mm")
                early_svg = early_mindmap_result.get("svg")

                def _make_excel():
                    flow_data = extra_sheets.get('flow', []) if extra_sheets else []
                    return _export_to_excel(items, f"prd_{task_id}", text[:50],
                                           extra_sheets=extra_sheets, audit_issues=audit_issues,
                                           flow_diagrams=flow_data, anchor=anchor)

                def _make_xmind():
                    data = _generate_mindmap_xmind(items, "智能骑行头盔 V1 功能框架")
                    path = EXPORT_DIR / f"mindmap_{task_id}.xmind"
                    path.write_bytes(data)
                    return str(path)

                def _make_mindmap_html():
                    if early_svg:
                        content = early_svg
                    else:
                        content = _generate_mindmap_svg(items, "智能骑行头盔 V1 功能框架")
                    path = EXPORT_DIR / f"mindmap_{task_id}.html"
                    path.write_text(content, encoding="utf-8")
                    return str(path)

                def _make_prd_html():
                    content = _generate_interactive_prd_html(items, extra_sheets, "智能骑行头盔 V1 PRD 规格书", anchor)
                    path = EXPORT_DIR / f"prd_interactive_{task_id}.html"
                    path.write_text(content, encoding="utf-8")
                    return str(path)

                with ThreadPoolExecutor(max_workers=4) as gen_pool:
                    f_excel = gen_pool.submit(_make_excel)
                    f_xmind = gen_pool.submit(_make_xmind)
                    f_mindmap = gen_pool.submit(_make_mindmap_html)
                    f_prd = gen_pool.submit(_make_prd_html)

                    try:
                        file_results["excel"] = f_excel.result()
                        print(f"  ✅ Excel: {file_results['excel']}")
                    except Exception as e:
                        print(f"[File] Excel 生成失败: {e}")
                    try:
                        file_results["xmind"] = f_xmind.result()
                        print(f"  ✅ XMind: {file_results['xmind']}")
                    except Exception as e:
                        print(f"[File] XMind 生成失败: {e}")
                    try:
                        file_results["mindmap_html"] = f_mindmap.result()
                        print(f"  ✅ Mindmap HTML: {file_results['mindmap_html']}")
                    except Exception as e:
                        print(f"[File] 脑图HTML 生成失败: {e}")
                    try:
                        file_results["prd_html"] = f_prd.result()
                        print(f"  ✅ PRD HTML: {file_results['prd_html']}")
                    except Exception as e:
                        print(f"[File] PRD HTML 生成失败: {e}")

                print(f"[FastTrack] {len(file_results)} 个文件并行生成完成")

                # Bug 4: 停止生成 .mm 文件（已注释）
                # if early_mm:
                #     mm_path = EXPORT_DIR / f"mindmap_{task_id}.mm"
                #     mm_path.write_text(early_mm, encoding="utf-8")
                #     file_results["mm"] = str(mm_path)
                #     print(f"  ✅ .mm 文件: {mm_path}")
                # else:
                #     try:
                #         mm_content = _generate_mindmap_mm(items, "智能骑行头盔 V1 功能框架")
                #         mm_path = EXPORT_DIR / f"mindmap_{task_id}.mm"
                #         mm_path.write_text(mm_content, encoding="utf-8")
                #         file_results["mm"] = str(mm_path)
                #         print(f"  ✅ .mm 文件: {mm_path}")
                #     except Exception as e:
                #         print(f"[File] .mm 生成失败: {e}")

                # 保存版本快照
                try:
                    versions_dir = EXPORT_DIR.parent / "prd_versions"
                    versions_dir.mkdir(parents=True, exist_ok=True)

                    version_id = datetime.now().strftime("%Y%m%d_%H%M%S")
                    snapshot = {
                        "version": version_id,
                        "timestamp": datetime.now().isoformat(),
                        "total_features": len(items),
                        "l1_count": sum(1 for i in items if i.get("level") == "L1"),
                        "l2_count": sum(1 for i in items if i.get("level") == "L2"),
                        "l3_count": sum(1 for i in items if i.get("level") == "L3"),
                        "priority_dist": {p: sum(1 for i in items if i.get("priority") == p) for p in ["P0", "P1", "P2", "P3"]},
                        "items": items,
                        "extra_sheets_counts": {k: len(v) for k, v in (extra_sheets or {}).items()},
                        "file_paths": file_results,
                    }

                    snap_path = versions_dir / f"prd_v_{version_id}.json"
                    snap_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
                    print(f"[Version] 版本快照已保存: {version_id}")
                except Exception as e:
                    print(f"[Version] 保存失败: {e}")

                # === 并行发送所有文件到飞书 ===
                send_files = []
                if file_results.get("excel"):
                    send_files.append(("Excel", file_results["excel"]))
                if file_results.get("xmind"):
                    send_files.append(("XMind脑图", file_results["xmind"]))
                if file_results.get("mindmap_html"):
                    send_files.append(("交互脑图", file_results["mindmap_html"]))
                if file_results.get("prd_html"):
                    send_files.append(("交互PRD", file_results["prd_html"]))
                # Bug 4: 停止发送 .mm 文件（已注释）
                # if file_results.get("mm"):
                #     send_files.append(("FreeMind脑图", file_results["mm"]))

                def _send_one(name_path):
                    name, path = name_path
                    try:
                        ok = _send_file_to_feishu(reply_target, path, reply_type)
                        return name, ok
                    except Exception as e:
                        print(f"[Send] {name} 失败: {e}")
                        return name, False

                sent_results = {}
                with ThreadPoolExecutor(max_workers=5) as send_pool:
                    futs = {send_pool.submit(_send_one, sf): sf[0] for sf in send_files}
                    for f in as_completed(futs):
                        name, ok = f.result()
                        sent_results[name] = ok
                        print(f"  {'✅' if ok else '❌'} {name} 发送{'成功' if ok else '失败'}")

                file_sent = sent_results.get("Excel", False)

                if file_sent:
                    summary = _format_items_as_tree(items)
                    _send_reply(
                        f"📊 功能PRD清单已生成({len(items)} 条),Excel + 脑图已发送.\n\n预览:\n{summary[:3000]}"
                    )
                else:
                    # 文件发送失败,发树形摘要
                    summary = _format_items_as_tree(items)
                    _send_reply(f"📋 功能PRD清单({len(items)} 条):\n\n{summary[:4000]}")
                    excel_path = file_results.get("excel", "未生成")
                    _send_reply(f"⚠️ 文件发送失败,Excel 已保存服务器: {excel_path}")

            except Exception as e:
                print(f"[FastTrack] Excel 失败: {e}")
                import traceback
                traceback.print_exc()
                # 降级发树形摘要
                summary = _format_items_as_tree(items)
                _send_reply(f"📋 功能清单({len(items)} 条):\n\n{summary[:4000]}")

        except Exception as e:
            print(f"[FastTrack] 异常: {e}")
            import traceback
            traceback.print_exc()
            _send_reply(f"生成失败: {str(e)[:200]}")

    # 后台线程执行
    threading.Thread(target=_process_in_background, daemon=True).start()
    return True


# === 测试入口 ===
if __name__ == "__main__":
    import os
    os.chdir(Path(__file__).parent.parent.parent)

    # 测试导入
    from scripts.feishu_handlers.structured_doc import try_structured_doc_fast_track
    print("模块导入成功")

    # 测试关键词检测
    test_texts = [
        "帮我生成智能头盔的功能清单",
        "这是一个短文本",
        "请输出PRD表格,包含HUD显示、语音交互、按键控制等功能模块",
    ]

    for t in test_texts:
        result = is_structured_doc_request(t)
        print(f"'{t[:30]}...' -> {result}")