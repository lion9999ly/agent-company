"""
@description: 研发任务处理 - LangGraph 多 Agent 调用 + 后台线程
@refactored_from: feishu_sdk_client.py
@last_modified: 2026-03-28
"""
import threading
import json
import uuid
import re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# 研发任务并发锁，同一时间只允许一个研发任务执行
_rd_lock = threading.Lock()
_rd_task_running = False


def is_rd_task(text: str) -> bool:
    """判断是否为需要多Agent协作的研发任务"""
    keywords = [
        "设计方案", "技术方案", "架构", "需求分析",
        "市场策略", "GTM", "竞品分析", "产品规划",
        "研发", "开发方案", "实现方案", "技术选型",
        "设计", "模块", "方案", "选型", "评估",
        "原型", "验证", "测试方案", "硬件", "固件",
        "通信", "传感器", "算法", "HUD", "蓝牙",
        "导航", "供应链", "认证", "量产"
    ]
    return any(kw in text for kw in keywords)


def is_rd_task_running() -> bool:
    """检查是否有研发任务正在运行"""
    return _rd_task_running


def run_rd_task_background(text: str, reply_target: str, reply_type: str, open_id: str, chat_id: str, send_reply):
    """后台线程执行研发任务，不阻塞主线程"""
    global _rd_task_running

    try:
        _rd_task_running = True
        reply = _call_langgraph(text, reply_target, reply_type)

        if reply:
            # 检测 JSON 并导出 Excel
            json_match = re.search(r'\[[\s\S]*\]', reply)
            xlsx_path = None
            items = None

            if json_match:
                try:
                    items = json.loads(json_match.group())
                    if isinstance(items, list) and len(items) > 0:
                        task_id = f"{uuid.uuid4().hex[:8]}"
                        xlsx_path = _export_to_excel(items, task_id, text)
                        print(f"[Export] Excel 生成: {xlsx_path}, {len(items)} 条功能")
                except Exception as e:
                    print(f"[Export] JSON 解析失败: {e}")
                    items = None
                    xlsx_path = None

            # 根据 Excel 生成结果决定发送内容
            if xlsx_path:
                from scripts.feishu_handlers.file_sender import send_file_to_feishu
                file_sent = send_file_to_feishu(reply_target, xlsx_path, reply_type)
                print(f"[Export] 飞书文件发送: {'成功' if file_sent else '失败'}")

                if file_sent:
                    send_reply(reply_target, f"📊 功能PRD清单已导出为 Excel（{len(items)} 条功能），文件已发送。")
                else:
                    summary = _format_items_as_tree(items)
                    send_reply(reply_target, summary[:4000])
                    send_reply(reply_target, f"⚠️ 飞书文件发送失败，Excel 已保存在服务器: {xlsx_path}")
            else:
                clean_reply = _clean_synthesis_output(reply)
                if clean_reply.startswith('[') and '{' in clean_reply:
                    try:
                        items_retry = json.loads(clean_reply)
                        if isinstance(items_retry, list):
                            clean_reply = _format_items_as_tree(items_retry)
                    except:
                        pass
                send_reply(reply_target, clean_reply[:4000])

        send_reply(reply_target, "[Rating] Please rate this solution:\nA. Ready to use\nB. Needs minor changes\nC. Right direction but not deep enough\nD. Wrong direction\nReply with letter, C/D please include reason.")

    except Exception as e:
        send_reply(reply_target, f"[Error] R&D task failed: {str(e)[:300]}")
        print(f"[RD_TASK] Background execution error: {e}")
    finally:
        _rd_task_running = False


def _call_langgraph(text: str, reply_target: str = None, reply_type: str = None) -> str:
    """调用 LangGraph 多 Agent 工作流"""
    try:
        from src.graph.router import app as langgraph_app
        from src.schema.state import AgentGlobalState

        initial_state = {
            "task_contract": {"task_goal": text, "task_id": f"task_{uuid.uuid4().hex[:8]}"},
            "sub_tasks": {},
            "execution": {},
            "control": {},
            "memory_context": "",
            "metadata": {"task_id": f"task_{uuid.uuid4().hex[:8]}"}
        }

        result = langgraph_app.invoke(initial_state)
        synthesis = result.get("execution", {}).get("synthesis_output", "")

        # 保存记忆
        _save_task_memory(text, synthesis, result)

        return synthesis
    except Exception as e:
        print(f"[LangGraph] Error: {e}")
        return f"执行失败: {str(e)}"


def _save_task_memory(task_goal: str, synthesis: str, result: dict):
    """保存任务记忆"""
    try:
        from datetime import datetime
        memory_dir = PROJECT_ROOT / ".ai-state" / "memory"
        memory_dir.mkdir(parents=True, exist_ok=True)

        task_id = f"task_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
        memory_file = memory_dir / f"{task_id}.json"

        memory_data = {
            "task_id": task_id,
            "task_goal": task_goal,
            "synthesis_output": synthesis[:4000],
            "timestamp": datetime.now().isoformat(),
            "result_summary": {
                "cto": result.get("execution", {}).get("cto_output", {}).get("output", "")[:500],
                "cmo": result.get("execution", {}).get("cmo_output", {}).get("output", "")[:500],
                "critic_decision": result.get("execution", {}).get("critic_decision", ""),
            }
        }

        memory_file.write_text(json.dumps(memory_data, ensure_ascii=False, indent=2), encoding="utf-8")

        # 更新 last_task_memory
        from scripts.feishu_handlers.commands import set_last_task_memory
        set_last_task_memory(task_id, str(memory_dir))

        print(f"[Memory] Task saved: {task_id}")
    except Exception as e:
        print(f"[Memory] Save failed: {e}")


def _clean_synthesis_output(result_text: str) -> str:
    """清理整合输出：提取 CPO 整合部分"""
    if not result_text or len(result_text.strip()) < 50:
        return result_text

    # 尝试提取关键内容
    clean = result_text

    # 去除 Agent 标记
    markers = ["[CTO]", "[CMO]", "[CDO]", "[Critic]", "【CTO】", "【CMO】", "【CDO】", "【Critic】"]
    for m in markers:
        clean = clean.replace(m, "")

    return clean.strip()


def _format_items_as_tree(items: list, max_items: int = 50) -> str:
    """将功能列表格式化为树形文本"""
    if not items:
        return ""

    lines = ["📊 功能清单\n"]
    for i, item in enumerate(items[:max_items], 1):
        name = item.get("name", item.get("功能名", item.get("title", f"项目{i}")))
        priority = item.get("priority", item.get("优先级", ""))
        desc = item.get("description", item.get("描述", ""))[:60]

        line = f"{i}. {name}"
        if priority:
            line += f" [{priority}]"
        lines.append(line)

        if desc:
            lines.append(f"   └─ {desc}")

    if len(items) > max_items:
        lines.append(f"\n... 共 {len(items)} 条，仅显示前 {max_items} 条")

    return "\n".join(lines)


def _export_to_excel(items: list, task_id: str, task_goal: str = "") -> str:
    """导出功能列表为 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "功能PRD"

        # 表头
        headers = ["序号", "功能名", "优先级", "描述", "验收标准"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 数据
        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=item.get("name", item.get("功能名", item.get("title", ""))))
            ws.cell(row=row, column=3, value=item.get("priority", item.get("优先级", "")))
            ws.cell(row=row, column=4, value=item.get("description", item.get("描述", "")))
            ws.cell(row=row, column=5, value=item.get("acceptance", item.get("验收标准", "")))

        # 调整列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30

        # 保存
        export_dir = PROJECT_ROOT / ".ai-state" / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = export_dir / f"prd_{task_id}.xlsx"
        wb.save(xlsx_path)

        return str(xlsx_path)
    except Exception as e:
        print(f"[Excel] Export failed: {e}")
        return ""


# === 导出接口 ===
__all__ = [
    "is_rd_task",
    "is_rd_task_running",
    "run_rd_task_background",
]